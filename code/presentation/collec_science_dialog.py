#!/usr/bin/env python3
"""
Dialogs for Collect-Science generation.
Layout inspire de "Autre type de source" :
  Gauche  : previsualisation du fichier COLISA (Excel ou donnees logiciel) + selection de lignes
  Droite  : panneau "Contenant" (type + numero, bouton + custom, information sur les lignes deja affectees)
"""
from __future__ import annotations

import hashlib
import json
import os
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from infrastructure.app_paths import settings_dir

from PySide6.QtCore import Qt, QAbstractTableModel, QModelIndex, Signal
from PySide6.QtGui import QColor, QFont, QBrush, QPalette
from PySide6.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QCheckBox,
    QComboBox,
    QDialog,
    QFrame,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QInputDialog,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QSizePolicy,
    QSplitter,
    QTableView,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

# ---------------------------------------------------------------------------
# Persistence
# ---------------------------------------------------------------------------
SETTINGS_DIR = settings_dir("DATACOLISA")
CONTAINERS_FILE = SETTINGS_DIR / "containers_history.json"


def _load_history() -> Dict[str, Any]:
    if not CONTAINERS_FILE.exists():
        return {}
    try:
        return json.loads(CONTAINERS_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _save_history(data: Dict[str, Any]) -> None:
    try:
        SETTINGS_DIR.mkdir(parents=True, exist_ok=True)
        CONTAINERS_FILE.write_text(
            json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
        )
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Types d'echantillons connus
# ---------------------------------------------------------------------------
SAMPLE_KEYS = [
    "ecailles_brutes", "montees", "empreintes", "otolithes",
    "nageoires", "opercules", "vertebres", "maxillaires",
    "chair_lyophilisee", "muscle", "fraction_inconnue",
]

DEFAULT_CONTAINER_TYPES = ["TIROIR", "BOITE"]

# ---------------------------------------------------------------------------
# Decoration poisson (subtile, utilisee dans les titres de groupbox)
# ---------------------------------------------------------------------------
FISH_ICON = "🐟"   # affiche une seule fois dans le titre principal du dialogue


# ---------------------------------------------------------------------------
# Stylesheet — mode jour ET nuit, aérée, lisible
# Les couleurs fixes sont utilisées uniquement pour les éléments d'accentuation
# (titres de groupbox, sélection, bouton principal). Tout le reste s'appuie
# sur la palette Qt afin de fonctionner correctement en thème sombre.
# ---------------------------------------------------------------------------
DIALOG_STYLE = """

/* ── Fenêtre ─────────────────────────────────────────────────────── */
QDialog {
    background: palette(window);
    color:      palette(windowText);
}

/* ── GroupBox ────────────────────────────────────────────────────── */
QGroupBox {
    border:        1px solid palette(mid);
    border-radius: 8px;
    margin-top:    14px;
    padding:       10px 8px 8px 8px;
    font-weight:   bold;
    font-size:     11px;
    color:         palette(windowText);
}
QGroupBox::title {
    subcontrol-origin:   margin;
    subcontrol-position: top left;
    left:         12px;
    padding:      1px 6px;
    border-radius: 4px;
    background:   #2C7DA0;
    color:        #FFFFFF;
    font-size:    11px;
    font-weight:  bold;
}

/* ── Tableaux ────────────────────────────────────────────────────── */
QTableView, QTableWidget {
    gridline-color:            palette(mid);
    alternate-background-color: palette(alternateBase);
    selection-background-color: #2C7DA0;
    selection-color:            #FFFFFF;
    border:        1px solid palette(mid);
    border-radius: 5px;
    background:    palette(base);
    color:         palette(text);
    font-size:     11px;
}
QHeaderView::section {
    background:  #1B4965;
    color:       #FFFFFF;
    font-weight: bold;
    font-size:   10px;
    padding:     5px 8px;
    border:      none;
    border-right: 1px solid #2C6E8A;
}
QHeaderView::section:vertical {
    background:  palette(button);
    color:       palette(buttonText);
    border:      none;
    border-bottom: 1px solid palette(mid);
    padding:     2px 6px;
    font-weight: normal;
    font-size:   10px;
}
QHeaderView::section:last {
    border-right: none;
}

/* ── Boutons ─────────────────────────────────────────────────────── */
QPushButton {
    border-radius: 5px;
    padding:       5px 14px;
    font-size:     11px;
    font-weight:   bold;
    border:        1px solid palette(mid);
    background:    palette(button);
    color:         palette(buttonText);
    min-height:    26px;
}
QPushButton:hover  { background: palette(midlight); }
QPushButton:pressed { background: palette(dark); }

QPushButton#btn_enregistrer {
    background:  #2C7DA0;
    color:       #FFFFFF;
    border:      none;
    padding:     7px 22px;
    font-size:   12px;
    min-height:  32px;
}
QPushButton#btn_enregistrer:hover  { background: #1B5E7A; }
QPushButton#btn_enregistrer:pressed { background: #144A60; }

QPushButton#btn_annuler {
    background:  palette(button);
    color:       palette(buttonText);
    padding:     7px 22px;
    font-size:   12px;
    min-height:  32px;
}
QPushButton#btn_annuler:hover { background: palette(midlight); }

/* ── ComboBox ────────────────────────────────────────────────────── */
QComboBox {
    border:        1px solid palette(mid);
    border-radius: 5px;
    padding:       4px 8px;
    background:    palette(base);
    color:         palette(text);
    font-size:     11px;
    min-height:    26px;
}
QComboBox:focus { border: 1px solid #2C7DA0; }
QComboBox::drop-down {
    subcontrol-origin:   padding;
    subcontrol-position: top right;
    width:       22px;
    border-left: 1px solid palette(mid);
}
QComboBox QAbstractItemView {
    background: palette(base);
    color:      palette(text);
    selection-background-color: #2C7DA0;
    selection-color: #FFFFFF;
    border: 1px solid palette(mid);
}

/* ── LineEdit ────────────────────────────────────────────────────── */
QLineEdit {
    border:        1px solid palette(mid);
    border-radius: 5px;
    padding:       4px 8px;
    background:    palette(base);
    color:         palette(text);
    font-size:     11px;
    min-height:    26px;
}
QLineEdit:focus { border: 1px solid #2C7DA0; }

/* ── Labels ──────────────────────────────────────────────────────── */
QLabel { color: palette(windowText); font-size: 11px; }

/* ── Splitter ────────────────────────────────────────────────────── */
QSplitter::handle {
    background: palette(mid);
    width:      4px;
    margin:     0px 2px;
}

/* ── Scrollbars ──────────────────────────────────────────────────── */
QScrollBar:vertical {
    background: palette(base); width: 10px; margin: 0;
}
QScrollBar::handle:vertical {
    background: palette(mid); border-radius: 4px; min-height: 20px;
}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0; }

QScrollBar:horizontal {
    background: palette(base); height: 10px; margin: 0;
}
QScrollBar::handle:horizontal {
    background: palette(mid); border-radius: 4px; min-width: 20px;
}
QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal { width: 0; }

"""


def _is_dark_mode() -> bool:
    app = QApplication.instance()
    if app is None:
        return False
    role = getattr(QPalette, "ColorRole", QPalette)
    window_role = getattr(role, "Window", None) or getattr(QPalette, "window", None)
    if window_role is None:
        return False
    return app.palette().color(window_role).lightness() < 128


def _container_colors(container_value: str, dark: bool) -> tuple[QColor, QColor]:
    text = str(container_value or "").strip()
    if not text:
        return QColor("#ECFDF5"), QColor("#065F46")

    digest = hashlib.md5(text.encode("utf-8")).hexdigest()
    hue = int(digest[:6], 16) % 360

    if dark:
        background = QColor.fromHsv(hue, 110, 85)
        foreground = QColor("#F8FAFC")
    else:
        background = QColor.fromHsv(hue, 70, 245)
        foreground = QColor("#1F2937")
    return background, foreground


# ===========================================================================
# Modele de previsualisation
# ===========================================================================

class ExcelPreviewModel(QAbstractTableModel):

    def __init__(self, parent=None):
        super().__init__(parent)
        self._headers: List[str] = []
        self._rows: List[List[Any]] = []
        self._container_col: int = 0

    def load_from_file(self, path: Path, sheet_name: Optional[str] = None) -> None:
        self.beginResetModel()
        self._headers = []
        self._rows = []
        try:
            wb = load_workbook(str(path), read_only=True, data_only=True)
            ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            wb.close()
            if not all_rows:
                self.endResetModel()
                return
            file_headers = [str(h).strip() if h is not None else "" for h in all_rows[0]]
            self._headers = file_headers + ["container_parent_identifier"]
            self._container_col = len(file_headers)
            for raw_row in all_rows[1:]:
                row_data = list(raw_row)
                while len(row_data) < len(file_headers):
                    row_data.append(None)
                row_data.append("")
                self._rows.append(row_data)
        except Exception as exc:
            self._headers = [f"Erreur : {exc}"]
        self.endResetModel()

    def load_from_rows(self, rows: List[Dict[str, Any]]) -> None:
        """Charge les donnees depuis le tableau interne du logiciel."""
        self.beginResetModel()
        self._headers = []
        self._rows = []
        if not rows:
            self.endResetModel()
            return
        all_keys = list(rows[0].keys())
        priority = [
            "ref", "num_individu", "code_espece", "date_capture",
            "code_type_echantillon", "lac_riviere", "pays_capture",
            "ecailles_brutes", "montees", "empreintes", "otolithes",
            "longueur_mm", "poids_g", "age_total", "categorie",
            "type_peche", "autre_oss", "observation_disponibilite",
            "pecheur", "maturite", "sexe",
        ]
        ordered = [k for k in priority if k in all_keys]
        ordered += [k for k in all_keys if k not in ordered and k != "selected"]
        self._headers = ordered + ["container_parent_identifier"]
        self._container_col = len(ordered)
        for r in rows:
            row_data = [r.get(k, "") for k in ordered]
            row_data.append("")
            self._rows.append(row_data)
        self.endResetModel()

    def set_container_for_rows(self, row_indices: List[int], value: str) -> None:
        for i in row_indices:
            if 0 <= i < len(self._rows):
                self._rows[i][self._container_col] = value
        if row_indices:
            self.dataChanged.emit(
                self.index(min(row_indices), 0),
                self.index(max(row_indices), self._container_col),
            )

    def all_containers(self) -> Dict[int, str]:
        return {
            i: str(row[self._container_col])
            for i, row in enumerate(self._rows)
            if row[self._container_col]
        }

    def rowCount(self, parent=QModelIndex()):
        return len(self._rows)

    def columnCount(self, parent=QModelIndex()):
        return len(self._headers)

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if orientation == Qt.Horizontal:
            if role == Qt.DisplayRole:
                return self._headers[section] if section < len(self._headers) else ""
            if role == Qt.BackgroundRole:
                return QBrush(QColor("#2E4057"))
            if role == Qt.ForegroundRole:
                return QBrush(QColor("#FFFFFF"))
        if orientation == Qt.Vertical:
            container_value = ""
            if 0 <= section < len(self._rows) and self._container_col < len(self._rows[section]):
                container_value = self._rows[section][self._container_col] or ""
            if role == Qt.DisplayRole:
                return str(section + 1)
            if container_value and role == Qt.BackgroundRole:
                dark = _is_dark_mode()
                background, _ = _container_colors(str(container_value), dark)
                return QBrush(background.darker(125) if dark else background.darker(112))
            if container_value and role == Qt.ForegroundRole:
                dark = _is_dark_mode()
                _, foreground = _container_colors(str(container_value), dark)
                return QBrush(foreground)
            if container_value and role == Qt.FontRole:
                f = QFont(); f.setBold(True); return f
        return None

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid():
            return None
        row, col = index.row(), index.column()
        value = self._rows[row][col] if col < len(self._rows[row]) else None
        container_value = self._rows[row][self._container_col] if self._container_col < len(self._rows[row]) else ""

        if role == Qt.DisplayRole:
            return "" if value is None else str(value)

        if role == Qt.BackgroundRole and container_value:
            dark = _is_dark_mode()
            background, _ = _container_colors(str(container_value), dark)
            if col == self._container_col:
                return QBrush(background.darker(125) if dark else background.darker(108))
            return QBrush(background)

        if role == Qt.ForegroundRole and container_value:
            dark = _is_dark_mode()
            _, foreground = _container_colors(str(container_value), dark)
            return QBrush(foreground)

        if role == Qt.FontRole and container_value:
            f = QFont(); f.setBold(True); return f

        return None

    def flags(self, index):
        return Qt.ItemIsEnabled | Qt.ItemIsSelectable


# ===========================================================================
# Vue previsualisation
# ===========================================================================

class ExcelPreviewView(QTableView):
    def __init__(self, model: ExcelPreviewModel, parent=None):
        super().__init__(parent)
        self.setModel(model)
        self.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.setAlternatingRowColors(True)
        self.setShowGrid(True)
        self.verticalHeader().setVisible(True)
        self.verticalHeader().setDefaultSectionSize(22)
        self.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.horizontalHeader().setMinimumSectionSize(60)
        self.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

    def selected_row_indices(self) -> List[int]:
        return sorted({idx.row() for idx in self.selectedIndexes()})


# ===========================================================================
# Contenant
# ===========================================================================

class ContainerEntry:
    def __init__(self, ctype: str, label: str, rows: Optional[List[int]] = None):
        self.ctype = ctype
        self.label = label
        self.rows: List[int] = rows or []

    @property
    def full_label(self) -> str:
        return f"{self.ctype} {self.label}".strip() if self.label.strip() else self.ctype


class ContainerPanel(QWidget):
    assignment_changed = Signal()

    def __init__(self, preview_model: ExcelPreviewModel, parent=None):
        super().__init__(parent)
        self._preview_model = preview_model
        self._container_types: List[str] = list(DEFAULT_CONTAINER_TYPES)
        self._entries: List[ContainerEntry] = []
        self._build_ui()

    def _build_ui(self):
        main = QVBoxLayout(self)
        main.setSpacing(0)
        main.setContentsMargins(0, 0, 0, 0)

        grp = QGroupBox(f"{FISH_ICON}  Contenants - TIROIR / BOITE")
        v = QVBoxLayout(grp)
        v.setSpacing(10)
        v.setContentsMargins(12, 14, 12, 12)

        # ── Type de contenant ──────────────────────────────────────
        lbl_type = QLabel("Type de contenant")
        lbl_type.setStyleSheet("font-weight:bold; font-size:11px; color:palette(windowText);")
        v.addWidget(lbl_type)

        type_row = QHBoxLayout()
        type_row.setSpacing(6)
        self._type_combo = QComboBox()
        self._type_combo.addItems(self._container_types)
        btn_plus = QPushButton("+")
        btn_plus.setToolTip("Ajouter un type personnalise")
        btn_plus.setFixedSize(30, 30)
        btn_plus.setStyleSheet(
            "QPushButton { background:#2A9D5C; color:white; font-weight:bold;"
            "              border-radius:5px; border:none; font-size:16px; }"
            "QPushButton:hover { background:#1E7A45; }"
        )
        btn_plus.clicked.connect(self._add_custom_type)
        type_row.addWidget(self._type_combo, 1)
        type_row.addWidget(btn_plus)
        v.addLayout(type_row)

        # ── Numéro ─────────────────────────────────────────────────
        lbl_num = QLabel("Numero")
        lbl_num.setStyleSheet("font-weight:bold; font-size:11px; color:palette(windowText);")
        v.addWidget(lbl_num)

        self._num_edit = QLineEdit()
        self._num_edit.setPlaceholderText("ex :  123  ou  11B")
        v.addWidget(self._num_edit)

        # ── Aperçu ─────────────────────────────────────────────────
        self._prev_lbl = QLabel("")
        self._prev_lbl.setAlignment(Qt.AlignCenter)
        self._prev_lbl.setMinimumHeight(28)
        self._prev_lbl.setStyleSheet(
            "font-weight:bold; font-size:12px; padding:4px 8px;"
            "border-radius:5px; background:#D1FAE5; color:#065F46;"
        )
        self._type_combo.currentTextChanged.connect(self._refresh_prev)
        self._num_edit.textChanged.connect(self._refresh_prev)
        v.addWidget(self._prev_lbl)

        # ── Bouton Assigner ─────────────────────────────────────────
        btn_assign = QPushButton("Assigner aux lignes selectionnees")
        btn_assign.setObjectName("btn_enregistrer")
        btn_assign.clicked.connect(self._assign)
        v.addWidget(btn_assign)

        # ── Séparateur ──────────────────────────────────────────────
        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setStyleSheet("color: palette(mid); margin: 4px 0;")
        v.addWidget(sep)

        # ── Tableau des correspondances ─────────────────────────────
        lbl_table = QLabel("Correspondances assignees")
        lbl_table.setStyleSheet("font-weight:bold; font-size:11px; color:palette(windowText);")
        v.addWidget(lbl_table)

        self._table = QTableWidget(0, 3)
        self._table.setHorizontalHeaderLabels(["Contenant", "Lignes", ""])
        self._table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self._table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self._table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Fixed)
        self._table.setColumnWidth(2, 34)
        self._table.verticalHeader().setVisible(False)
        self._table.setSelectionMode(QAbstractItemView.SingleSelection)
        self._table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self._table.setAlternatingRowColors(True)
        self._table.setMinimumHeight(120)
        v.addWidget(self._table, 1)

        main.addWidget(grp, 1)
        self._refresh_prev()

    def _refresh_prev(self):
        ctype = self._type_combo.currentText().strip()
        num   = self._num_edit.text().strip()
        full_label = f"{ctype} {num}".strip() if num else ctype
        self._prev_lbl.setText(f"->  {full_label}" if full_label else "")
        self._prev_lbl.setStyleSheet(
            "font-weight:bold; font-size:12px; padding:4px 8px; border-radius:5px;"
            "background:#D1FAE5; color:#065F46;"
        )

    def _add_custom_type(self):
        text, ok = QInputDialog.getText(self, "Nouveau type", "Nom du type de contenant :")
        if ok and text.strip():
            name = text.strip()
            if name not in self._container_types:
                self._container_types.append(name)
                self._type_combo.addItem(name)
            self._type_combo.setCurrentText(name)

    def _assign(self):
        view = self._find_preview_view()
        if view is None:
            return
        selected = view.selected_row_indices()
        if not selected:
            QMessageBox.information(self, "Aucune selection",
                "Selectionne d'abord des lignes dans le tableau de gauche.")
            return
        ctype = self._type_combo.currentText().strip()
        num = self._num_edit.text().strip()
        if not ctype:
            QMessageBox.warning(self, "Type manquant", "Choisis un type de contenant.")
            return
        full_label = f"{ctype} {num}".strip() if num else ctype

        for e in self._entries:
            e.rows = [r for r in e.rows if r not in selected]

        self._entries.append(ContainerEntry(ctype, num, sorted(selected)))
        self._entries = [e for e in self._entries if e.rows]
        self._preview_model.set_container_for_rows(selected, full_label)
        self._refresh_table()
        self.assignment_changed.emit()

    def _find_preview_view(self) -> Optional[ExcelPreviewView]:
        p = self.parent()
        while p is not None:
            v = p.findChild(ExcelPreviewView)
            if v:
                return v
            p = p.parent() if hasattr(p, "parent") else None
        return None

    def _refresh_table(self):
        self._table.setRowCount(0)
        for i, entry in enumerate(self._entries):
            self._table.insertRow(i)
            item = QTableWidgetItem(entry.full_label)
            item.setForeground(QBrush(QColor("#2E86AB")))
            f = QFont(); f.setBold(True); item.setFont(f)
            self._table.setItem(i, 0, item)
            item_r = QTableWidgetItem(", ".join(str(r+1) for r in entry.rows))
            item_r.setTextAlignment(Qt.AlignCenter)
            self._table.setItem(i, 1, item_r)
            btn = QPushButton("X")
            btn.setFixedSize(22, 22)
            btn.setStyleSheet("QPushButton{background:#dc2626;color:white;font-weight:bold;"
                              "border-radius:3px;}QPushButton:hover{background:#b91c1c;}")
            btn.clicked.connect(lambda c, idx=i: self._remove_entry(idx))
            cell = QWidget(); hl = QHBoxLayout(cell)
            hl.setContentsMargins(2, 1, 2, 1); hl.addWidget(btn)
            self._table.setCellWidget(i, 2, cell)
            self._table.setRowHeight(i, 24)

    def _remove_entry(self, index: int):
        if 0 <= index < len(self._entries):
            self._preview_model.set_container_for_rows(self._entries[index].rows, "")
            self._entries.pop(index)
            self._refresh_table()
            self.assignment_changed.emit()

    def get_entries(self) -> List[ContainerEntry]:
        return list(self._entries)


# ===========================================================================
# Dialog principal
# ===========================================================================

class CollecScienceSourceDialog(QDialog):
    """
    Dialog unique Collec-Science.
    Accepte soit colisa_path (fichier Excel) soit software_rows (donnees logiciel).
    """

    def __init__(
        self,
        parent=None,
        colisa_path: Optional[Path] = None,
        sheet_name: Optional[str] = None,
        software_rows: Optional[List[Dict[str, Any]]] = None,
        source_label: Optional[str] = None,
    ):
        super().__init__(parent)
        self.setWindowTitle("Collec-Science - Selection et contenant")
        self.setMinimumSize(1100, 640)
        self.setModal(True)
        self.setStyleSheet(DIALOG_STYLE)

        self._colisa_path = colisa_path
        self._software_rows = software_rows or []
        self._preview_model = ExcelPreviewModel()

        if source_label:
            if software_rows is not None:
                self._preview_model.load_from_rows(software_rows)
            elif colisa_path and colisa_path.exists():
                self._preview_model.load_from_file(colisa_path, sheet_name)
            self._source_label = source_label
        elif software_rows is not None:
            self._preview_model.load_from_rows(software_rows)
            self._source_label = f"{len(software_rows)} echantillon(s) depuis le logiciel"
        elif colisa_path and colisa_path.exists():
            self._preview_model.load_from_file(colisa_path, sheet_name)
            self._source_label = f"Fichier : {colisa_path.name}"
        else:
            self._source_label = "Aucune source chargee"

        self._build_ui()

    def _build_ui(self):
        main = QVBoxLayout(self)
        main.setSpacing(12)
        main.setContentsMargins(16, 14, 16, 14)

        # ── Titre + source ──────────────────────────────────────────
        titre = QLabel(f"{FISH_ICON}  Collec-Science - Preparation de l'export")
        titre.setStyleSheet(
            "font-size:14px; font-weight:bold; color:palette(windowText);"
            "padding-bottom:2px;"
        )
        main.addWidget(titre)

        src = QLabel(f"Source :  {self._source_label}")
        src.setStyleSheet(
            "font-size:10px; font-style:italic; color:palette(mid);"
            "padding-bottom:4px;"
        )
        main.addWidget(src)

        sep_top = QFrame()
        sep_top.setFrameShape(QFrame.HLine)
        sep_top.setStyleSheet("color: palette(mid);")
        main.addWidget(sep_top)

        # ═══════════════════════════════════════════════════════════
        # BANDE DU HAUT — Partie 1  |  Partie 2
        # ═══════════════════════════════════════════════════════════
        top_band = QHBoxLayout()
        top_band.setSpacing(14)

        # ── Partie 1 : Contenants ───────────────────────────────────
        grp_cont = QGroupBox("1 - Contenants")
        v_cont = QVBoxLayout(grp_cont)
        v_cont.setSpacing(6)
        v_cont.setContentsMargins(12, 14, 12, 10)

        instructions = QLabel(
            "1. Selectionne les lignes dans le tableau ci-dessous\n"
            "2. Choisis le type et le numero de contenant a droite\n"
            "3. Clique <b>Assigner</b> - la colonne "
            "<b>container_parent_identifier</b> est remplie"
        )
        instructions.setTextFormat(Qt.RichText)
        instructions.setWordWrap(True)
        instructions.setStyleSheet(
            "font-size:11px; padding:6px 10px; border-radius:5px;"
            "background:palette(alternateBase); color:palette(windowText);"
            "font-weight:normal; line-height:160%;"
        )
        v_cont.addWidget(instructions)
        top_band.addWidget(grp_cont, 1)

        # ── Partie 2 : Colonne md_num_individu ─────────────────────
        # Uniquement depuis un fichier Excel externe ; depuis le logiciel le
        # num_individu est déjà renseigné par l'import → on ne crée pas le widget.
        self._md_num_individu_combo: Optional[QComboBox] = None

        if not self._software_rows:
            grp_md = QGroupBox("2 - Colonne md_num_individu")
            v_md = QVBoxLayout(grp_md)
            v_md.setSpacing(8)
            v_md.setContentsMargins(12, 14, 12, 10)

            desc_md = QLabel(
                "Quelle colonne de l'Excel source doit alimenter\n"
                "le champ <b>md_num_individu</b> dans le fichier genere ?"
            )
            desc_md.setTextFormat(Qt.RichText)
            desc_md.setWordWrap(True)
            desc_md.setStyleSheet(
                "font-size:11px; padding:6px 10px; border-radius:5px;"
                "background:palette(alternateBase); color:palette(windowText);"
                "font-weight:normal; line-height:160%;"
            )
            v_md.addWidget(desc_md)

            lbl_col = QLabel("Colonne choisie :")
            lbl_col.setStyleSheet("font-size:11px; font-weight:bold; color:palette(windowText);")
            v_md.addWidget(lbl_col)

            self._md_num_individu_combo = QComboBox()
            self._md_num_individu_combo.addItem(
                "Colonne fixe \"Numero individu\" (par defaut)", userData=None
            )

            excel_src = self._colisa_path if self._colisa_path is not None else None
            if excel_src and excel_src.exists():
                try:
                    wb_tmp = load_workbook(str(excel_src), read_only=True, data_only=True)
                    ws_tmp = wb_tmp.active
                    header_row = next(
                        ws_tmp.iter_rows(min_row=1, max_row=1, values_only=True), ()
                    )
                    wb_tmp.close()
                    for col_idx, hdr in enumerate(header_row):
                        label = str(hdr).strip() if hdr is not None else f"(col {col_idx + 1})"
                        self._md_num_individu_combo.addItem(
                            f"Col {col_idx + 1} - {label}", userData=col_idx
                        )
                except Exception:
                    pass

            v_md.addWidget(self._md_num_individu_combo)
            v_md.addStretch()
            top_band.addWidget(grp_md, 1)

        main.addLayout(top_band)

        # ═══════════════════════════════════════════════════════════
        # SPLITTER : tableau sélection | panel contenants
        # ═══════════════════════════════════════════════════════════
        splitter = QSplitter(Qt.Horizontal)
        splitter.setChildrenCollapsible(False)

        # Gauche — tableau de sélection des lignes
        left = QWidget()
        lv = QVBoxLayout(left)
        lv.setContentsMargins(0, 0, 0, 0)
        lv.setSpacing(4)
        grp_l = QGroupBox("Echantillons - selection des lignes")
        gl = QVBoxLayout(grp_l)
        gl.setSpacing(6)
        gl.setContentsMargins(8, 12, 8, 8)
        hint = QLabel("Clic | Maj+clic pour une plage | Ctrl+clic pour plusieurs lignes")
        hint.setStyleSheet("font-size:10px; color:palette(mid); font-style:italic;")
        gl.addWidget(hint)
        self._preview_view = ExcelPreviewView(self._preview_model)
        gl.addWidget(self._preview_view)
        lv.addWidget(grp_l)
        splitter.addWidget(left)

        # Droite — panel d'assignation des contenants
        right = QWidget()
        rv = QVBoxLayout(right)
        rv.setContentsMargins(0, 0, 0, 0)
        self._container_panel = ContainerPanel(self._preview_model, parent=self)
        rv.addWidget(self._container_panel)
        splitter.addWidget(right)

        splitter.setSizes([680, 380])
        main.addWidget(splitter, 1)

        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setStyleSheet("color: palette(mid); margin: 2px 0;")
        main.addWidget(sep)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(10)
        btn_row.addStretch()
        btn_annuler = QPushButton("  Annuler  ")
        btn_annuler.setObjectName("btn_annuler")
        btn_annuler.clicked.connect(self.reject)
        btn_ok = QPushButton("  Enregistrer  ")
        btn_ok.setObjectName("btn_enregistrer")
        btn_ok.clicked.connect(self._on_ok)
        btn_row.addWidget(btn_annuler)
        btn_row.addWidget(btn_ok)
        main.addLayout(btn_row)

    def _on_ok(self):
        self.accept()

    def get_container_assignments(self) -> Dict[int, str]:
        return self._preview_model.all_containers()

    def get_selected_row_indices(self) -> List[int]:
        selected = self._preview_view.selected_row_indices()
        if selected:
            return selected
        if self._software_rows:
            return list(range(len(self._software_rows)))
        return []

    def get_selected_software_rows(self) -> List[Dict[str, Any]]:
        if not self._software_rows:
            return []
        selected = self.get_selected_row_indices()
        return [self._software_rows[i] for i in selected if 0 <= i < len(self._software_rows)]

    def build_containers_dict(self) -> Dict[str, Any]:
        """
        Construit le dict containers pour generer_collec_science().
        Format : {sample_key: "1-5=BOITE 123\\n6-10=TIROIR 11B"}
        """
        entries = self._container_panel.get_entries() if hasattr(self, "_container_panel") else []
        if not entries:
            return {}

        from generer_collec_science import (
            COLISA_REQUIRED_HEADERS,
            infer_sample_key_from_type,
            normalize_header,
            valeur_present,
        )

        headers = getattr(self._preview_model, "_headers", [])
        rows = getattr(self._preview_model, "_rows", [])

        def _row_dict(row_idx: int) -> Dict[str, Any]:
            if row_idx < 0 or row_idx >= len(rows):
                return {}
            values = rows[row_idx]
            row: Dict[str, Any] = {}
            for col_idx in range(min(len(headers), len(values))):
                header = headers[col_idx]
                if header == "container_parent_identifier":
                    continue
                value = values[col_idx]
                row[header] = value

                normalized = normalize_header(header)
                if normalized in SAMPLE_KEYS:
                    row[normalized] = value
                    continue

                for key, aliases in COLISA_REQUIRED_HEADERS.items():
                    normalized_aliases = {normalize_header(alias) for alias in aliases}
                    if normalized in normalized_aliases:
                        row[key] = value
                        break
            return row

        def _sample_keys_for_row(row_idx: int) -> List[str]:
            row = _row_dict(row_idx)
            explicit_keys = [
                key for key in SAMPLE_KEYS
                if valeur_present(row.get(key))
            ]
            if explicit_keys:
                return explicit_keys
            inferred = infer_sample_key_from_type(row.get("code_type_echantillon"))
            return [inferred] if inferred in SAMPLE_KEYS else []

        containers_by_sample: Dict[str, str] = {}
        for entry in entries:
            if not entry.rows:
                continue
            grouped_rows: Dict[str, List[int]] = {}
            for row_idx in sorted(entry.rows):
                preview_row_number = row_idx + 1
                for sample_key in _sample_keys_for_row(row_idx):
                    grouped_rows.setdefault(sample_key, []).append(preview_row_number)

            for sample_key, sample_rows in grouped_rows.items():
                rows_s = sorted(sample_rows)
                rule_lines: List[str] = []
                start = end = rows_s[0]
                for r in rows_s[1:]:
                    if r == end + 1:
                        end = r
                    else:
                        rule_lines.append(f"{start}-{end}={entry.full_label}")
                        start = end = r
                rule_lines.append(f"{start}-{end}={entry.full_label}")

                existing_text = containers_by_sample.get(sample_key, "")
                new_text = "\n".join(rule_lines)
                containers_by_sample[sample_key] = (
                    f"{existing_text}\n{new_text}".strip() if existing_text else new_text
                )

        return containers_by_sample

    def use_fixed_md_num_individu_column(self) -> bool:
        """Retro-compatibilite : True si aucune colonne specifique n'est choisie."""
        return self.get_md_num_individu_column_index() is None

    def get_md_num_individu_column_index(self) -> Optional[int]:
        """
        Retourne l'index (0-based) de la colonne choisie pour md_num_individu,
        ou None si on doit utiliser la colonne fixe fallback (comportement historique).
        """
        if self._md_num_individu_combo is None:
            return None
        return self._md_num_individu_combo.currentData()


# ===========================================================================
# AnomaliesDialog - requis par le_visage.py (inchange)
# ===========================================================================

class AnomaliesDialog(QDialog):
    Accepted = 1

    def __init__(self, parent=None, sans_date: list | None = None, sans_espece: list | None = None):
        super().__init__(parent)
        self.setWindowTitle("Anomalies detectees")
        self.setMinimumWidth(420)
        self.setModal(True)
        self.forcer_injection = False
        sans_date = sans_date or []
        sans_espece = sans_espece or []

        main = QVBoxLayout(self)
        main.setSpacing(12)
        main.setContentsMargins(16, 16, 16, 16)

        titre = QLabel("Des lignes ont des donnees manquantes")
        titre.setStyleSheet("font-weight:bold;font-size:13px;color:#b45309;")
        main.addWidget(titre)

        if sans_date:
            grp = QGroupBox(f"Sans date ({len(sans_date)} ligne(s))")
            grp.setStyleSheet("QGroupBox{color:#dc2626;font-weight:bold;}")
            v = QVBoxLayout(grp)
            refs = ", ".join(sans_date[:20])
            if len(sans_date) > 20:
                refs += f" ... (+{len(sans_date)-20} autres)"
            lbl = QLabel(refs); lbl.setWordWrap(True); v.addWidget(lbl)
            main.addWidget(grp)

        if sans_espece:
            grp2 = QGroupBox(f"Sans espece ({len(sans_espece)} ligne(s))")
            grp2.setStyleSheet("QGroupBox{color:#dc2626;font-weight:bold;}")
            v2 = QVBoxLayout(grp2)
            refs2 = ", ".join(sans_espece[:20])
            if len(sans_espece) > 20:
                refs2 += f" ... (+{len(sans_espece)-20} autres)"
            lbl2 = QLabel(refs2); lbl2.setWordWrap(True); v2.addWidget(lbl2)
            main.addWidget(grp2)

        main.addWidget(QLabel("Voulez-vous quand meme les injecter ?"))

        btn_row = QHBoxLayout()
        btn_oui = QPushButton("Oui, tout injecter")
        btn_oui.setFixedHeight(34)
        btn_oui.setStyleSheet("font-weight:bold;background-color:#16a34a;color:white;")
        btn_non = QPushButton("Non, ignorer ces lignes")
        btn_non.setFixedHeight(34)
        btn_non.setStyleSheet("font-weight:bold;background-color:#dc2626;color:white;")
        btn_ann = QPushButton("Annuler l'import")
        btn_ann.setFixedHeight(34)
        btn_oui.clicked.connect(self._forcer)
        btn_non.clicked.connect(self.accept)
        btn_ann.clicked.connect(self.reject)
        btn_row.addWidget(btn_oui)
        btn_row.addWidget(btn_non)
        btn_row.addWidget(btn_ann)
        main.addLayout(btn_row)

    def _forcer(self):
        self.forcer_injection = True
        self.accept()


