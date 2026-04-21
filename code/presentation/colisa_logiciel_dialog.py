"""Dialog for preparing rows before generating COLISA logiciel format."""

from __future__ import annotations

import unicodedata
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook

from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QCheckBox,
    QComboBox,
    QDialog,
    QGridLayout,
    QHBoxLayout,
    QLabel,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
)


def _normalize(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return " ".join(text.split())


def load_template_options(template_path: Path) -> Dict[str, List[str]]:
    options = {
        "sites_atelier": [],
        "types_echantillon": [],
        "types_labels": {},
        "stades": [],
        "sens_migratoires": [],
        "maturites": [],
        "sexes": [],
    }
    if not template_path.exists():
        return options

    workbook = load_workbook(template_path, read_only=True, data_only=True)
    try:
        for sheet_name in workbook.sheetnames:
            norm_name = _normalize(sheet_name)
            worksheet = workbook[sheet_name]

            if "site" in norm_name and "atelier" in norm_name:
                values = _read_first_column(worksheet)
                options["sites_atelier"] = values
            elif "type" in norm_name and "echantillon" in norm_name:
                rows = list(worksheet.iter_rows(min_row=2, values_only=True))
                labels: Dict[str, str] = {}
                values: List[str] = []
                for row in rows:
                    code = str(row[0] or "").strip()
                    label = str(row[1] or "").strip()
                    if not code:
                        continue
                    values.append(code)
                    labels[code] = label
                options["types_echantillon"] = values
                options["types_labels"] = labels
            elif "stade" in norm_name:
                options["stades"] = _read_first_column(worksheet)
            elif "sens" in norm_name and "migr" in norm_name:
                options["sens_migratoires"] = _read_first_column(worksheet)
            elif "maturit" in norm_name:
                options["maturites"] = _read_first_column(worksheet)
            elif norm_name == "sexes":
                options["sexes"] = _read_first_column(worksheet)
        return options
    finally:
        workbook.close()


def _read_first_column(worksheet) -> List[str]:
    values: List[str] = []
    seen = set()
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        value = str((row[0] if row else "") or "").strip()
        if not value or value in seen:
            continue
        seen.add(value)
        values.append(value)
    return values


def _row_observations(row: Dict[str, Any]) -> str:
    parts: List[str] = []
    for key in ("observations", "information_disponibilite"):
        value = str(row.get(key) or "").strip()
        if value in {"+", "++", "+++"}:
            continue
        if value and value not in parts:
            parts.append(value)
    return " | ".join(parts)


class SampleTypeChoiceCombo(QComboBox):
    def __init__(self, values: List[str], labels: Dict[str, str], current_value: str, parent=None):
        super().__init__(parent)
        self.setEditable(True)
        self.setInsertPolicy(QComboBox.NoInsert)
        self._labels = labels
        self.addItem("", "")
        for value in values:
            label = labels.get(value, "")
            display = f"{value} - {label}" if label else value
            self.addItem(display, value)
        self.set_current_value(current_value)

    def set_current_value(self, current_value: str) -> None:
        value = str(current_value or "").strip()
        for index in range(self.count()):
            if str(self.itemData(index) or "") == value:
                self.setCurrentIndex(index)
                return
        self.setEditText(value)

    def current_code(self) -> str:
        data = self.currentData()
        if data is not None and str(data).strip():
            return str(data).strip()
        text = self.currentText().strip()
        if " - " in text:
            return text.split(" - ", 1)[0].strip()
        return text


class ColisaLogicielPreparationDialog(QDialog):
    def __init__(self, rows: List[Dict[str, Any]], template_path: Path, default_site_atelier: str = "", parent=None):
        super().__init__(parent)
        self._rows = [dict(row) for row in rows]
        self._options = load_template_options(template_path)
        self._default_site_atelier = default_site_atelier
        self.setWindowTitle("Apercu du fichier COLISA a generer")
        self.setMinimumSize(1400, 650)

        if parent and parent.styleSheet():
            self.setStyleSheet(parent.styleSheet())

        self._build_ui()
        self._load_rows()

    def _build_ui(self) -> None:
        main = QVBoxLayout(self)
        help_label = QLabel(
            "Apercu du fichier qui va etre cree. Tu peux cocher les lignes a generer et appliquer des valeurs en lot."
        )
        help_label.setWordWrap(True)
        main.addWidget(help_label)

        bulk_box = QVBoxLayout()
        bulk_form = QGridLayout()
        self.bulk_site = self._combo(self._options["sites_atelier"], self._default_site_atelier)
        self.bulk_code_site = self._combo(["V1235003"], "V1235003")
        self.bulk_type = SampleTypeChoiceCombo(self._options["types_echantillon"], self._options["types_labels"], "")
        self.bulk_stade = self._combo(self._options["stades"], "")
        self.bulk_sens = self._combo(self._options["sens_migratoires"], "")
        self.bulk_maturite = self._combo([], "")
        self.bulk_sexe = self._combo(self._options["sexes"], "")
        self.bulk_otg = self._combo(["", "0", "1"], "")
        self.bulk_otd = self._combo(["", "0", "1"], "")
        self.bulk_obs = self._combo([], "")

        bulk_fields = [
            ("Site", self.bulk_site),
            ("Code site", self.bulk_code_site),
            ("Type echantillon", self.bulk_type),
            ("Stade", self.bulk_stade),
            ("Sens migratoire", self.bulk_sens),
            ("Maturite", self.bulk_maturite),
            ("Sexe", self.bulk_sexe),
            ("Otolithe G", self.bulk_otg),
            ("Otolithe D", self.bulk_otd),
            ("Observations", self.bulk_obs),
        ]
        for index, (label, widget) in enumerate(bulk_fields):
            bulk_form.addWidget(QLabel(label), index // 4, (index % 4) * 2)
            bulk_form.addWidget(widget, index // 4, (index % 4) * 2 + 1)
        bulk_box.addLayout(bulk_form)

        bulk_buttons = QHBoxLayout()
        btn_all = QPushButton("Tout selectionner")
        btn_none = QPushButton("Vider selection")
        btn_apply = QPushButton("Appliquer aux lignes cochees")
        btn_all.clicked.connect(lambda: self._set_all_checked(True))
        btn_none.clicked.connect(lambda: self._set_all_checked(False))
        btn_apply.clicked.connect(self._apply_bulk_to_checked)
        bulk_buttons.addWidget(btn_all)
        bulk_buttons.addWidget(btn_none)
        bulk_buttons.addWidget(btn_apply)
        bulk_buttons.addStretch()
        bulk_box.addLayout(bulk_buttons)
        main.addLayout(bulk_box)

        self.table = QTableWidget(0, 13)
        self.table.setEditTriggers(QTableWidget.AllEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectItems)
        self.table.setHorizontalHeaderLabels([
            "OK",
            "Code echantillon",
            "Numero individu",
            "Code type echantillon",
            "Site atelier",
            "Code site",
            "Code stade",
            "Sens migratoire",
            "Maturite",
            "Sexe",
            "Otolithe G",
            "Otolithe D",
            "Observations",
        ])
        self.table.horizontalHeader().setStretchLastSection(True)
        main.addWidget(self.table, 1)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_cancel = QPushButton("Annuler")
        btn_ok = QPushButton("Generer")
        btn_cancel.clicked.connect(self.reject)
        btn_ok.clicked.connect(self.accept)
        btn_row.addWidget(btn_cancel)
        btn_row.addWidget(btn_ok)
        main.addLayout(btn_row)

    def _combo(self, values: List[str], current_value: str) -> QComboBox:
        combo = QComboBox()
        combo.setEditable(True)
        combo.setInsertPolicy(QComboBox.NoInsert)
        combo.addItem("")
        for value in values:
            combo.addItem(value)
        combo.setCurrentText(str(current_value or ""))
        return combo

    def _load_rows(self) -> None:
        self.table.setRowCount(len(self._rows))
        for row_index, row in enumerate(self._rows):
            checkbox = QTableWidgetItem("")
            checkbox.setFlags(Qt.ItemIsEnabled | Qt.ItemIsUserCheckable)
            checkbox.setCheckState(Qt.Checked)
            self.table.setItem(row_index, 0, checkbox)

            self._set_readonly_item(row_index, 1, row.get("code_echantillon", "") or row.get("ref", ""))
            self._set_readonly_item(row_index, 2, row.get("num_individu", ""))

            self._set_editable_item(row_index, 3, str(row.get("code_type_echantillon", "")))
            self._set_editable_item(row_index, 4, str(row.get("site_atelier", "") or self._default_site_atelier))
            self._set_editable_item(row_index, 5, str(row.get("code_site", "V1235003")))
            self._set_editable_item(row_index, 6, str(row.get("code_stade", "")))
            self._set_editable_item(row_index, 7, str(row.get("sens_migratoire", "")))
            self._set_editable_item(row_index, 8, str(row.get("maturite", "")))
            self._set_editable_item(row_index, 9, str(row.get("sexe", "")))
            self._set_editable_item(row_index, 10, str(row.get("otolithe_gauche", "")))
            self._set_editable_item(row_index, 11, str(row.get("otolithe_droit", "")))
            self._set_editable_item(row_index, 12, _row_observations(row))

        self.table.resizeColumnsToContents()

    def _set_all_checked(self, checked: bool) -> None:
        state = Qt.Checked if checked else Qt.Unchecked
        for row_index in range(self.table.rowCount()):
            item = self.table.item(row_index, 0)
            if item is not None:
                item.setCheckState(state)

    def _checked_row_indexes(self) -> List[int]:
        out: List[int] = []
        for row_index in range(self.table.rowCount()):
            item = self.table.item(row_index, 0)
            if item is not None and item.checkState() == Qt.Checked:
                out.append(row_index)
        return out

    def _apply_bulk_to_checked(self) -> None:
        for row_index in self._checked_row_indexes():
            self._set_item_text_if_not_empty(row_index, 4, self.bulk_site.currentText().strip())
            self._set_item_text_if_not_empty(row_index, 5, self.bulk_code_site.currentText().strip())
            self._set_item_text_if_not_empty(row_index, 3, self.bulk_type.currentText().strip())
            self._set_item_text_if_not_empty(row_index, 6, self.bulk_stade.currentText().strip())
            self._set_item_text_if_not_empty(row_index, 7, self.bulk_sens.currentText().strip())
            self._set_item_text_if_not_empty(row_index, 8, self.bulk_maturite.currentText().strip())
            self._set_item_text_if_not_empty(row_index, 9, self.bulk_sexe.currentText().strip())
            self._set_item_text_if_not_empty(row_index, 10, self.bulk_otg.currentText().strip())
            self._set_item_text_if_not_empty(row_index, 11, self.bulk_otd.currentText().strip())
            self._set_item_text_if_not_empty(row_index, 12, self.bulk_obs.currentText().strip())

    def _set_readonly_item(self, row_index: int, col_index: int, value: Any) -> None:
        item = QTableWidgetItem(str(value or ""))
        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
        self.table.setItem(row_index, col_index, item)

    def _set_editable_item(self, row_index: int, col_index: int, value: Any) -> None:
        item = QTableWidgetItem(str(value or ""))
        self.table.setItem(row_index, col_index, item)

    def _set_item_text_if_not_empty(self, row_index: int, col_index: int, value: str) -> None:
        if not str(value or "").strip():
            return
        item = self.table.item(row_index, col_index)
        if item is None:
            item = QTableWidgetItem("")
            self.table.setItem(row_index, col_index, item)
        item.setText(str(value).strip())

    def get_rows(self) -> List[Dict[str, Any]]:
        out: List[Dict[str, Any]] = []
        for row_index in self._checked_row_indexes():
            row = self._rows[row_index]
            updated = dict(row)
            type_text = self.table.item(row_index, 3).text().strip() if self.table.item(row_index, 3) else ""
            updated["code_type_echantillon"] = type_text.split(" - ", 1)[0].strip() if " - " in type_text else type_text
            updated["site_atelier"] = self.table.item(row_index, 4).text().strip() if self.table.item(row_index, 4) else ""
            updated["code_site"] = self.table.item(row_index, 5).text().strip() if self.table.item(row_index, 5) else ""
            updated["code_stade"] = self.table.item(row_index, 6).text().strip() if self.table.item(row_index, 6) else ""
            updated["sens_migratoire"] = self.table.item(row_index, 7).text().strip() if self.table.item(row_index, 7) else ""
            updated["maturite"] = self.table.item(row_index, 8).text().strip() if self.table.item(row_index, 8) else ""
            updated["sexe"] = self.table.item(row_index, 9).text().strip() if self.table.item(row_index, 9) else ""
            updated["otolithe_gauche"] = self.table.item(row_index, 10).text().strip() if self.table.item(row_index, 10) else ""
            updated["otolithe_droit"] = self.table.item(row_index, 11).text().strip() if self.table.item(row_index, 11) else ""
            updated["observations"] = self.table.item(row_index, 12).text().strip() if self.table.item(row_index, 12) else ""
            out.append(updated)
        return out
