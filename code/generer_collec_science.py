#!/usr/bin/env python3
"""
Genere le fichier Collec-Science (Excel) depuis un fichier COLISA importe.
"""
from __future__ import annotations

import csv
import datetime
import math
import re
import unicodedata
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

from openpyxl import Workbook, load_workbook
from domain.value_objects import DateCapture
from infrastructure.file_value_normalizer import coerce_numeric_string
from infrastructure.internal_target_workbook import build_numero_identification_value

COUNTRY_CODES = {
    "france": "FR", "suisse": "CH", "switzerland": "CH",
    "allemagne": "DE", "germany": "DE", "italie": "IT",
    "italy": "IT", "espagne": "ES", "spain": "ES",
    "belgique": "BE", "belgium": "BE",
}

SAMPLE_TYPES = {
    "ecailles_brutes": (1, "ECAILLES BRUTES 1", None),
    "montees": (2, "ECAILLE MONTEE 2", "-MON"),
    "empreintes": (3, "EMPREINTE 3", None),
    "opercules": (4, "OPERCULES 4", None),
    "otolithes": (5, "OTOLITHE 5", "-OT"),
    "vertebres": (6, "VERTEBRE 6", None),
    "maxillaires": (7, "MAXILLAIRES 7", None),
    "chair_lyophilisee": (8, "CHAIR LYOPHILISEE 8", None),
    "nageoires": (9, "NAGEOIRE 9", None),
    "muscle": (10, "MUSCLE 10", None),
    "fraction_inconnue": (11, "FRACTION INCONNUE 11", None),
}

CHILD_SAMPLE_KEYS = {"montees", "otolithes"}

CSV_FILENAME_BY_SAMPLE_KEY = {
    "ecailles_brutes": "ecaille_brute.csv",
    "montees": "ecaille_montee.csv",
    "empreintes": "empreinte.csv",
    "opercules": "opercule.csv",
    "otolithes": "otolithe.csv",
    "vertebres": "vertebre.csv",
    "maxillaires": "maxillaire.csv",
    "chair_lyophilisee": "chair_lyophilisee.csv",
    "nageoires": "nageoire.csv",
    "muscle": "muscle.csv",
    "fraction_inconnue": "fraction_inconnue.csv",
}

HEADERS = [
    "sample_identifier", "collection_id", "sample_type_id", "sample_status_id",
    "country_code", "country_origin_code", "referent_id", "sampling_date",
    "sample_multiple_value", "sample_parent_identifier", "container_parent_identifier",
    "md_taxon", "md_longueur", "md_riviere", "md_num_individu",
]

# Dans le nouveau format COLISA à 40 colonnes, "Numero individu" est en colonne 19.
FALLBACK_NUM_INDIVIDU_COLUMN_INDEX = 18  # zero-based Excel column 19

COLISA_REQUIRED_HEADERS = {
    "code_echantillon": ["code echantillon"],
    "code_type_echantillon": ["code type echantillon"],
    "code_espece": ["code espece"],
    "pays_capture": ["pays capture"],
    "date_capture": ["date capture"],
    "lac_riviere": ["lac/riviere", "lac riviere"],
    "type_peche": ["type peche/engin", "type peche"],
    "num_individu": [
        "numero individu",
        "numero individue",
        "numero individu (numero de capture)",
        "numero individu numero de capture",
        "num individu",
    ],
    "longueur_mm": ["longueur totale (mm)", "longueur totale"],
    "age_total": ["age total"],
    "ecailles_brutes": ["ecailles brutes"],
    "montees": ["montees"],
    "empreintes": ["empreintes"],
    "otolithes": ["otolithes"],
    "opercules": ["opercules"],
}

CONTAINER_FIXED_TYPES = {
    "ecailles_brutes": "TIROIR",
    "montees": "BOITE",
    "otolithes": "BOITE",
}

TYPE_ECHANTILLON_TO_SAMPLE_KEY = [
    ("ecailles_brutes", ("eb", "brute", "brutes", "ec brute", "ec brutes", "ecaille brute", "ecailles brutes")),
    ("montees", ("ec", "em", "montee", "montees", "ec montee", "ec montees", "ecaille montee", "ecailles montees")),
    ("empreintes", ("ep", "empreinte", "empreintes")),
    ("otolithes", ("ot", "oto", "otolithe", "otolithes")),
    ("opercules", ("op", "ope", "opercule", "opercules")),
    ("vertebres", ("ver", "vertebre", "vertebres")),
    ("maxillaires", ("max", "maxillaire", "maxillaires")),
    ("chair_lyophilisee", ("chl", "chair lyophilisee", "lyophilisee")),
    ("nageoires", ("nag", "nageoire", "nageoires")),
    ("muscle", ("mu", "mus", "muscle")),
    ("fraction_inconnue", ("fn", "fraction inconnue")),
]

SAMPLE_CODE_TO_SAMPLE_KEY = {
    "AN": "nageoires",
    "DN": "nageoires",
    "NN": "nageoires",
    "PN": "nageoires",
    "QN": "nageoires",
    "OT": "otolithes",
    "ON": "opercules",
    "OP": "opercules",
    "OPE": "opercules",
    "VN": "vertebres",
    "VER": "vertebres",
    "MN": "maxillaires",
    "MAX": "maxillaires",
    "MU": "muscle",
    "MI": "muscle",
    "MP": "muscle",
    "FN": "fraction_inconnue",
    "EC": "montees",
    "EM": "montees",
    "EB": "ecailles_brutes",
}


def pays_to_code(pays_str: Any) -> str:
    if not pays_str:
        return ""
    value = str(pays_str).strip()
    code = COUNTRY_CODES.get(value.lower(), "")
    if code:
        return code
    if len(value) == 2:
        return value.upper()
    return value


def valeur_present(val: Any) -> bool:
    if val is None:
        return False
    if isinstance(val, bool):
        return bool(val)
    if isinstance(val, str):
        normalized = val.strip()
        if not normalized:
            return False
        if normalized.upper() in {"NON", "NO", "FALSE", "FAUX", "N", "0"}:
            return False
        if normalized.upper() in {"OUI", "YES", "TRUE", "VRAI", "Y", "X"}:
            return True
        normalized = normalized.replace(",", ".")
        try:
            return float(normalized) >= 1
        except (ValueError, TypeError):
            return True
    try:
        return float(val) >= 1
    except (ValueError, TypeError):
        return False


def infer_sample_key_from_type(code_type_echantillon: Any) -> Optional[str]:
    raw_text = normalize_text(code_type_echantillon).upper()
    if raw_text in SAMPLE_CODE_TO_SAMPLE_KEY:
        return SAMPLE_CODE_TO_SAMPLE_KEY[raw_text]

    normalized = normalize_header(code_type_echantillon)
    if not normalized:
        return None

    compact = normalized.replace(" ", "").upper()
    if compact in SAMPLE_CODE_TO_SAMPLE_KEY:
        return SAMPLE_CODE_TO_SAMPLE_KEY[compact]

    for sample_key, aliases in TYPE_ECHANTILLON_TO_SAMPLE_KEY:
        for alias in aliases:
            if normalize_header(alias) in normalized:
                return sample_key

    # Quand le type est bien renseigné mais ne correspond à aucune famille
    # Collect-Science connue, on exporte quand même l'échantillon dans la
    # catégorie générique au lieu de le perdre silencieusement.
    return None


def resolve_present_sample_keys_from_dict(data_row: Dict[str, Any]) -> Set[str]:
    explicit = {
        key
        for key in SAMPLE_TYPES
        if valeur_present(data_row.get(key))
    }
    if explicit:
        return explicit

    inferred = infer_sample_key_from_type(data_row.get("code_type_echantillon"))
    return {inferred} if inferred else set()


def resolve_present_sample_keys_from_excel_row(
    data_row: Tuple[Any, ...],
    col_map: Dict[str, int],
    code_type_echantillon: Any,
) -> Set[str]:
    explicit = {
        key
        for key in SAMPLE_TYPES
        if valeur_present(get_row_value(data_row, col_map, key))
    }
    if explicit:
        return explicit

    inferred = infer_sample_key_from_type(code_type_echantillon)
    return {inferred} if inferred else set()


def build_skip_reason(
    *,
    code_echantillon: Any,
    num_individu: Any,
    code_type_echantillon: Any,
    date_val: Any,
    espece: Any,
    present_sample_keys: Set[str],
) -> Optional[str]:
    if not normalize_text(code_echantillon) and not normalize_text(num_individu):
        return "numero individu / code echantillon manquant"
    if not normalize_text(date_val):
        return "date capture manquante"
    if not normalize_text(espece):
        return "code espece manquant"
    if not present_sample_keys:
        type_value = normalize_text(code_type_echantillon)
        if type_value:
            return (
                f"type echantillon '{type_value}' non reconnu pour Collect-Science. "
                "Choisissez manuellement le sample type (1, 2, 3, ...)."
            )
        return "aucun echantillon renseigne"
    return None


def append_skip_detail(
    details: List[str],
    *,
    row_index: int,
    code_echantillon: Any,
    num_individu: Any,
    reason: str,
) -> None:
    identifier = (
        normalize_text(code_echantillon)
        or normalize_text(num_individu)
        or f"ligne {row_index}"
    )
    details.append(f"{identifier}: {reason}")


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_sampling_date(value: Any) -> Any:
    parsed = DateCapture.from_excel(value)
    if parsed:
        return parsed.date
    return value


def build_code_echantillon_value(
    lac_riviere: Any,
    code_type_echantillon: Any,
    date_capture: Any,
    age_total: Any,
    numero_individu: Any,
    type_peche: Any = None,
) -> str:
    lac_part = normalize_text(lac_riviere)[:2]
    type_source = type_peche if normalize_text(type_peche) else code_type_echantillon
    type_part = normalize_text(type_source)[:1]
    date_value = normalize_sampling_date(date_capture)
    if isinstance(date_value, datetime.datetime):
        date_value = date_value.date()
    if isinstance(date_value, datetime.date):
        date_part = date_value.strftime("%d%m%Y")
    else:
        date_part = ""
    num_part = normalize_text(numero_individu)
    if not (lac_part or type_part or date_part or num_part):
        return ""
    return f"{lac_part}{type_part}{date_part}-{num_part}".upper()


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.replace("-", " ").replace("_", " ").replace("'", " ")
    return " ".join(text.split())


def build_colisa_column_map(worksheet) -> Dict[str, int]:
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    normalized_headers = [normalize_header(value) for value in header_row]
    col_map: Dict[str, int] = {}

    for key, aliases in COLISA_REQUIRED_HEADERS.items():
        normalized_aliases = [normalize_header(alias) for alias in aliases]
        for alias in normalized_aliases:
            for index, header in enumerate(normalized_headers):
                if header == alias:
                    col_map[key] = index
                    break
            if key in col_map:
                break

    if "num_individu" not in col_map and FALLBACK_NUM_INDIVIDU_COLUMN_INDEX < len(normalized_headers):
        col_map["num_individu"] = FALLBACK_NUM_INDIVIDU_COLUMN_INDEX

    return col_map


def get_row_value(data_row: Tuple[Any, ...], col_map: Dict[str, int], key: str) -> Any:
    index = col_map.get(key)
    if index is None or index < 0 or index >= len(data_row):
        return None
    return data_row[index]


def get_excel_num_individu_value(
    data_row: Tuple[Any, ...],
    col_map: Dict[str, int],
    prefer_fixed_column: bool = True,
    explicit_column_index: Optional[int] = None,
) -> Any:
    # Priorité 1 : colonne choisie explicitement par l'utilisateur dans le dialog
    if explicit_column_index is not None and 0 <= explicit_column_index < len(data_row):
        value = data_row[explicit_column_index]
        if normalize_text(value):
            return value
    # Priorité 2 : colonne fixe fallback (comportement historique)
    if prefer_fixed_column and FALLBACK_NUM_INDIVIDU_COLUMN_INDEX < len(data_row):
        fallback_value = data_row[FALLBACK_NUM_INDIVIDU_COLUMN_INDEX]
        if normalize_text(fallback_value):
            return fallback_value
    return get_row_value(data_row, col_map, "num_individu")


def ensure_visible_output_sheet(wb_out: Workbook) -> None:
    """
    Guarantee that the workbook always keeps at least one visible sheet.

    When no sample type is present in the selection, all generated sample sheets
    would otherwise be removed and openpyxl refuses to save the workbook.
    """
    if wb_out.sheetnames:
        return

    ws = wb_out.create_sheet("Collect-Science")
    for col_idx, header in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=header)


def normalize_container_label(value: Any) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    upper = text.upper()
    if upper.startswith("BOITE"):
        suffix = text[5:].strip()
        return "BOITE" if not suffix else f"BOITE {suffix}"
    if upper.startswith("TIROIR"):
        suffix = text[6:].strip()
        return "TIROIR" if not suffix else f"TIROIR {suffix}"
    return text


def parse_container_rules(raw_value: Optional[Any], sample_key: str) -> List[Tuple[int, int, str]]:
    if not raw_value:
        return []

    fixed_type = CONTAINER_FIXED_TYPES.get(sample_key)
    if isinstance(raw_value, dict):
        mode = str(raw_value.get("mode") or "single")
        container_type = normalize_container_label(raw_value.get("container_type") or fixed_type or "TIROIR")
        entries = raw_value.get("entries") or []
        rules: List[Tuple[int, int, str]] = []
        position = 1

        for entry in entries:
            label = normalize_container_label(entry.get("label"))
            if not label:
                continue
            quantity = entry.get("quantity")
            try:
                quantity_int = int(quantity) if quantity not in (None, "") else None
            except (TypeError, ValueError):
                quantity_int = None

            if fixed_type:
                if not label.lower().startswith(("tiroir", "boite")):
                    label = normalize_container_label(f"{fixed_type} {label}")
            elif not label.lower().startswith(("tiroir", "boite")):
                label = normalize_container_label(f"{container_type} {label}")
            else:
                label = normalize_container_label(label)

            if mode == "multiple" and quantity_int and quantity_int > 0:
                rules.append((position, position + quantity_int - 1, label))
                position += quantity_int
            else:
                rules.append((1, math.inf, label))
                break
        return rules

    rules: List[Tuple[int, int, str]] = []
    for raw_line in str(raw_value).splitlines():
        line = raw_line.strip()
        if not line:
            continue

        match = re.match(r"^(\d+)\s*-\s*(\d+)\s*=\s*(.+)$", line)
        if match:
            start = int(match.group(1))
            end = int(match.group(2))
            label = normalize_container_label(match.group(3).strip())
            if fixed_type and not label.lower().startswith(("tiroir", "boite")):
                label = normalize_container_label(f"{fixed_type} {label}")
            rules.append((start, end, label))
            continue

        label = normalize_container_label(line)
        if fixed_type and not label.lower().startswith(("tiroir", "boite")):
            label = normalize_container_label(f"{fixed_type} {label}")
        rules.append((1, math.inf, label))

    return rules


def resolve_container_value(rules: List[Tuple[int, int, str]], row_number: int) -> Optional[str]:
    for start, end, label in rules:
        if start <= row_number <= end:
            return label
    return None


def _format_csv_row(row: Tuple[Any, ...]) -> List[str]:
    formatted: List[str] = []
    for value in row:
        if value is None:
            formatted.append("")
        elif isinstance(value, float):
            if value.is_integer():
                formatted.append(str(int(value)))
            else:
                formatted.append(format(value, "f").rstrip("0").rstrip("."))
        elif isinstance(value, int):
            formatted.append(str(value))
        elif isinstance(value, datetime.datetime):
            formatted.append(value.strftime("%d/%m/%Y"))
        elif isinstance(value, datetime.date):
            formatted.append(value.strftime("%d/%m/%Y"))
        else:
            formatted.append(str(value))
    return formatted


def _write_collect_science_csv_files(output_path: Path, workbook) -> List[str]:
    output_dir = output_path.parent
    csv_files: List[str] = []

    for sample_key, (_, sheet_name, _) in SAMPLE_TYPES.items():
        if sheet_name not in workbook.sheetnames:
            continue

        ws = workbook[sheet_name]
        data_rows: List[List[str]] = []
        for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_index == 1:
                continue
            formatted = _format_csv_row(row)
            if any(value.strip() for value in formatted):
                data_rows.append(formatted)

        if not data_rows:
            continue

        csv_name = CSV_FILENAME_BY_SAMPLE_KEY.get(sample_key, f"{sample_key}.csv")
        csv_path = output_dir / csv_name
        with open(csv_path, "w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle, delimiter=";")
            writer.writerow(HEADERS)
            writer.writerows(data_rows)
        csv_files.append(csv_path.name)

    return csv_files


def _sample_multiple_value_for_key(sample_key: str, default_value: int) -> Optional[int]:
    if sample_key == "ecailles_brutes":
        return default_value
    return None


def _resolve_md_num_individu(num_individu: Any, code_echantillon: Any, t_code: str) -> Optional[str]:
    value = normalize_text(num_individu) or normalize_text(code_echantillon) or normalize_text(t_code)
    return value or None


def _resolve_md_num_individu_from_excel(num_individu: Any, code_echantillon: Any, t_code: str) -> Optional[str]:
    value = normalize_text(num_individu)
    return value or None


def build_md_num_individu_value(
    lac_riviere: Any,
    code_type_echantillon: Any,
    date_capture: Any,
    numero_individu: Any,
    type_peche: Any = None,
    code_echantillon: Any = None,
    t_code: str = "",
    prefer_existing: bool = True,
) -> Optional[str]:
    existing = normalize_text(numero_individu)
    if prefer_existing and existing:
        return existing.upper()

    built = build_numero_identification_value(
        lac_riviere=lac_riviere,
        code_type_echantillon=code_type_echantillon,
        date_capture=date_capture,
        numero_individu=numero_individu,
        type_peche=type_peche,
    )
    if built:
        return built

    if prefer_existing:
        return _resolve_md_num_individu(existing, code_echantillon, t_code)
    return _resolve_md_num_individu(code_echantillon, code_echantillon, t_code)


def read_sample_counts_from_colisa(
    colisa_path: Path,
    colisa_sheet: str = "Feuil1 ",
    allowed_num_individus: Optional[Set[str]] = None,
) -> Dict[str, int]:
    allowed_values = {normalize_text(v) for v in (allowed_num_individus or set()) if normalize_text(v)}
    wb_src = load_workbook(str(colisa_path), read_only=True, data_only=True)
    try:
        sheet_name = _resolve_sheet_name(wb_src, colisa_sheet)
        ws_src = wb_src[sheet_name]
        col_map = build_colisa_column_map(ws_src)

        counts: Dict[str, int] = {key: 0 for key in ("ecailles_brutes", "montees", "empreintes", "otolithes")}
        for data_row in ws_src.iter_rows(min_row=2, values_only=True):
            num_individu = normalize_text(get_row_value(data_row, col_map, "num_individu"))
            if allowed_values and num_individu not in allowed_values:
                continue

            for key in counts:
                if valeur_present(get_row_value(data_row, col_map, key)):
                    counts[key] += 1

        return {key: value for key, value in counts.items() if value > 0}
    finally:
        wb_src.close()


def generer_collec_science(
    colisa_path: Path,
    output_path: Path,
    collection_id: int = 1,
    sample_status_id: int = 1,
    referent_id: int = 1,
    sample_multiple_value: int = 5,
    containers: Dict[str, Optional[str]] = None,
    forcer_anomalies: bool = False,
    colisa_sheet: str = "Feuil1 ",
    allowed_num_individus: Optional[Set[str]] = None,
    prefer_fixed_num_individu_column: bool = True,
    md_num_individu_column_index: Optional[int] = None,
) -> Dict[str, Any]:
    containers = containers or {}
    allowed_values = {normalize_text(v) for v in (allowed_num_individus or set()) if normalize_text(v)}
    container_rules = {key: parse_container_rules(value, key) for key, value in containers.items()}

    wb_src = load_workbook(str(colisa_path), read_only=True, data_only=True)
    sheet_name = _resolve_sheet_name(wb_src, colisa_sheet)
    ws_src = wb_src[sheet_name]
    col_map = build_colisa_column_map(ws_src)
    all_rows = list(ws_src.iter_rows(min_row=2, values_only=True))
    wb_src.close()

    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    sheets: Dict[str, Any] = {}
    for key, (type_id, sheet_label, suffix) in SAMPLE_TYPES.items():
        ws = wb_out.create_sheet(sheet_label)
        sheets[key] = ws
        for col_idx, header in enumerate(HEADERS, start=1):
            ws.cell(row=1, column=col_idx, value=header)

    row_counters: Dict[str, int] = {key: 2 for key in SAMPLE_TYPES}
    sample_positions: Dict[str, int] = {key: 0 for key in SAMPLE_TYPES}
    rows_written = 0
    skipped_details: List[str] = []

    for row_index, data_row in enumerate(all_rows, start=2):
        if not data_row:
            continue

        code_echantillon = get_row_value(data_row, col_map, "code_echantillon")
        code_type_echantillon = get_row_value(data_row, col_map, "code_type_echantillon")
        espece = get_row_value(data_row, col_map, "code_espece") or ""
        pays_texte = get_row_value(data_row, col_map, "pays_capture") or ""
        country_code = pays_to_code(pays_texte)
        date_val = get_row_value(data_row, col_map, "date_capture")
        lac_riviere = get_row_value(data_row, col_map, "lac_riviere") or ""
        type_peche = get_row_value(data_row, col_map, "type_peche") or ""
        longueur = get_row_value(data_row, col_map, "longueur_mm")
        num_individu = get_excel_num_individu_value(
            data_row,
            col_map,
            prefer_fixed_column=prefer_fixed_num_individu_column,
            explicit_column_index=md_num_individu_column_index,
        ) or ""
        age_total = get_row_value(data_row, col_map, "age_total")
        t_code = normalize_text(code_echantillon) or build_code_echantillon_value(
            lac_riviere=lac_riviere,
            code_type_echantillon=code_type_echantillon,
            date_capture=date_val,
            age_total=age_total,
            numero_individu=num_individu,
            type_peche=type_peche,
        )
        if not t_code:
            append_skip_detail(
                skipped_details,
                row_index=row_index,
                code_echantillon=code_echantillon,
                num_individu=num_individu,
                reason="code echantillon impossible a construire",
            )
            continue
        num_individu_norm = t_code

        if allowed_values and num_individu_norm not in allowed_values:
            continue

        present_sample_keys = resolve_present_sample_keys_from_excel_row(
            data_row,
            col_map,
            code_type_echantillon,
        )
        date_val = normalize_sampling_date(date_val)
        skip_reason = None
        if not forcer_anomalies:
            skip_reason = build_skip_reason(
                code_echantillon=code_echantillon,
                num_individu=num_individu,
                code_type_echantillon=code_type_echantillon,
                date_val=date_val,
                espece=espece,
                present_sample_keys=present_sample_keys,
            )
        if skip_reason:
            append_skip_detail(
                skipped_details,
                row_index=row_index,
                code_echantillon=code_echantillon,
                num_individu=num_individu,
                reason=skip_reason,
            )
            continue

        ecaille_brute_present = "ecailles_brutes" in present_sample_keys
        boss_identifier = t_code if ecaille_brute_present else None

        for key, (type_id, sheet_label, suffix) in SAMPLE_TYPES.items():
            ws = sheets[key]
            current_row = row_counters[key]

            present = key in present_sample_keys

            if not present:
                continue

            sample_positions[key] += 1
            parent_id = boss_identifier if key in CHILD_SAMPLE_KEYS else None
            sample_id = t_code if suffix is None or parent_id is None else f"{t_code}{suffix}"
            preview_row_number = row_index - 1
            container_val = resolve_container_value(container_rules.get(key, []), preview_row_number)
            sample_multiple_value_cell = _sample_multiple_value_for_key(key, sample_multiple_value)
            md_num_individu = build_md_num_individu_value(
                lac_riviere=lac_riviere,
                code_type_echantillon=code_type_echantillon,
                date_capture=date_val,
                numero_individu=num_individu,
                type_peche=type_peche,
                code_echantillon=code_echantillon,
                t_code=t_code,
            )

            ligne = [
                sample_id, collection_id, type_id, sample_status_id,
                country_code, country_code, referent_id, date_val,
                sample_multiple_value_cell, parent_id, container_val,
                str(espece) if espece else None,
                coerce_numeric_string(longueur),
                str(lac_riviere) if lac_riviere else None,
                md_num_individu,
            ]

            for col_idx, val in enumerate(ligne, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                if col_idx == 8 and isinstance(val, (datetime.datetime, datetime.date)):
                    cell.number_format = "DD/MM/YYYY"
            row_counters[key] += 1
            rows_written += 1

    to_delete = [SAMPLE_TYPES[key][1] for key in SAMPLE_TYPES if row_counters[key] == 2]
    for title in to_delete:
        if title in wb_out.sheetnames:
            del wb_out[title]

    ensure_visible_output_sheet(wb_out)

    wb_out.save(str(output_path))
    csv_files = _write_collect_science_csv_files(output_path, wb_out)
    return {
        "excel": str(output_path),
        "csv_files": csv_files,
        "rows_written": rows_written,
        "skipped_details": skipped_details,
    }


def generer_collec_science_depuis_rows(
    rows: List[Dict[str, Any]],
    output_path: Path,
    collection_id: int = 1,
    sample_status_id: int = 1,
    referent_id: int = 1,
    sample_multiple_value: int = 5,
    containers: Dict[str, Optional[str]] = None,
    forcer_anomalies: bool = False,
) -> Dict[str, Any]:
    """
    Génère le fichier Collec-Science (Excel + CSV) directement depuis les rows
    du logiciel (List[Dict]), sans passer par aucun fichier COLISA.
    Utilisé pour le mode logiciel ET le mode 'depuis un fichier Excel source'.
    """
    containers = containers or {}
    container_rules = {key: parse_container_rules(value, key) for key, value in containers.items()}

    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    sheets: Dict[str, Any] = {}
    for key, (type_id, sheet_label, suffix) in SAMPLE_TYPES.items():
        ws = wb_out.create_sheet(sheet_label)
        sheets[key] = ws
        for col_idx, header in enumerate(HEADERS, start=1):
            ws.cell(row=1, column=col_idx, value=header)

    row_counters: Dict[str, int] = {key: 2 for key in SAMPLE_TYPES}
    sample_positions: Dict[str, int] = {key: 0 for key in SAMPLE_TYPES}
    rows_written = 0
    skipped_details: List[str] = []

    for row_index, data_row in enumerate(rows, start=1):
        if not data_row:
            continue

        code_echantillon = data_row.get("code_echantillon")
        code_type_echantillon = data_row.get("code_type_echantillon")
        espece = data_row.get("code_espece") or ""
        pays_texte = data_row.get("pays_capture") or ""
        country_code = pays_to_code(pays_texte)
        date_val = data_row.get("date_capture")
        lac_riviere = data_row.get("lac_riviere") or ""
        type_peche = data_row.get("type_peche") or ""
        longueur = data_row.get("longueur_mm")
        num_individu = data_row.get("num_individu") or ""
        age_total = data_row.get("age_total")

        t_code = normalize_text(code_echantillon) or build_code_echantillon_value(
            lac_riviere=lac_riviere,
            code_type_echantillon=code_type_echantillon,
            date_capture=date_val,
            age_total=age_total,
            numero_individu=num_individu,
            type_peche=type_peche,
        )
        if not t_code:
            append_skip_detail(
                skipped_details,
                row_index=row_index,
                code_echantillon=code_echantillon,
                num_individu=num_individu,
                reason="code echantillon impossible a construire",
            )
            continue

        num_individu_norm = t_code

        present_sample_keys = resolve_present_sample_keys_from_dict(data_row)
        date_val = normalize_sampling_date(date_val)
        skip_reason = None
        if not forcer_anomalies:
            skip_reason = build_skip_reason(
                code_echantillon=code_echantillon,
                num_individu=num_individu,
                code_type_echantillon=code_type_echantillon,
                date_val=date_val,
                espece=espece,
                present_sample_keys=present_sample_keys,
            )
        if skip_reason:
            append_skip_detail(
                skipped_details,
                row_index=row_index,
                code_echantillon=code_echantillon,
                num_individu=num_individu,
                reason=skip_reason,
            )
            continue

        ecaille_brute_present = "ecailles_brutes" in present_sample_keys
        boss_identifier = t_code if ecaille_brute_present else None

        for key, (type_id, sheet_label, suffix) in SAMPLE_TYPES.items():
            ws = sheets[key]
            current_row = row_counters[key]

            present = key in present_sample_keys
            if not present:
                continue

            sample_positions[key] += 1
            parent_id = boss_identifier if key in CHILD_SAMPLE_KEYS else None
            sample_id = t_code if suffix is None or parent_id is None else f"{t_code}{suffix}"
            preview_row_number = row_index
            container_val = resolve_container_value(container_rules.get(key, []), preview_row_number)
            sample_multiple_value_cell = _sample_multiple_value_for_key(key, sample_multiple_value)
            md_num_individu = build_md_num_individu_value(
                lac_riviere=lac_riviere,
                code_type_echantillon=code_type_echantillon,
                date_capture=date_val,
                numero_individu=num_individu,
                type_peche=type_peche,
                code_echantillon=code_echantillon,
                t_code=t_code,
            )

            ligne = [
                sample_id, collection_id, type_id, sample_status_id,
                country_code, country_code, referent_id, date_val,
                sample_multiple_value_cell, parent_id, container_val,
                str(espece) if espece else None,
                coerce_numeric_string(longueur),
                str(lac_riviere) if lac_riviere else None,
                md_num_individu,
            ]

            for col_idx, val in enumerate(ligne, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                if col_idx == 8 and isinstance(val, (datetime.datetime, datetime.date)):
                    cell.number_format = "DD/MM/YYYY"
            row_counters[key] += 1
            rows_written += 1

    # Supprimer les feuilles vides
    to_delete = [SAMPLE_TYPES[key][1] for key in SAMPLE_TYPES if row_counters[key] == 2]
    for title in to_delete:
        if title in wb_out.sheetnames:
            del wb_out[title]

    ensure_visible_output_sheet(wb_out)

    wb_out.save(str(output_path))

    # Générer les CSV dans le même dossier
    csv_files = _write_collect_science_csv_files(output_path, wb_out)
    return {
        "excel": str(output_path),
        "csv_files": csv_files,
        "rows_written": rows_written,
        "skipped_details": skipped_details,
    }
def _resolve_sheet_name(workbook, colisa_sheet: str) -> str:
    if colisa_sheet in workbook.sheetnames:
        return colisa_sheet
    for sheet_name in workbook.sheetnames:
        if sheet_name.strip().lower() == colisa_sheet.strip().lower():
            return sheet_name
    if workbook.sheetnames:
        return workbook.sheetnames[0]
    raise ValueError("Le fichier Excel ne contient aucune feuille exploitable.")


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 3:
        print("Usage: python generer_collec_science.py <colisa.xlsx> <output.xlsx>")
        sys.exit(1)

    generer_collec_science(
        colisa_path=Path(sys.argv[1]),
        output_path=Path(sys.argv[2]),
    )
