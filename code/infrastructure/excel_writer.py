"""Excel writing operations using openpyxl library."""

from __future__ import annotations

import datetime as dt
import re
from pathlib import Path
from typing import Any, Dict, List, Optional

from config.constants import DEFAULT_CODE_WIDTH


class ExcelWriter:
    """
    Handles writing and modifying Excel files using openpyxl library.

    This class isolates all Excel writing operations, including
    formula propagation, context copying, and workbook management.
    """

    def __init__(self, openpyxl_module):
        """
        Initialize ExcelWriter with openpyxl module.

        Args:
            openpyxl_module: The openpyxl module (injected for testability)
        """
        self.openpyxl = openpyxl_module

    def load_workbook(self, file_path: Path, read_only: bool = False, data_only: bool = False):
        """
        Load an Excel workbook.

        Args:
            file_path: Path to the .xlsx file
            read_only: Whether to open in read-only mode
            data_only: Whether to read only data (not formulas)

        Returns:
            Workbook object

        Raises:
            IOError: If file cannot be opened
        """
        try:
            return self.openpyxl.load_workbook(
                file_path,
                read_only=read_only,
                data_only=data_only
            )
        except Exception as e:
            raise IOError(f"Cannot open workbook {file_path}: {e}") from e

    def save_workbook(self, workbook, file_path: Path) -> None:
        """
        Save a workbook to file.

        Args:
            workbook: Workbook object to save
            file_path: Path where to save the file

        Raises:
            IOError: If file cannot be saved
        """
        try:
            # Ensure parent directory exists
            file_path.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(file_path)
        except PermissionError as e:
            raise IOError(
                f"\u274c Impossible de sauvegarder le fichier :\n{file_path.name}\n\n"
                f"Le fichier est ouvert dans Microsoft Excel.\n"
                f"Ferme-le dans Excel puis relance l'import."
            ) from e
        except Exception as e:
            raise IOError(f"Cannot save workbook to {file_path}: {e}") from e

    def propagate_formula_from_previous_row(
        self,
        ws,
        target_row: int,
        col_index: Optional[int],
        min_row: int = 2
    ) -> None:
        """
        Copy formula from previous row and adjust references.

        This looks for the nearest row above with a formula in the same column
        and copies it to the target row, adjusting cell references.

        Args:
            ws: Worksheet object
            target_row: Row to write formula to
            col_index: Column index (1-based) or None to skip
            min_row: Minimum row to search (typically header_row + 1)
        """
        if not col_index or target_row <= min_row:
            return

        current_cell = ws.cell(target_row, col_index)
        if current_cell.value not in (None, ""):
            return  # Cell already has a value

        # Search for source formula in previous rows
        source_row = None
        source_formula = None
        for r in range(target_row - 1, min_row - 1, -1):
            candidate = ws.cell(r, col_index).value
            if isinstance(candidate, str) and candidate.startswith("="):
                source_row = r
                source_formula = candidate
                break

        if source_row is None or source_formula is None:
            return

        # Try to translate formula with openpyxl Translator
        try:
            from openpyxl.formula.translate import Translator
            from openpyxl.utils import get_column_letter

            col_letter = get_column_letter(col_index)
            origin = f"{col_letter}{source_row}"
            dest = f"{col_letter}{target_row}"
            current_cell.value = Translator(source_formula, origin=origin).translate_formula(dest)
        except Exception:
            # Fallback for localized Excel formulas: update row references manually.
            row_pattern = re.compile(rf"(?<![A-Z])([A-Z]{{1,3}}){source_row}(?!\d)")
            current_cell.value = row_pattern.sub(rf"\1{target_row}", source_formula)

    def propagate_all_formulas(
        self,
        ws,
        target_row: int,
        min_row: int = 2
    ) -> None:
        """
        Propagate all formulas from previous rows for the entire row.

        This scans all columns and propagates any formulas found.

        Args:
            ws: Worksheet object
            target_row: Row to write formulas to
            min_row: Minimum row to search
        """
        for col in range(1, ws.max_column + 1):
            self.propagate_formula_from_previous_row(ws, target_row, col, min_row)

    def propagate_incremental_code(
        self,
        ws,
        target_row: int,
        col_index: Optional[int],
        min_row: int = 2
    ) -> None:
        """
        Generate incremental code from previous row.

        If previous row has a code like "T00042", generates "T00043".
        This is a fallback for when formula propagation doesn't work
        (e.g., data_only workbooks).

        Args:
            ws: Worksheet object
            target_row: Row to write code to
            col_index: Column index (1-based) or None to skip
            min_row: Minimum row to search
        """
        if not col_index or target_row <= min_row:
            return

        cell = ws.cell(target_row, col_index)
        if cell.value not in (None, ""):
            return  # Cell already has a value

        pattern = re.compile(r"^(.*?)(\d+)$")
        for r in range(target_row - 1, min_row - 1, -1):
            prev_val = self._normalize(ws.cell(r, col_index).value)
            if not prev_val:
                continue
            m = pattern.match(prev_val)
            if not m:
                continue
            prefix, num = m.group(1), m.group(2)
            next_num = str(int(num) + 1).zfill(len(num))
            cell.value = f"{prefix}{next_num}"
            return

    def assign_next_code_if_missing(
        self,
        ws,
        target_row: int,
        code_col: Optional[int],
        seq_state: Dict[str, Any]
    ) -> None:
        """
        Assign next sequential code if cell is empty.

        Uses a sequence state dict to track the current code number.

        Args:
            ws: Worksheet object
            target_row: Row to write code to
            code_col: Column index (1-based) or None to skip
            seq_state: Dict with 'prefix', 'num', 'width' keys
        """
        if not code_col:
            return

        cell = ws.cell(target_row, code_col)
        if self._normalize(cell.value):
            return  # Cell already has a value

        seq_state["num"] += 1
        prefix = seq_state.get("prefix", "T") or "T"
        width = int(seq_state.get("width", DEFAULT_CODE_WIDTH))
        cell.value = f"{prefix}{str(seq_state['num']).zfill(width)}"

    def copy_context_fields(
        self,
        ws,
        target_row: int,
        header_map: Dict[str, int],
        source_rows: Optional[List[int]] = None,
        min_row: int = 2,
        expected_type: str = "",
        expected_ref_prefix: str = ""
    ) -> None:
        """
        Copy context fields from previous rows.

        Context fields are metadata that applies across multiple samples
        (e.g., site, gestionnaire) and should be inherited when empty.

        Args:
            ws: Worksheet object
            target_row: Row to copy values to
            header_map: Dict mapping field names to column indices
            source_rows: Optional list of rows to search (in priority order)
            min_row: Minimum row to search
            expected_type: Only copy from rows with this type
            expected_ref_prefix: Only copy from rows with this REF prefix
        """
        # Context fields to copy (structural metadata only)
        context_keys = [
            "code_unite_gestionnaire",
            "site_atelier",
            "numero_correspondant",
            "lac_riviere",
            "categorie",
            "type_peche",
            "autre_oss",
            "observation_disponibilite"
        ]

        type_col = header_map.get("code_type_echantillon")
        num_col = header_map.get("num_individu")

        # Determine search order
        if source_rows:
            candidates = [r for r in source_rows if min_row <= r < target_row]
            candidates = sorted(candidates, reverse=True)
        else:
            candidates = list(range(target_row - 1, min_row - 1, -1))

        # Find suitable source row
        source_row = None
        for r in candidates:
            # Check type matches if specified
            if type_col and expected_type:
                row_type = self._normalize(ws.cell(r, type_col).value)
                if row_type != self._normalize(expected_type):
                    continue

            # Check REF prefix matches if specified
            if num_col and expected_ref_prefix:
                num_val = self._normalize(ws.cell(r, num_col).value)
                from domain.value_objects import RefCode  # Will be created in Phase 3
                parts = self._parse_ref_parts_simple(num_val)
                prefix = parts[0] if parts else ""
                if prefix != expected_ref_prefix:
                    continue

            # Check if this row has any context data
            has_context = False
            for key in context_keys:
                col = header_map.get(key)
                if col and self._normalize(ws.cell(r, col).value):
                    has_context = True
                    break

            if has_context:
                source_row = r
                break

        if source_row is None:
            return

        # Copy context values
        for key in context_keys:
            col = header_map.get(key)
            if not col:
                continue
            current = self._normalize(ws.cell(target_row, col).value)
            if current:
                continue  # Don't overwrite existing value
            prev = ws.cell(source_row, col).value
            if self._normalize(prev):
                ws.cell(target_row, col).value = prev

    def set_cell_format(self, ws, row: int, col: int, number_format: str) -> None:
        """
        Set number format for a cell.

        Args:
            ws: Worksheet object
            row: Row index (1-based)
            col: Column index (1-based)
            number_format: Format string (e.g., "dd/mm/yy", "@")
        """
        ws.cell(row, col).number_format = number_format

    @staticmethod
    def _normalize(value: Any) -> str:
        """Normalize a cell value to string."""
        if value is None:
            return ""
        if isinstance(value, float):
            if value.is_integer():
                return str(int(value))
            return format(value, "f").rstrip("0").rstrip(".")
        return str(value).strip()

    @staticmethod
    def _parse_ref_parts_simple(code: str) -> Optional[tuple]:
        """Simple REF parsing (will be replaced by RefCode value object)."""
        code = code.strip().upper()
        m = re.match(r"^([A-Z]+)\s*0*(\d+)$", code)
        if not m:
            return None
        return m.group(1), int(m.group(2))
