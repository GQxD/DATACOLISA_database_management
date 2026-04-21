"""Generate the built-in COLISA workbook used by the application."""

from __future__ import annotations

import unicodedata
import datetime as dt
from pathlib import Path
from typing import Tuple

from config.constants import DEFAULT_TARGET_SHEET


TOTAL_COLUMNS = 41

# En-tetes codes en dur — structure identique aux fichiers COLISA de reference
HEADER_POSITIONS = {
    1:  "Code unit\u00e9 gestionnaire",
    2:  "Site Atelier",
    3:  "Numero du correspondant",
    4:  "Code type echantillon",
    5:  "Code \u00e9chantillon",
    6:  "Code esp\u00e8ce",
    7:  "Sous-esp\u00e8ce ",
    8:  "Organisme pr\u00e9leveur",
    9:  "Nom de l'op\u00e9rateur",
    10: "Pays capture ",
    11: "Date capture",
    12: "Lac/riviere",
    13: "Lieu de capture / debarquement",
    14: "Cat\u00e9gorie p\u00eacheur ",
    15: "Type p\u00eache/engin ",
    16: "Maille (mm)",
    # 17 : colonne vide (r\u00e9serv\u00e9e)
    18: "CODE IDENTIFICATION",
    19: "Numero individu (numero de capture)",
    20: "Longueur totale (mm)",
    21: "Poids (g)",
    22: "Code stade",
    23: "Code maturit\u00e9 sexuelle",
    24: "Generer pour site colisa",
    25: "Code sexe",
    26: "Pr\u00e9sence de l'otolithe gauche (0 si non, 1 si oui)",
    27: "Pr\u00e9sence de l'otolithe droite (0 si non, 1 si oui)",
    28: "Nombre d' opercules en \u00e9tat",
    29: "Information stockage ",
    30: "Observation disponibilit\u00e9",
    31: "Autre \u00e9chantillon osseuses collect\u00e9e sur l'individu OUI/NON",
    32: "Age total",
    33: "Age rivi\u00e8re ",
    34: "Age lac",
    35: "Nombre de fraie",
    36: "Ecailles regen\u00e9r\u00e9es ? (0 si non, 1 si oui)",
    37: "Observations",
    38: "Ecailles brutes",
    39: "Mont\u00e9es",
    40: "Empreintes",
    41: "Otolithes",
}

# Donnees de la feuille "Type echantillon" (codes officiels COLISA)
TYPE_ECHANTILLON_DATA = [
    ("AN", "Abdomen de poisson", 291),
    ("BI", "Bile", 233),
    ("GN", "Bouche de poisson", 293),
    ("BN", "Branchie de poisson", 10),
    ("VN", "Colonne vert\u00e9brale de poisson", 306),
    ("HN", "Dos de Poisson", 294),
    ("EC", "Ecaille de poisson", 103),
    ("ES", "Estomac", None),
    ("FO", "Foie de poisson", 104),
    ("FN", "Fraction inconnue de poisson", 100),
    ("GO", "Gonade de poisson", 245),
    ("GR", "Graisse de poisson", 107),
    ("HM", "H\u00e9mocytes de poissons", 241),
    ("LE", "L\u00e8vre de poisson", 295),
    ("MN", "M\u00e2choire de poisson", 296),
    ("MU", "Muscle de poisson", 102),
    ("MP", "Muscle et peau de poisson", 272),
    ("MI", "Muscle, muscle+tissu adipeux face interne de peau", 164),
    ("QN", "Nageoire caudale de poisson", 299),
    ("NN", "Nageoire de poisson", 297),
    ("DN", "Nageoire dorsale de poisson", 292),
    ("PN", "Nageoire pectorale de poisson", 305),
    ("YN", "Oeil de poisson", 302),
    ("ON", "Opercules", 298),
    ("XN", "Orifice anal de poisson", 308),
    ("XU", "Orifice urog\u00e9nital de poisson", 301),
    ("OT", "Otolithes", None),
    ("KN", "P\u00e9doncule caudal de poisson", 304),
    ("CN", "Poisson entier", 101),
    ("WE", "Poisson etete et equeute", 155),
    ("WV", "Poisson sans visc\u00e8res ni gonades", 243),
    ("RE", "Rein de poisson", 105),
    ("NF", "Syst\u00e8me nerveux de poisson", 106),
    ("TN", "T\u00eate de poisson", 300),
    ("WN", "Tronc de poisson", 307),
]


def build_numero_identification_value(
    lac_riviere: object,
    code_type_echantillon: object,
    date_capture: object,
    numero_individu: object,
    type_peche: object = None,
) -> str:
    """Build the Numéro d'identification value: {LAC2}{TYPE1}{DDMMYYYY}-{NUMERO}."""
    lac_part = str(lac_riviere or "").strip()[:2]
    type_source = type_peche if str(type_peche or "").strip() else code_type_echantillon
    type_part = str(type_source or "").strip()[:1]
    num_part = str(numero_individu or "").strip()

    date_part = ""
    if isinstance(date_capture, dt.datetime):
        date_capture = date_capture.date()
    if isinstance(date_capture, dt.date):
        date_part = date_capture.strftime("%d%m%Y")

    if not (lac_part or type_part or date_part or num_part):
        return ""

    return f"{lac_part}{type_part}{date_part}-{num_part}".upper()


def build_code_echantillon_value(
    lac_riviere: object,
    code_type_echantillon: object,
    date_capture: object,
    age_total: object,
    numero_individu: object,
    type_peche: object = None,
    force_prefix: str = "T",
) -> str:
    """Build the Code echantillon : {PREFIXE}{NUMERO}.
    Le prefixe est force a 'T' par defaut (configurable via force_prefix).
    """
    num_part = str(numero_individu or "").strip()
    if not num_part:
        return ""
    prefix = force_prefix.strip().upper() if force_prefix and str(force_prefix).strip() else "T"
    return f"{prefix}{num_part}"


def create_internal_target_workbook(output_path: Path, openpyxl_module, template_path: Path | None = None) -> Path:
    """Create the built-in COLISA workbook used as import base."""
    if template_path and template_path.exists():
        return create_target_workbook_from_template(output_path, openpyxl_module, template_path)

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill

    workbook = openpyxl_module.Workbook()
    worksheet = workbook.active
    worksheet.title = DEFAULT_TARGET_SHEET

    # ── Colonnes format nombre et date ────────────────────────────────────────
    # Colonnes format nombre entier : Code unite (1), Num correspondant (3),
    # Ecailles brutes (38), Montees (39), Empreintes (40), Otolithes (41)
    _NUMBER_FORMAT_COLS = {1, 3, 38, 39, 40, 41}
    # Colonne date capture (11)
    _DATE_FORMAT_COLS = {11}

    # ── Style en-tete (identique aux fichiers COLISA de reference) ────────────
    header_font = Font(name="Times New Roman", bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_border = Border(
        bottom=Side(border_style="medium"),
    )

    # ── Ecriture des en-tetes et style ────────────────────────────────────────
    for col_index in range(1, TOTAL_COLUMNS + 1):
        header = HEADER_POSITIONS.get(col_index, "")
        cell = worksheet.cell(1, col_index)
        cell.value = header if header else None
        cell.font = header_font
        cell.alignment = header_align
        cell.border = header_border

    worksheet.row_dimensions[1].height = 73.5
    worksheet.freeze_panes = "A2"

    # ── Largeur colonnes + formats pre-appliques sur 2000 lignes ──────────────
    PRE_FORMAT_ROWS = 2000
    for col_index in range(1, TOTAL_COLUMNS + 1):
        header = HEADER_POSITIONS.get(col_index, "")
        width = max(14, min(len(header) + 4, 42)) if header else 8
        letter = _column_letter(col_index)
        worksheet.column_dimensions[letter].width = width

        if col_index in _NUMBER_FORMAT_COLS:
            fmt = "0"
        elif col_index in _DATE_FORMAT_COLS:
            fmt = "DD/MM/YYYY"
        else:
            continue

        for row_index in range(2, PRE_FORMAT_ROWS + 2):
            worksheet.cell(row_index, col_index).number_format = fmt

    # ── Feuil2 (vide) ──────────────────────────────────────────────────────────
    workbook.create_sheet("Feuil2")

    # ── Type echantillon ───────────────────────────────────────────────────────
    type_sheet = workbook.create_sheet("Type echantillon")
    type_sheet.cell(1, 1).value = "Code type echantillon"
    type_sheet.cell(1, 2).value = "Description"
    type_sheet.cell(1, 3).value = "code sandre"
    for row_index, (code, desc, sandre) in enumerate(TYPE_ECHANTILLON_DATA, start=2):
        type_sheet.cell(row_index, 1).value = code
        type_sheet.cell(row_index, 2).value = desc
        type_sheet.cell(row_index, 3).value = sandre
    type_sheet.column_dimensions["A"].width = 10
    type_sheet.column_dimensions["B"].width = 52
    type_sheet.column_dimensions["C"].width = 14

    # ── Especes ────────────────────────────────────────────────────────────────
    esp_sheet = workbook.create_sheet("Esp\u00e8ces")
    for col, hdr in enumerate(["Code esp\u00e8ce", "libell\u00e9", "code SANDRE", "code TAXREF"], start=1):
        esp_sheet.cell(1, col).value = hdr
    esp_sheet.column_dimensions["A"].width = 14
    esp_sheet.column_dimensions["B"].width = 30

    # ── Type peche ─────────────────────────────────────────────────────────────
    tp_sheet = workbook.create_sheet("Type p\u00eache ")
    tp_sheet.cell(1, 1).value = "Ligne "
    tp_sheet.column_dimensions["A"].width = 20

    # ── Stade ──────────────────────────────────────────────────────────────────
    stade_sheet = workbook.create_sheet("Stade")
    for col, hdr in enumerate(["CodeStade", "DescriptionCodeStade", "CodeSandre"], start=1):
        stade_sheet.cell(1, col).value = hdr
    stade_sheet.column_dimensions["A"].width = 12
    stade_sheet.column_dimensions["B"].width = 30

    # ── Sous espece ────────────────────────────────────────────────────────────
    sous_sheet = workbook.create_sheet("Sous esp\u00e8ce")
    sous_sheet.cell(1, 1).value = "Pal\u00e9e"
    sous_sheet.column_dimensions["A"].width = 20

    # ── Maturite sexuelle ──────────────────────────────────────────────────────
    mat_sheet = workbook.create_sheet("Maturit\u00e9 sexuelle")
    mat_sheet.cell(1, 1).value = "Code maturit\u00e9 sexuelle"
    mat_sheet.cell(1, 2).value = "Libell\u00e9"
    mat_sheet.column_dimensions["A"].width = 24
    mat_sheet.column_dimensions["B"].width = 30

    # ── Sexe ───────────────────────────────────────────────────────────────────
    sexe_sheet = workbook.create_sheet("Sexe")
    sexe_sheet.cell(1, 1).value = "Code sexe"
    sexe_sheet.cell(1, 2).value = "Libell\u00e9"
    sexe_sheet.column_dimensions["A"].width = 12
    sexe_sheet.column_dimensions["B"].width = 20

    # ── Type engin technique ───────────────────────────────────────────────────
    engin_sheet = workbook.create_sheet("Type engin technique ")
    engin_sheet.cell(1, 1).value = "Pics "
    engin_sheet.column_dimensions["A"].width = 20

    # ── type peche engins ──────────────────────────────────────────────────────
    pe_sheet = workbook.create_sheet("type p\u00eache engins")
    pe_sheet.cell(1, 1).value = "Type de p\u00eache /engins"
    pe_sheet.column_dimensions["A"].width = 24

    # ── Categorie de pecheurs ──────────────────────────────────────────────────
    cat_sheet = workbook.create_sheet("Categorie de pecheurs")
    cat_sheet.cell(1, 1).value = "Amateur "
    cat_sheet.column_dimensions["A"].width = 20

    # ── Sites Atelier ──────────────────────────────────────────────────────────
    site_sheet = workbook.create_sheet("Sites Atelier")
    site_sheet.cell(1, 1).value = "Nom du site atelier"
    site_sheet.column_dimensions["A"].width = 24

    # ── Observation disponibilite ──────────────────────────────────────────────
    obs_sheet = workbook.create_sheet("Observation disponibilit\u00e9")
    obs_sheet.cell(1, 1).value = "Code type echantillon"
    obs_sheet.cell(1, 2).value = "Description"
    obs_sheet.column_dimensions["A"].width = 24
    obs_sheet.column_dimensions["B"].width = 30

    # ── Sens migratoire ────────────────────────────────────────────────────────
    sens_sheet = workbook.create_sheet("Sens migratoire")
    sens_sheet.cell(1, 1).value = "CodeMigration"
    sens_sheet.cell(1, 2).value = "Description CodeMigration"
    sens_sheet.column_dimensions["A"].width = 16
    sens_sheet.column_dimensions["B"].width = 28

    # ── Code marque individuelle ───────────────────────────────────────────────
    marque_sheet = workbook.create_sheet("Code marque ind")
    marque_sheet.cell(1, 1).value = "Code marque individuelle"
    marque_sheet.cell(1, 2).value = "Libell\u00e9"
    marque_sheet.column_dimensions["A"].width = 26
    marque_sheet.column_dimensions["B"].width = 30

    # ── Correspondants ─────────────────────────────────────────────────────────
    corr_sheet = workbook.create_sheet("Correspondants")
    for col, hdr in enumerate(["Numero correspondant", "Nom", "Pr\u00e9nom", "Adresse", "T\u00e9l\u00e9phone", "Mail"], start=1):
        corr_sheet.cell(1, col).value = hdr
    corr_sheet.column_dimensions["A"].width = 22
    corr_sheet.column_dimensions["B"].width = 18
    corr_sheet.column_dimensions["C"].width = 18
    corr_sheet.column_dimensions["D"].width = 30
    corr_sheet.column_dimensions["E"].width = 16
    corr_sheet.column_dimensions["F"].width = 28

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()
    return output_path


def create_target_workbook_from_template(output_path: Path, openpyxl_module, template_path: Path) -> Path:
    """Clone the provided template workbook while keeping only headers/structure."""
    workbook = openpyxl_module.load_workbook(template_path)
    try:
        target_sheet = workbook[DEFAULT_TARGET_SHEET] if DEFAULT_TARGET_SHEET in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        _clear_worksheet_data_keep_header(target_sheet)

        for sheet_name in workbook.sheetnames:
            if sheet_name == target_sheet.title:
                continue
            if normalize_sheet_name(sheet_name) == "type echantillon":
                continue
            _clear_worksheet_data_keep_header(workbook[sheet_name])

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return output_path
    finally:
        workbook.close()


def _clear_worksheet_data_keep_header(worksheet) -> None:
    """Remove row data while preserving the first header row and sheet layout."""
    max_row = worksheet.max_row
    max_col = worksheet.max_column
    if max_row <= 1:
        return

    for row_index in range(2, max_row + 1):
        for col_index in range(1, max_col + 1):
            cell = worksheet.cell(row_index, col_index)
            cell.value = None
            cell._comment = None
            cell.hyperlink = None


def normalize_sheet_name(value: object) -> str:
    return _normalize_header(value)


def validate_collect_science_source_workbook(workbook, sheet_name: str | None = None) -> Tuple[bool, str]:
    """Validate that an Excel file follows the COLISA structure needed by Collect-Science."""
    target_sheet = sheet_name or DEFAULT_TARGET_SHEET

    if target_sheet in workbook.sheetnames:
        worksheet = workbook[target_sheet]
    else:
        worksheet = workbook[workbook.sheetnames[0]] if workbook.sheetnames else None

    if worksheet is None:
        return False, "Le fichier Excel ne contient aucune feuille exploitable."

    header_row = [worksheet.cell(1, col_index).value for col_index in range(1, min(worksheet.max_column + 1, 50))]
    normalized_headers = [_normalize_header(value) for value in header_row if _normalize_header(value)]

    # Recherche flexible : cherche si les mots-clés sont présents dans les en-têtes
    # Format : (mots-clés requis, mots-clés optionnels/alternatifs)
    required_keywords = {
        "Numero individu": (["individu"], ["numero"]),
        "Code espece": (["espece"], ["code"]),
        "Pays capture": (["pays"], ["capture"]),
        "Date capture": (["date"], ["capture"]),
        "Lac/riviere": (["lac", "riviere"], []),
        "Longueur totale": (["longueur"], ["totale", "mm"]),
        "Ecailles brutes": (["ecailles"], ["brutes"]),
        "Montees": (["montees"], []),
        "Empreintes": (["empreintes"], []),
        "Otolithes": (["otolithes"], ["otolithe"]),
        "Code echantillon": (["echantillon"], ["code"]),
    }

    def _matches_keywords(header: str, keywords_tuple: tuple) -> bool:
        """Check if required keywords are present, and at least one optional if provided."""
        required, optional = keywords_tuple
        # Tous les mots-clés requis doivent être présents
        has_required = all(keyword in header for keyword in required)
        if not has_required:
            return False
        # Si pas d'optionnels, c'est bon
        if not optional:
            return True
        # Sinon, au moins un optionnel doit être présent
        return any(keyword in header for keyword in optional)

    missing_headers = []
    for label, keywords_tuple in required_keywords.items():
        found = any(_matches_keywords(norm_header, keywords_tuple) for norm_header in normalized_headers)
        if not found:
            missing_headers.append(label)

    if missing_headers:
        missing_text = ", ".join(missing_headers[:4])
        if len(missing_headers) > 4:
            missing_text += ", ..."
        return (
            False,
            "Le fichier Excel choisi ne contient pas les colonnes COLISA attendues. "
            f"Colonnes manquantes: {missing_text}.",
        )

    return True, ""


def _column_letter(index: int) -> str:
    """Convert a 1-based column index to an Excel column letter."""
    letters = []
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        letters.append(chr(65 + remainder))
    return "".join(reversed(letters))


def _normalize_header(value: object) -> str:
    """Normalize header to a simple searchable form."""
    if value is None:
        return ""
    import re
    text = str(value).strip().lower()
    # Normaliser les accents
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    # Enlever tout ce qui n'est pas alphanumérique ou espace
    text = re.sub(r'[^a-z0-9\s]', ' ', text)
    # Enlever les espaces multiples
    return " ".join(text.split())
