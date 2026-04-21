#!/usr/bin/env python3
"""Generate the 'format COLISA logiciel' workbook from software rows."""

from __future__ import annotations

import datetime
import unicodedata
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook

import datacolisa_importer as core
from domain.value_objects import DateCapture
from infrastructure.file_value_normalizer import coerce_colisa_header_value


DEFAULT_CODE_SITE = "V1235003"
DEFAULT_COUNTRY_CODES = {
    "france": "FR",
    "suisse": "CH",
    "switzerland": "CH",
    "allemagne": "DE",
    "germany": "DE",
    "italie": "IT",
    "italy": "IT",
}


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace("_", " ").replace("'", " ")
    return " ".join(text.split())


def _country_code(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if len(text) == 2:
        return text.upper()
    return DEFAULT_COUNTRY_CODES.get(text.lower(), text.upper())


def _build_header_map(worksheet) -> Dict[str, int]:
    return {
        _normalize_header(worksheet.cell(1, col_index).value): col_index
        for col_index in range(1, worksheet.max_column + 1)
        if _normalize_header(worksheet.cell(1, col_index).value)
    }


def _cell_value(row_values: List[Any], header_map: Dict[str, int], header_label: str) -> Any:
    col_index = header_map.get(_normalize_header(header_label))
    if not col_index:
        return None
    zero_based = col_index - 1
    if zero_based < 0 or zero_based >= len(row_values):
        return None
    return row_values[zero_based]


def _set_if_present(worksheet, row_index: int, header_map: Dict[str, int], header_label: str, value: Any) -> None:
    col_index = header_map.get(_normalize_header(header_label))
    if col_index:
        worksheet.cell(row_index, col_index).value = coerce_colisa_header_value(header_label, value)


def _clear_data_rows(worksheet) -> None:
    if worksheet.max_row <= 1:
        return
    for row_index in range(2, worksheet.max_row + 1):
        for col_index in range(1, worksheet.max_column + 1):
            worksheet.cell(row_index, col_index).value = None


def _sampling_date(value: Any):
    parsed = DateCapture.from_excel(value)
    return parsed.date if parsed else value


def _otolithe_value(value: Any) -> int:
    text = str(value or "").strip().upper()
    if not text or text in {"0", "NON", "NO", "FALSE", "FAUX", "N"}:
        return 0
    return 1


def _sample_code(row: Dict[str, Any]) -> str:
    return str(
        row.get("code_echantillon")
        or row.get("ref")
        or row.get("num_individu")
        or ""
    ).strip()


def _capture_number(row: Dict[str, Any], sample_code: str) -> str:
    return str(row.get("num_individu") or sample_code or "").strip()


def _source_observations(row: Dict[str, Any]) -> str:
    parts: List[str] = []
    for key in ("observations", "information_disponibilite"):
        value = str(row.get(key) or "").strip()
        if value in {"+", "++", "+++"}:
            continue
        if value and value not in parts:
            parts.append(value)
    return " | ".join(parts)


def _clean_export_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        text = value.strip()
        if text in {"-", "N"}:
            return ""
        return text
    return value


def _rows_from_header_map(data_rows: List[List[Any]], header_map: Dict[str, int]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for values in data_rows:
        row_values = list(values)
        code_echantillon = _cell_value(row_values, header_map, "Code echantillon")
        num_individu = _cell_value(row_values, header_map, "Numero individu (numero de capture)")
        if not code_echantillon and not num_individu:
            continue

        ot_gauche = _cell_value(row_values, header_map, "Presence de l'otolithe gauche (0 si non, 1 si oui)")
        ot_droit = _cell_value(row_values, header_map, "Presence de l'otolithe droite (0 si non, 1 si oui)")
        otolithes = ""
        if str(ot_gauche or "").strip() == "1" or str(ot_droit or "").strip() == "1":
            otolithes = "1"

        observation_disponibilite = str(_cell_value(row_values, header_map, "Observation disponibilite") or "").strip()
        observations = str(_cell_value(row_values, header_map, "Observations") or "").strip()

        rows.append({
            "selected": True,
            "ref": str(code_echantillon or num_individu or "").strip(),
            "code_echantillon": str(code_echantillon or "").strip(),
            "code_type_echantillon": str(_cell_value(row_values, header_map, "Code type echantillon") or "").strip(),
            "categorie": str(_cell_value(row_values, header_map, "Categorie pecheur") or "").strip(),
            "type_peche": str(_cell_value(row_values, header_map, "Type peche/engin") or "").strip(),
            "autre_oss": str(_cell_value(row_values, header_map, "Autre echantillon osseuses collectee sur l'individu OUI/NON") or "").strip(),
            "ecailles_brutes": "",
            "montees": "",
            "empreintes": observation_disponibilite,
            "otolithes": otolithes,
            "observation_disponibilite": observation_disponibilite,
            "num_individu": str(num_individu or "").strip(),
            "date_capture": _cell_value(row_values, header_map, "Date capture"),
            "code_espece": str(_cell_value(row_values, header_map, "Code espece") or "").strip(),
            "lac_riviere": str(_cell_value(row_values, header_map, "Lac/riviere") or "").strip(),
            "pays_capture": str(_cell_value(row_values, header_map, "Pays capture") or "").strip(),
            "pecheur": str(_cell_value(row_values, header_map, "Nom de l'operateur") or "").strip(),
            "longueur_mm": _cell_value(row_values, header_map, "Longueur totale (mm)"),
            "poids_g": _cell_value(row_values, header_map, "Poids (g)"),
            "maturite": str(_cell_value(row_values, header_map, "Code maturite sexuelle") or "").strip(),
            "sexe": str(_cell_value(row_values, header_map, "Code sexe") or "").strip(),
            "age_total": str(_cell_value(row_values, header_map, "Age total") or "").strip(),
            "code_stade": str(_cell_value(row_values, header_map, "Code stade") or "").strip(),
            "site_atelier": str(_cell_value(row_values, header_map, "Site Atelier") or "").strip(),
            "numero_correspondant": str(_cell_value(row_values, header_map, "Numero du correspondant") or "").strip(),
            "organisme": str(_cell_value(row_values, header_map, "Organisme preleveur") or "").strip(),
            "nom_operateur": str(_cell_value(row_values, header_map, "Nom de l'operateur") or "").strip(),
            "otolithe_gauche": str(ot_gauche or "").strip(),
            "otolithe_droit": str(ot_droit or "").strip(),
            "observations": observations or observation_disponibilite,
            "code_site": DEFAULT_CODE_SITE,
        })
    return rows


def lire_rows_depuis_excel_colisa(colisa_path: Path) -> List[Dict[str, Any]]:
    """Read a COLISA-brut style workbook and convert rows to the app row format."""
    try:
        workbook = load_workbook(colisa_path, read_only=True, data_only=True)
        try:
            worksheet = workbook["Feuil1 "] if "Feuil1 " in workbook.sheetnames else workbook[workbook.sheetnames[0]]
            header_map = _build_header_map(worksheet)
            data_rows = [list(values) for values in worksheet.iter_rows(min_row=2, values_only=True)]
            rows = _rows_from_header_map(data_rows, header_map)
            if rows:
                return rows
        finally:
            workbook.close()
    except Exception:
        pass

    sheet_names = core.get_workbook_sheet_names(colisa_path)
    if not sheet_names:
        return []
    source_rows, _datemode = core.read_any_source_rows(colisa_path, sheet_names[0])
    if not source_rows:
        return []
    header_map = {
        _normalize_header(value): index + 1
        for index, value in enumerate(source_rows[0])
        if _normalize_header(value)
    }
    return _rows_from_header_map(source_rows[1:], header_map)


def generer_colisa_logiciel_depuis_rows(
    rows: List[Dict[str, Any]],
    template_path: Path,
    output_path: Path,
    default_code_unite_gestionnaire: str = "0042",
    default_site_atelier: str = "",
    default_numero_correspondant: str = "",
    default_organisme: str = "INRAE",
) -> Dict[str, Any]:
    if not template_path.exists():
        raise FileNotFoundError(f"Template introuvable: {template_path}")

    workbook = load_workbook(template_path)
    try:
        worksheet = workbook["Echantillons"] if "Echantillons" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        _clear_data_rows(worksheet)
        header_map = _build_header_map(worksheet)

        target_row = 2
        for row in rows:
            if not row:
                continue

            sample_code = _sample_code(row)
            if not sample_code:
                continue

            capture_number = _capture_number(row, sample_code)
            date_capture = _sampling_date(row.get("date_capture"))
            year_value = ""
            month_value = ""
            if isinstance(date_capture, datetime.datetime):
                date_capture = date_capture.date()
            if isinstance(date_capture, datetime.date):
                year_value = date_capture.year
                month_value = date_capture.month

            _set_if_present(worksheet, target_row, header_map, "UUID", None)
            _set_if_present(worksheet, target_row, header_map, "Code unite gestionnaire", default_code_unite_gestionnaire or "0042")
            _set_if_present(worksheet, target_row, header_map, "Pays", _clean_export_value(_country_code(row.get("pays_capture"))))
            _set_if_present(worksheet, target_row, header_map, "Site atelier", _clean_export_value(row.get("site_atelier", default_site_atelier)))
            _set_if_present(worksheet, target_row, header_map, "Code type echantillon", _clean_export_value(row.get("code_type_echantillon", "")))
            _set_if_present(worksheet, target_row, header_map, "Code echantillon", sample_code)
            _set_if_present(worksheet, target_row, header_map, "Code site", _clean_export_value(row.get("code_site", DEFAULT_CODE_SITE)))
            _set_if_present(worksheet, target_row, header_map, "Numero du correspondant", _clean_export_value(default_numero_correspondant))
            _set_if_present(worksheet, target_row, header_map, "Code espece", _clean_export_value(row.get("code_espece", "")))
            _set_if_present(worksheet, target_row, header_map, "Code stade", _clean_export_value(row.get("code_stade", "")))
            _set_if_present(worksheet, target_row, header_map, "Sens migratoire", _clean_export_value(row.get("sens_migratoire", "")))
            _set_if_present(worksheet, target_row, header_map, "Code maturite sexuelle", _clean_export_value(row.get("maturite", "")))
            _set_if_present(worksheet, target_row, header_map, "Particularite (becard, smolt, presmolt)", _clean_export_value(row.get("particularite", "")))
            _set_if_present(worksheet, target_row, header_map, "Code sexe", _clean_export_value(row.get("sexe", "")))
            _set_if_present(worksheet, target_row, header_map, "Code marque individuelle (Pits, IV ...)", _clean_export_value(row.get("code_marque_individuelle", "")))
            _set_if_present(worksheet, target_row, header_map, "Numero individu (numero de capture)", capture_number)
            _set_if_present(worksheet, target_row, header_map, "Date de capture (JJ/MM/AAAA)", date_capture)
            _set_if_present(worksheet, target_row, header_map, "Annee", year_value)
            _set_if_present(worksheet, target_row, header_map, "Mois", month_value)
            _set_if_present(worksheet, target_row, header_map, "Nombre", _clean_export_value(row.get("nombre", "")))
            _set_if_present(worksheet, target_row, header_map, "Organisme preleveur", _clean_export_value(default_organisme or "INRAE"))
            _set_if_present(worksheet, target_row, header_map, "Nom de l'operateur", _clean_export_value(row.get("nom_operateur", "")))
            _set_if_present(worksheet, target_row, header_map, "Longueur totale (mm)", _clean_export_value(row.get("longueur_mm", "")))
            _set_if_present(worksheet, target_row, header_map, "Longueur totale estimee ? (0 pour non, 1 pour oui)", _clean_export_value(row.get("longueur_totale_estimee", "")))
            _set_if_present(worksheet, target_row, header_map, "Longueur fourche (mm)", _clean_export_value(row.get("longueur_fourche_mm", "")))
            _set_if_present(worksheet, target_row, header_map, "Longueur fourche estimee ? (0 pour non, 1 pour oui)", _clean_export_value(row.get("longueur_fourche_estimee", "")))
            _set_if_present(worksheet, target_row, header_map, "Poids (g)", _clean_export_value(row.get("poids_g", "")))
            _set_if_present(worksheet, target_row, header_map, "Numero de marque individuelle", _clean_export_value(row.get("numero_marque_individuelle", "")))
            _set_if_present(worksheet, target_row, header_map, "Code age", _clean_export_value(row.get("code_age") or row.get("age_total", "")))
            _set_if_present(worksheet, target_row, header_map, "Poux de mer (0 pour non, 1 pour oui)", row.get("poux_de_mer", 0))
            _set_if_present(worksheet, target_row, header_map, "Presence de l'otolithe gauche ? (0 pour non, 1 pour oui)", row.get("otolithe_gauche", _otolithe_value(row.get("otolithes"))))
            _set_if_present(worksheet, target_row, header_map, "Presence de l'otolithe droit ? (0 pour non, 1 pour oui)", row.get("otolithe_droit", _otolithe_value(row.get("otolithes"))))
            _set_if_present(worksheet, target_row, header_map, "Observations", _clean_export_value(_source_observations(row)))
            _set_if_present(worksheet, target_row, header_map, "Lien Externe", _clean_export_value(row.get("lien_externe", "")))

            target_row += 1

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return {"excel": str(output_path), "rows_written": max(0, target_row - 2)}
    finally:
        workbook.close()
