#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import os
import re
import shutil
import sys
from pathlib import Path
from typing import Any, Dict, List

from PySide6.QtCore import Qt, QModelIndex
from PySide6.QtGui import QAction, QColor, QIcon, QPainter, QPen, QPixmap
from PySide6.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QCheckBox,
    QComboBox,
    QDialog,
    QFrame,
    QFileDialog,
    QGridLayout,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QInputDialog,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QScrollArea,
    QSizePolicy,
    QTableView,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

import datacolisa_importer as core
from infrastructure.app_paths import app_assets_dir, data_dir, ensure_runtime_dirs, exports_dir, runtime_dir, settings_dir
from infrastructure.file_value_normalizer import coerce_internal_value
from infrastructure.internal_target_workbook import (
    create_internal_target_workbook,
    validate_collect_science_source_workbook,
)
from presentation.table_model import ImportTableModel
from presentation.table_delegate import ComboBoxDelegate, EditableComboBoxDelegate, MappedEditableComboBoxDelegate

# Style QComboBox compatible thème jour ET nuit
COMBOBOX_STYLE = """
QComboBox {
    min-height: 24px;
    padding: 2px 4px;
}
QComboBox::drop-down {
    subcontrol-origin: padding;
    subcontrol-position: top right;
    width: 18px;
}
QComboBox::down-arrow {
    border-left: 4px solid transparent;
    border-right: 4px solid transparent;
    border-top: 5px solid currentColor;
}
"""

TYPE_PECHE_OPTIONS = ["", "LIGNE", "FILET", "TRAINE", "SONDE"]
CATEGORIE_OPTIONS = ["", "PRO", "AMATEUR", "SCIENTIFIQUE"]
OUI_NON_OPTIONS = ["", "OUI", "NON"]
OBSERVATION_OPTIONS = ["", "+", "++", "+++"]
NUMERIC_OPTIONS = [""] + [str(i) for i in range(1, 11)]
ECAILLES_BRUTES_OPTIONS = [""] + [str(i) for i in range(1, 21)]
APP_VERSION = "1.3.0"
APP_ORGANISATION = "INRAE"
APP_AUTHOR = "Anatole Aubin et Quentin Godeaux"
SETTINGS_DIR = settings_dir("DATACOLISA")
SETTINGS_FILE = SETTINGS_DIR / "ui_pyside6_settings.json"
SAMPLE_TYPE_LABELS = {
    "BI": "Bile",
    "GN": "Bouche de poisson",
    "BN": "Branchie de poisson",
    "VN": "Colonne vertebrale de poisson",
    "HN": "Dos de Poisson",
    "EC": "Ecaille de poisson",
    "ES": "Estomac",
    "FO": "Foie de poisson",
    "FN": "Fraction inconnue de poisson",
    "GO": "Gonade de poisson",
    "GR": "Graisse de poisson",
    "HM": "Hemocytes de poissons",
    "LE": "Levre de poisson",
    "MN": "Machoire de poisson",
    "MU": "Muscle de poisson",
    "MP": "Muscle et peau de poisson",
    "MI": "Muscle, muscle+tissu adipeux face interne de peau",
    "QN": "Nageoire caudale de poisson",
    "NN": "Nageoire de poisson",
    "DN": "Nageoire dorsale de poisson",
    "PN": "Nageoire pectorale de poisson",
    "YN": "Oeil de poisson",
    "ON": "Opercules",
    "XN": "Orifice anal de poisson",
    "XU": "Orifice urogenital de poisson",
    "OT": "Otolithes",
    "KN": "Pedoncule caudal de poisson",
    "CN": "Poisson entier",
    "WE": "Poisson etete et equeute",
    "WV": "Poisson sans visceres ni gonades",
    "RE": "Rein de poisson",
    "NF": "Systeme nerveux de poisson",
    "TN": "Tete de poisson",
    "WN": "Tronc de poisson",
}


def format_sample_type_display(code: str) -> str:
    label = SAMPLE_TYPE_LABELS.get(code, "").strip()
    return f"{code} - {label}" if label else code


def normalize_sample_type_code(value: str) -> str:
    text = str(value or "").strip()
    normalized = core.normalize(text).upper()
    if normalized in {"EC MONTEE", "EC MONTEE", "EC BRUTE", "ECAILLE MONTEE", "ECAILLE BRUTE", "ECAILLES MONTEES", "ECAILLES BRUTES"}:
        return "EC"
    if " - " in text:
        return text.split(" - ", 1)[0].strip()
    return text

def _famille_ec_ot(code_type: str) -> str:
    """Retourne 'EC', 'OT' ou '' selon le type d'echantillon."""
    t = str(code_type or "").strip().upper()
    if "EC" in t or "ECAILLE" in t:
        return "EC"
    if "OT" in t or "OTOLITH" in t:
        return "OT"
    return ""


def fusionner_lignes_ec_ot(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Fusionne les lignes EC et OT ayant le meme num_individu en une seule ligne.

    Regles :
    - Seulement pour EC (ecaille) et OT (otolithe). Les autres types restent inchanges.
    - Toutes les lignes EC/OT du meme individu (qu'elles soient EC, OT, montee, brute)
      sont fusionnees en UNE seule ligne.
    - Priorite du type : si EC et OT coexistent, la ligne finale prend le type EC.
    - Les colonnes ecailles_brutes, montees, empreintes, otolithes sont cumulees.
    - La premiere ligne du groupe sert de base pour tous les autres champs.
    - L'ordre d'apparition est preserve (position de la premiere occurrence).
    """
    COLS_COMPTAGE = ["ecailles_brutes", "montees", "empreintes", "otolithes"]

    result: List[Dict[str, Any]] = []
    # num_individu -> index dans result (une seule entree par individu pour EC/OT)
    index_par_individu: Dict[str, int] = {}

    for row in rows:
        famille = _famille_ec_ot(row.get("code_type_echantillon", ""))
        num_ind = str(row.get("num_individu", "") or "").strip()

        if not famille or not num_ind:
            # Type non concerne ou pas de num_individu : ligne normale
            result.append(dict(row))
            continue

        if num_ind in index_par_individu:
            # Un partenaire EC/OT existe deja pour cet individu : fusionner
            base = result[index_par_individu[num_ind]]

            # EC prend le dessus sur OT si l'un des deux est EC
            base_famille = _famille_ec_ot(base.get("code_type_echantillon", ""))
            if famille == "EC" and base_famille == "OT":
                base["code_type_echantillon"] = row.get("code_type_echantillon", base["code_type_echantillon"])

            # Fusion des colonnes de comptage
            for col in COLS_COMPTAGE:
                val_base = str(base.get(col, "") or "").strip()
                val_new = str(row.get(col, "") or "").strip()
                if (not val_base or val_base == "0") and val_new and val_new != "0":
                    base[col] = val_new
                elif val_base and val_base != "0" and val_new and val_new != "0":
                    try:
                        base[col] = str(max(int(val_base), int(val_new)))
                    except ValueError:
                        pass  # Conserver val_base en cas de valeur non numerique
            # Pas de nouvelle ligne ajoutee : la ligne entrante est absorbee
        else:
            # Premiere occurrence EC/OT pour cet individu
            index_par_individu[num_ind] = len(result)
            result.append(dict(row))

    return result


def qitem(text: Any, editable: bool = True) -> QTableWidgetItem:
    it = QTableWidgetItem(str(text) if text is not None else "")
    if not editable:
        it.setFlags(it.flags() & ~Qt.ItemIsEditable)
    return it


def _app_base_dir() -> Path:
    return runtime_dir()


def _app_data_dir() -> Path:
    return data_dir()


def _app_logo_path() -> Path:
    return app_assets_dir() / "colisa_fr.png"


def _app_exports_dir() -> Path:
    return exports_dir()


def _default_source_start_dir(current: Path | None, fallback: Path) -> str:
    if current:
        try:
            if str(current).strip():
                parent = current.parent
                if parent.exists():
                    return str(parent)
        except Exception:
            pass
    return str(fallback)


class MainWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()

        ensure_runtime_dirs()

        # Setup logging
        from config.logging_config import setup_logging
        app_base = _app_base_dir()
        setup_logging(log_level="INFO", log_file=app_base / "datacolisa.log", console=True)
        import logging
        self.logger = logging.getLogger(__name__)
        self.logger.info("Application started")

        self.setWindowTitle("DATACOLISA - Gestion des Données Scientifiques")
        self.resize(1680, 920)

        # Apply modern stylesheet (mode nuit force)
        from presentation.styles import get_stylesheet
        self.setStyleSheet(get_stylesheet(dark_mode=True))

        self.rows: List[Dict[str, Any]] = []
        self.type_options: List[str] = []
        # Cache des individus EC/OT deja presents dans le fichier de sortie
        # Structure : { num_individu -> {"row", "code_type_echantillon", "montees", "ecailles_brutes", "otolithes", "empreintes"} }
        self._individus_sortie: Dict[str, Dict[str, str]] = {}
        self.missing_codes: List[str] = []
        self._table_sync_guard = False
        self.dark_mode = True
        self._post_import_pipeline = False
        self._import_start_numero = 0

        self.source_path = Path()
        self.source_sheet = core.DEFAULT_SOURCE_SHEET
        self.source_mode = "pac_final"
        self.source_mapping: Dict[str, Any] = {}
        self.target_sheet = core.DEFAULT_TARGET_SHEET
        # Dossier de sortie portable/local au logiciel
        self.imports_dir = _app_exports_dir()
        self.target_path = _app_data_dir() / "COLISA_base_integree.xlsx"
        # Template embarqué dans le code — plus besoin de fichier externe
        from infrastructure.embedded_assets import get_colisa_logiciel_template_path as _get_tpl
        self.colisa_logiciel_template_path = _get_tpl()

        self.out_path = self.imports_dir / "COLISA en cours.xlsx"
        self.history_path = self.imports_dir / "import_history.json"
        self.selection_csv = self.imports_dir / "selection_import.csv"

        self._build_menu()

        root = QWidget()
        self.setCentralWidget(root)
        lay = QVBoxLayout(root)
        lay.setSpacing(8)
        lay.setContentsMargins(10, 10, 10, 10)

        lay.addWidget(self._build_context_strip())
        lay.addWidget(self._build_workspace_panel(), 1)
        lay.addWidget(self._build_bottom_panel())

        self._load_settings()
        self._refresh_type_options()
        self._refresh_context_labels()
        self._apply_theme(True)
        self.switch_theme.blockSignals(True)
        self.switch_theme.setChecked(True)
        self.switch_theme.blockSignals(False)
        self.switch_theme.setEnabled(False)

        from PySide6.QtCore import QTimer
        QTimer.singleShot(200, self._verifier_fichier_colisa_au_demarrage)

    def _verifier_fichier_colisa_au_demarrage(self) -> None:
        """Au démarrage, affiche un dialog pour choisir/confirmer les fichiers ENTREE et SORTIE."""
        from PySide6.QtWidgets import (
            QDialog, QVBoxLayout, QHBoxLayout, QLabel,
            QLineEdit, QPushButton, QDialogButtonBox, QFrame
        )
        from presentation.dialogs import ErrorDialog

        dlg = QDialog(self)
        dlg.setWindowTitle("DATACOLISA - Fichiers de travail")
        dlg.setMinimumWidth(580)
        dlg.setModal(True)

        lay = QVBoxLayout(dlg)
        lay.setSpacing(14)
        lay.setContentsMargins(16, 16, 16, 16)

        # ── ENTREE ──────────────────────────────────────────────────────────
        lbl_entree = QLabel("ENTREE - Fichier source (PAC Final / Excel) :")
        lbl_entree.setStyleSheet("font-weight: bold;")
        lay.addWidget(lbl_entree)

        row_in = QHBoxLayout()
        ed_source = QLineEdit(str(self.source_path) if str(self.source_path).strip() else "")
        ed_source.setMinimumWidth(380)
        row_in.addWidget(ed_source)

        btn_parcourir_in = QPushButton("Parcourir...")

        def _parcourir_in():
            current_text = ed_source.text().strip()
            current = Path(current_text) if current_text else None
            start_dir = _default_source_start_dir(current, self.imports_dir)
            p, _ = QFileDialog.getOpenFileName(
                dlg, "Choisir le fichier source", start_dir,
                "Excel (*.xls *.xlsx);;Tous (*.*)"
            )
            if p:
                ed_source.setText(p)

        btn_parcourir_in.clicked.connect(_parcourir_in)
        row_in.addWidget(btn_parcourir_in)
        lay.addLayout(row_in)

        # séparateur
        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setFrameShadow(QFrame.Sunken)
        lay.addWidget(sep)

        # ── SORTIE ──────────────────────────────────────────────────────────
        lbl_sortie = QLabel("SORTIE - Fichier COLISA en cours :")
        lbl_sortie.setStyleSheet("font-weight: bold;")
        lay.addWidget(lbl_sortie)

        row_out = QHBoxLayout()
        ed_path = QLineEdit(str(self.out_path))
        ed_path.setMinimumWidth(380)
        row_out.addWidget(ed_path)

        btn_parcourir_out = QPushButton("Parcourir...")

        def _parcourir_out():
            current = Path(ed_path.text().strip())
            start_dir = str(current.parent) if current.parent.exists() else str(self.imports_dir)
            p, _ = QFileDialog.getOpenFileName(dlg, "Choisir un fichier COLISA existant", start_dir, "Excel (*.xlsx)")
            if p:
                ed_path.setText(p)

        btn_parcourir_out.clicked.connect(_parcourir_out)
        row_out.addWidget(btn_parcourir_out)
        lay.addLayout(row_out)

        btn_nouveau = QPushButton("Creer un nouveau fichier SORTIE vierge...")

        # Conteneur pour transmettre le code de départ depuis la fonction imbriquée
        _starting_code: list[str] = [""]

        def _nouveau():
            current = Path(ed_path.text().strip())
            start = str(current.parent) if current.parent.exists() else str(self.imports_dir)
            p, _ = QFileDialog.getSaveFileName(dlg, "Nouveau fichier COLISA", start + "/COLISA en cours.xlsx", "Excel (*.xlsx)")
            if not p:
                return
            dest = Path(p)
            if dest.suffix.lower() != ".xlsx":
                dest = dest.with_suffix(".xlsx")
            if dest.exists():
                from presentation.dialogs import WarningDialog
                WarningDialog.show(dlg, "Nouveau fichier", f"Ce fichier existe déjà :\n{dest}")
                return
            try:
                base_path = self._ensure_internal_target_base()
                dest.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(base_path, dest)
                ed_path.setText(str(dest))
                # Demander le code échantillon de départ
                code, ok = QInputDialog.getText(
                    dlg,
                    "Code échantillon de départ",
                    "Par quel code échantillon commencer ?\n(ex : T001, CA950...)",
                    text=self.ed_start.text().strip(),
                )
                if ok and code.strip():
                    _starting_code[0] = code.strip().upper()
            except Exception as exc:
                ErrorDialog.show(dlg, "Nouveau fichier", f"Impossible de créer :\n{exc}")

        btn_nouveau.clicked.connect(_nouveau)
        lay.addWidget(btn_nouveau)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        lay.addWidget(btns)

        if dlg.exec() != QDialog.Accepted:
            return

        # Appliquer ENTREE
        source_text = ed_source.text().strip()
        new_source = Path(source_text) if source_text else Path()
        if new_source != self.source_path:
            self.source_path = new_source

        # Appliquer SORTIE
        chosen = Path(ed_path.text().strip())
        if not chosen.suffix:
            chosen = chosen.with_suffix(".xlsx")
        self.out_path = chosen

        # Appliquer le code échantillon de départ si renseigné
        if _starting_code[0]:
            self.ed_start.setText(_starting_code[0])

        self._refresh_context_labels()
        self._refresh_type_options()
        self._save_settings()

    def _build_menu(self) -> None:
        mb = self.menuBar()
        menu_file = mb.addMenu("Fichier")

        act_source = QAction("PAC final...", self)
        act_source.triggered.connect(self.select_source_file)
        menu_file.addAction(act_source)

        act_other_excel = QAction("Autre type de source...", self)
        act_other_excel.triggered.connect(self.select_other_excel_source)
        menu_file.addAction(act_other_excel)

        menu_file.addSeparator()
        act_reset_out = QAction("Refaire COLISA en cours", self)
        act_reset_out.triggered.connect(self.refaire_fichier_import)
        menu_file.addAction(act_reset_out)

        menu_file.addSeparator()
        act_about = QAction("A propos", self)
        act_about.triggered.connect(self.show_about)
        menu_file.addAction(act_about)

        act_manual = QAction("Manuel d'utilisation", self)
        act_manual.triggered.connect(self.show_manual)
        menu_file.addAction(act_manual)

        menu_file.addSeparator()
        act_quit = QAction("Quitter", self)
        act_quit.triggered.connect(self.close)
        menu_file.addAction(act_quit)

    def _build_context_strip(self) -> QWidget:
        w = QWidget()
        h = QHBoxLayout(w)
        h.setContentsMargins(0, 0, 0, 0)
        h.setSpacing(8)

        self.lbl_source = QLabel()
        self.lbl_target = QLabel()
        self.lbl_paths = QLabel()
        for lbl in (self.lbl_source, self.lbl_target, self.lbl_paths):
            lbl.setWordWrap(True)

        info = QFrame()
        info.setObjectName("contextCard")
        info_v = QVBoxLayout(info)
        info_v.setContentsMargins(10, 8, 10, 8)
        info_v.setSpacing(2)
        info_v.addWidget(self.lbl_source)
        info_v.addWidget(self.lbl_target)
        info_v.addWidget(self.lbl_paths)

        self.switch_theme = QCheckBox("Mode nuit")
        self.switch_theme.setObjectName("themeSwitch")
        self.switch_theme.toggled.connect(self._on_theme_toggled)
        self.switch_theme.hide()

        h.addWidget(info, 3)
        return w

    def _build_workspace_panel(self) -> QWidget:
        w = QWidget()
        v = QVBoxLayout(w)
        v.setContentsMargins(0, 0, 0, 0)
        v.setSpacing(8)

        v.addWidget(self._build_bulk_panel())

        h = QHBoxLayout()
        h.setContentsMargins(0, 0, 0, 0)
        h.setSpacing(12)

        left = QWidget()
        left.setMaximumWidth(420)
        left_lay = QVBoxLayout(left)
        left_lay.setContentsMargins(0, 0, 0, 0)
        left_lay.setSpacing(8)
        left_lay.addWidget(self._build_top_panel())
        left_lay.addWidget(self._build_metrics_panel())
        left_lay.addStretch()

        h.addWidget(left, 1)
        h.addWidget(self._build_table(), 3)
        v.addLayout(h, 1)
        return w

    def _build_top_panel(self) -> QWidget:
        box = QGroupBox("Parametres d'import")
        box.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        g = QGridLayout(box)
        g.setHorizontalSpacing(10)
        g.setVerticalSpacing(6)
        g.setColumnStretch(0, 0)
        g.setColumnStretch(1, 1)
        g.setColumnStretch(2, 0)
        g.setColumnStretch(3, 1)

        self.ed_start = QLineEdit("CA961")
        self.ed_end = QLineEdit("CA989")

        self.cb_duplicate = QComboBox(); self.cb_duplicate.addItems(["alert", "ignore", "replace"])
        self.cb_duplicate.setCurrentText("alert")


        self.ed_org = QLineEdit("INRAE")
        self.ed_country = QLineEdit("France")
        self.ed_code_unite = QLineEdit("0042")
        self.ed_site_atelier = QLineEdit("THONON")
        self.ed_num_correspondant = QLineEdit("4")

        # Largeurs fixes pour garantir visibilité complète
        self.ed_start.setMinimumWidth(124)
        self.ed_end.setMinimumWidth(124)
        self.cb_duplicate.setMinimumWidth(124)
        self.ed_org.setMinimumWidth(124)
        self.ed_country.setMinimumWidth(124)
        self.ed_code_unite.setMinimumWidth(124)
        self.ed_site_atelier.setMinimumWidth(124)
        self.ed_num_correspondant.setMinimumWidth(124)

        btn_add_type = QPushButton("Ajouter type")
        btn_add_type.setMinimumWidth(140)
        btn_add_type.setFixedHeight(30)
        btn_add_type.clicked.connect(self.add_new_type)

        btn_load = QPushButton("📂 Charger plage")
        btn_load.setMinimumWidth(150)
        btn_load.setFixedHeight(32)
        btn_load.clicked.connect(self.load_range)


        labels = [
            ("Numero individu debut", self.ed_start),
            ("Numero individu fin", self.ed_end),
            ("Doublons", self.cb_duplicate),
            ("Organisme", self.ed_org),
            ("Pays", self.ed_country),
            ("Code unite", self.ed_code_unite),
            ("Site atelier", self.ed_site_atelier),
            ("Numero correspondant", self.ed_num_correspondant),
        ]

        row = 0
        for lbl, w in labels:
            g.addWidget(QLabel(lbl), row, 0, alignment=Qt.AlignLeft)
            g.addWidget(w, row, 1)  # Pas d'alignment pour permettre expansion
            row += 1

        helper = QLabel("Change surtout la plage.")
        helper.setWordWrap(True)
        helper.setObjectName("panelHelp")

        g.addWidget(helper, 0, 2, 2, 2)
        g.addWidget(btn_add_type, 2, 2, alignment=Qt.AlignLeft)

        g.addWidget(btn_load, row, 0, 1, 2, alignment=Qt.AlignLeft)
        return box

    def _build_help_panel(self) -> QWidget:
        box = QGroupBox("Aide rapide")
        box.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        v = QVBoxLayout(box)
        v.setContentsMargins(8, 6, 8, 6)
        v.setSpacing(6)

        note = QLabel("Besoin d'aide ? Ouvre Fichier > Manuel d'utilisation.")
        note.setWordWrap(True)
        note.setObjectName("workflowNote")
        v.addWidget(note)
        return box

    def _build_metrics_panel(self) -> QWidget:
        box = QGroupBox("Resume")
        box.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        h = QHBoxLayout(box)
        self.lbl_count = QLabel("Lignes: 0")
        self.lbl_pending = QLabel("A reimporter: 0")
        self.lbl_missing = QLabel("Codes manquants: 0")
        btn_missing = QPushButton("Voir codes manquants")
        btn_missing.clicked.connect(self.show_missing_codes)
        h.addWidget(self.lbl_count)
        h.addWidget(self.lbl_pending)
        h.addWidget(self.lbl_missing)
        h.addWidget(btn_missing)
        return box

    def _build_bulk_panel(self) -> QWidget:
        container = QWidget()
        outer = QVBoxLayout(container)
        outer.setContentsMargins(0, 0, 0, 0)

        box = QGroupBox("Actions multi-lignes (sur cases Selection)")
        box.setSizePolicy(QSizePolicy.MinimumExpanding, QSizePolicy.Fixed)
        box.setMinimumWidth(1500)
        box_layout = QVBoxLayout(box)
        box_layout.setContentsMargins(14, 12, 14, 12)
        box_layout.setSpacing(10)

        self.bulk_type = QComboBox()
        self.bulk_type.setEditable(True)
        self.bulk_type.setInsertPolicy(QComboBox.NoInsert)
        self.bulk_categorie = QComboBox(); self.bulk_categorie.addItems(CATEGORIE_OPTIONS)
        self.bulk_type_peche = QComboBox(); self.bulk_type_peche.addItems(TYPE_PECHE_OPTIONS)
        self.bulk_autre = QComboBox(); self.bulk_autre.addItems(OUI_NON_OPTIONS)
        self.bulk_ecailles = QComboBox(); self.bulk_ecailles.addItems(ECAILLES_BRUTES_OPTIONS)
        self.bulk_montees = QComboBox(); self.bulk_montees.addItems(NUMERIC_OPTIONS)
        self.bulk_empreintes = QComboBox(); self.bulk_empreintes.addItems(NUMERIC_OPTIONS)
        self.bulk_otolithes = QComboBox(); self.bulk_otolithes.addItems(NUMERIC_OPTIONS)
        self.bulk_observation = QComboBox(); self.bulk_observation.addItems(OBSERVATION_OPTIONS)
        for cb in [self.bulk_type, self.bulk_categorie, self.bulk_type_peche,
                   self.bulk_autre, self.bulk_ecailles, self.bulk_montees, self.bulk_empreintes, self.bulk_otolithes, self.bulk_observation]:
            cb.setMinimumWidth(135)

        fields_grid = QGridLayout()
        fields_grid.setHorizontalSpacing(18)
        fields_grid.setVerticalSpacing(12)

        for index, (lbl, w) in enumerate([
            ("Type echantillon", self.bulk_type),
            ("Categorie", self.bulk_categorie),
            ("Type peche", self.bulk_type_peche),
            ("Autre oss", self.bulk_autre),
            ("Ecailles", self.bulk_ecailles),
            ("Montees", self.bulk_montees),
            ("Empreintes", self.bulk_empreintes),
            ("Otolithes", self.bulk_otolithes),
            ("Observation", self.bulk_observation),
        ]):
            row = index // 4
            col = index % 4
            field_widget = QWidget()
            field_layout = QVBoxLayout(field_widget)
            field_layout.setContentsMargins(0, 0, 0, 0)
            field_layout.setSpacing(5)
            field_layout.addWidget(QLabel(lbl))
            field_layout.addWidget(w)
            fields_grid.addWidget(field_widget, row, col)
        box_layout.addLayout(fields_grid)


        btn_sel_all = QPushButton("Tout selectionner")
        btn_sel_none = QPushButton("Vider selection")

        btn_apply = QPushButton("Appliquer")
        btn_sel_all.clicked.connect(lambda: self._select_all(True))
        btn_sel_none.clicked.connect(lambda: self._select_all(False))
        btn_apply.clicked.connect(self.apply_bulk)

        buttons_row = QHBoxLayout()
        buttons_row.addWidget(btn_sel_all)
        buttons_row.addWidget(btn_sel_none)
        buttons_row.addWidget(btn_apply)
        buttons_row.addStretch()
        buttons_row.setSpacing(10)
        box_layout.addLayout(buttons_row)

        scroll = QScrollArea()
        scroll.setWidgetResizable(False)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setWidget(box)

        outer.addWidget(scroll)
        return container

    def _build_table(self) -> QWidget:
        if hasattr(self, "table") and self.table is not None:
            placeholder = QWidget()
            placeholder.setFixedHeight(0)
            return placeholder

        # Create model and view (performance fix: QTableView + Model pattern)
        self.table_model = ImportTableModel()
        self.table = QTableView()
        self.table.setModel(self.table_model)

        # Configure view
        self.table.setAlternatingRowColors(True)
        self.table.setSortingEnabled(False)
        self.table.setShowGrid(True)
        self.table.setWordWrap(False)
        self.table.setCornerButtonEnabled(False)
        self.table.setHorizontalScrollMode(QTableView.ScrollPerPixel)
        self.table.setVerticalScrollMode(QTableView.ScrollPerPixel)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.horizontalHeader().setHighlightSections(False)
        self.table.verticalHeader().setDefaultSectionSize(26)
        self.table.setEditTriggers(
            QAbstractItemView.DoubleClicked
            | QAbstractItemView.SelectedClicked
            | QAbstractItemView.EditKeyPressed
            | QAbstractItemView.AnyKeyPressed
        )
        self.table.setSelectionBehavior(QTableView.SelectRows)
        self.table.setSelectionMode(QTableView.SingleSelection)
        self.table.clicked.connect(self._on_table_clicked)

        # Set up delegates for combo box columns (create editor only when editing!)
        self._setup_table_delegates()
        self._configure_table_columns()

        # Connect signals
        self.table_model.dataChanged.connect(self._on_table_data_changed)

        return self.table

    def _configure_table_columns(self) -> None:
        """Keep important columns readable while fitting in the window."""
        from presentation.table_model import COLUMNS

        header = self.table.horizontalHeader()
        resize_to_contents = {
            "selected",
            "ref",
            "date_capture",
            "code_espece",
            "status",
        }
        widths = {
            "selected": 44,
            "ref": 92,
            "code_type_echantillon": 260,
            "categorie": 100,
            "type_peche": 110,
            "autre_oss": 90,
            "ecailles_brutes": 82,
            "montees": 82,
            "empreintes": 82,
            "otolithes": 82,
            "observation_disponibilite": 110,
            "num_individu": 110,
            "date_capture": 108,
            "code_espece": 96,
            "lac_riviere": 130,
            "pays_capture": 100,
            "pecheur": 120,
            "longueur_mm": 92,
            "poids_g": 82,
            "maturite": 82,
            "sexe": 72,
            "age_total": 82,
            "status": 110,
            "errors": 240,
        }

        for index, name in enumerate(COLUMNS):
            if name in resize_to_contents:
                header.setSectionResizeMode(index, QHeaderView.ResizeToContents)
            elif name == "errors":
                header.setSectionResizeMode(index, QHeaderView.Stretch)
            else:
                header.setSectionResizeMode(index, QHeaderView.Interactive)
                self.table.setColumnWidth(index, widths.get(name, 100))

    def _on_table_clicked(self, index: QModelIndex) -> None:
        """Make the Selection checkbox reliably toggle on click."""
        if not index.isValid() or index.column() != 0:
            return
        current = self.table_model.data(index, Qt.CheckStateRole)
        next_value = Qt.Unchecked if current == Qt.Checked else Qt.Checked
        self.table_model.setData(index, next_value, Qt.CheckStateRole)

    def _setup_table_delegates(self) -> None:
        """Set up delegates for combo box columns (performance optimization)."""
        from presentation.table_model import COLUMNS

        # Get column indices
        col_indices = {name: idx for idx, name in enumerate(COLUMNS)}

        # Set delegates for combo box columns
        self.table.setItemDelegateForColumn(
            col_indices["code_type_echantillon"],
            MappedEditableComboBoxDelegate(
                [(format_sample_type_display(code), code) for code in self.type_options],
                self.table,
            )
        )
        self.table.setItemDelegateForColumn(
            col_indices["categorie"],
            ComboBoxDelegate(CATEGORIE_OPTIONS, self.table)
        )
        self.table.setItemDelegateForColumn(
            col_indices["type_peche"],
            EditableComboBoxDelegate(TYPE_PECHE_OPTIONS, self.table)
        )
        self.table.setItemDelegateForColumn(
            col_indices["autre_oss"],
            ComboBoxDelegate(OUI_NON_OPTIONS, self.table)
        )
        self.table.setItemDelegateForColumn(
            col_indices["ecailles_brutes"],
            EditableComboBoxDelegate(ECAILLES_BRUTES_OPTIONS, self.table)
        )
        self.table.setItemDelegateForColumn(
            col_indices["montees"],
            EditableComboBoxDelegate(NUMERIC_OPTIONS, self.table)
        )
        self.table.setItemDelegateForColumn(
            col_indices["empreintes"],
            EditableComboBoxDelegate(NUMERIC_OPTIONS, self.table)
        )
        self.table.setItemDelegateForColumn(
            col_indices["otolithes"],
            EditableComboBoxDelegate(NUMERIC_OPTIONS, self.table)
        )
        self.table.setItemDelegateForColumn(
            col_indices["observation_disponibilite"],
            EditableComboBoxDelegate(OBSERVATION_OPTIONS, self.table)
        )
    def _build_bottom_panel(self) -> QWidget:
        w = QWidget()
        h = QHBoxLayout(w)
        h.setSpacing(12)

        btn_import = QPushButton("Lancer import")
        btn_import.setObjectName("btn_import")
        btn_import.setMinimumWidth(180)
        btn_import.setFixedHeight(38)
        btn_import.clicked.connect(self._choisir_cible_puis_importer)

        btn_collec = QPushButton("Generer Collec-Science")
        btn_collec.setObjectName("btn_collec")
        btn_collec.setMinimumWidth(200)
        btn_collec.setFixedHeight(38)
        btn_collec.clicked.connect(self.generer_collec_science)

        btn_colisa_logiciel = QPushButton("Generer COLISA logiciel")
        btn_colisa_logiciel.setObjectName("btn_colisa_logiciel")
        btn_colisa_logiciel.setMinimumWidth(220)
        btn_colisa_logiciel.setFixedHeight(38)
        btn_colisa_logiciel.clicked.connect(self.generer_format_colisa_logiciel)

        btn_pipeline = QPushButton("Export complet")
        btn_pipeline.setObjectName("btn_pipeline")
        btn_pipeline.setMinimumWidth(180)
        btn_pipeline.setFixedHeight(38)
        btn_pipeline.setToolTip("Lance l'import, puis genere Collec-Science, puis genere le format COLISA logiciel")
        btn_pipeline.clicked.connect(self.generer_pipeline_complet)

        h.addWidget(btn_import)
        h.addWidget(btn_collec)
        h.addWidget(btn_colisa_logiciel)
        h.addWidget(btn_pipeline)
        h.addStretch()

        self.lbl_status = QLabel("Pret")
        self.lbl_status.setStyleSheet("font-weight: bold; padding: 8px;")
        h.addWidget(self.lbl_status)
        return w

    def _refresh_context_labels(self) -> None:
        source_label = "PAC final" if self.source_mode == "pac_final" else "Autre type de source"
        source_name = self.source_path.name if str(self.source_path).strip() else "Aucun fichier source"
        self.lbl_source.setText(f"Entree : {source_name}")
        self.lbl_target.setText(f"Sortie : {self.out_path.name}")
        self.lbl_paths.setText(f"Format : {source_label}")

    def _load_settings(self) -> None:
        if not SETTINGS_FILE.exists():
            return
        try:
            payload = json.loads(SETTINGS_FILE.read_text(encoding="utf-8"))
        except Exception:
            return

        # source_path : on ne restaure que si le fichier existe sur ce PC
        saved_src = payload.get("source_path")
        if isinstance(saved_src, str) and saved_src.strip():
            p = Path(saved_src)
            if p.exists():
                self.source_path = p

        src_sheet = payload.get("source_sheet")
        src_mode = payload.get("source_mode")
        src_mapping = payload.get("source_mapping")
        if isinstance(src_sheet, str) and src_sheet.strip():
            self.source_sheet = src_sheet.strip()
        if isinstance(src_mode, str) and src_mode.strip():
            self.source_mode = src_mode.strip()
        if isinstance(src_mapping, dict):
            self.source_mapping = src_mapping

        start_ref = payload.get("start_ref")
        end_ref = payload.get("end_ref")
        org = payload.get("default_org")
        country = payload.get("default_country")
        code_unite = payload.get("default_code_unite_gestionnaire")
        site_atelier = payload.get("default_site_atelier")
        num_correspondant = payload.get("default_numero_correspondant")
        dup_mode = payload.get("duplicate_mode")

        if isinstance(start_ref, str):
            self.ed_start.setText(start_ref)
        if isinstance(end_ref, str):
            self.ed_end.setText(end_ref)
        if isinstance(org, str):
            self.ed_org.setText(org)
        if isinstance(country, str):
            self.ed_country.setText(country)
        if isinstance(site_atelier, str) and site_atelier.strip():
            self.ed_site_atelier.setText(site_atelier)
        if isinstance(num_correspondant, str) and num_correspondant.strip():
            self.ed_num_correspondant.setText(num_correspondant)
        if isinstance(dup_mode, str) and self.cb_duplicate.findText(dup_mode) >= 0:
            self.cb_duplicate.setCurrentText(dup_mode)

        self.dark_mode = True
        self.ed_code_unite.setText("0042")
        self._reset_output_paths()

        saved_out_path = payload.get("out_path")
        saved_history_path = payload.get("history_path")
        saved_selection_csv = payload.get("selection_csv")
        # On ne restaure les chemins que s'ils sont accessibles sur ce PC
        if isinstance(saved_out_path, str) and saved_out_path.strip():
            p = Path(saved_out_path)
            if p.parent.exists():
                self.out_path = p
        if isinstance(saved_history_path, str) and saved_history_path.strip():
            p = Path(saved_history_path)
            if p.parent.exists():
                self.history_path = p
        if isinstance(saved_selection_csv, str) and saved_selection_csv.strip():
            p = Path(saved_selection_csv)
            if p.parent.exists():
                self.selection_csv = p

    def _reset_output_paths(self) -> None:
        """Force all generated files into the local portable output folder."""
        self.out_path = self.imports_dir / "COLISA en cours.xlsx"
        self.history_path = self.imports_dir / "import_history.json"
        self.selection_csv = self.imports_dir / "selection_import.csv"

    def _save_settings(self) -> None:
        try:
            SETTINGS_DIR.mkdir(parents=True, exist_ok=True)
            payload = {
                "source_path": str(self.source_path),
                "source_sheet": self.source_sheet,
                "source_mode": self.source_mode,
                "source_mapping": self.source_mapping,
                "out_path": str(self.out_path),
                "history_path": str(self.history_path),
                "selection_csv": str(self.selection_csv),
                "start_ref": self.ed_start.text().strip(),
                "end_ref": self.ed_end.text().strip(),
                "default_org": self.ed_org.text().strip(),
                "default_country": self.ed_country.text().strip(),
                "default_code_unite_gestionnaire": self.ed_code_unite.text().strip(),
                "default_site_atelier": self.ed_site_atelier.text().strip(),
                "default_numero_correspondant": self.ed_num_correspondant.text().strip(),
                "duplicate_mode": self.cb_duplicate.currentText(),
                "dark_mode": True,
            }
            SETTINGS_FILE.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception:
            return

    def _on_theme_toggled(self, checked: bool) -> None:
        self._apply_theme(True)
        self._save_settings()
        if hasattr(self, "lbl_status"):
            self.lbl_status.setText("Theme: nuit")

    def _update_app_icon(self) -> None:
        logo_path = _app_logo_path()
        if logo_path.exists():
            self.setWindowIcon(QIcon(str(logo_path)))
            return
        pix = self._build_logo_pixmap(self.dark_mode)
        self.setWindowIcon(QIcon(pix))

    def _build_logo_pixmap(self, dark_mode: bool) -> QPixmap:
        pix = QPixmap(256, 256)
        pix.fill(Qt.transparent)
        p = QPainter(pix)
        p.setRenderHint(QPainter.Antialiasing)

        bg = QColor("#0f172a") if dark_mode else QColor("#f8fafc")
        card = QColor("#111827") if dark_mode else QColor("#ffffff")
        border = QColor("#334155") if dark_mode else QColor("#94a3b8")
        accent = QColor("#22c55e") if dark_mode else QColor("#166534")
        glass = QColor("#374151") if dark_mode else QColor("#f8fafc")

        p.setPen(QPen(border, 1.2))
        p.setBrush(card)
        p.drawRoundedRect(12, 12, 232, 232, 28, 28)
        p.fillRect(14, 14, 228, 36, bg)

        p.setPen(QPen(border, 6))
        p.setBrush(QColor("#1f2937") if dark_mode else QColor("#e2e8f0"))
        p.drawRoundedRect(42, 58, 96, 140, 12, 12)

        p.setBrush(glass)
        for y in (72, 112, 152):
            p.drawRoundedRect(50, y, 80, 30, 8, 8)
            p.setBrush(accent)
            p.drawEllipse(114, y + 10, 8, 8)
            p.setBrush(glass)

        p.end()
        return pix

    def _apply_theme(self, dark_mode: bool) -> None:
        """Apply modern theme stylesheet."""
        self.dark_mode = True
        from presentation.styles import get_stylesheet
        self.setStyleSheet(get_stylesheet(dark_mode=True))
        self._update_app_icon()

    def show_about(self) -> None:
        date_str = dt.date.today().strftime("%d/%m/%Y")
        text = (
            "DATACOLISA\n"
            "Logiciel de gestion et d'import de donnees scientifiques\n\n"
            f"Version : {APP_VERSION}\n"
            f"Auteurs : {APP_AUTHOR}\n"
            f"Organisation : {APP_ORGANISATION}\n"
            f"Date : {date_str}\n\n"
            "Ce logiciel a ete conçu et realise par Anatole Aubin et Quentin Godeaux."
        )
        QMessageBox.information(self, "A propos", text)

    def show_manual(self) -> None:
        text = (
            "Manuel d'utilisation DATACOLISA\n\n"
            "1. Import principal\n"
            "- Choisir la source dans le menu Fichier.\n"
            "- Charger la plage.\n"
            "- Verifier les lignes dans le tableau.\n"
            "- Utiliser Actions multi-lignes ou Types individuels si besoin.\n"
            "- Lancer l'import pour remplir COLISA en cours.\n\n"
            "2. Generer format COLISA\n"
            "- Depuis le logiciel : utilise les lignes selectionnees dans le tableau.\n"
            "- Depuis un fichier Excel : le fichier choisi doit deja etre un fichier COLISA brut.\n"
            "- La fenetre de preparation montre le fichier qui va etre cree.\n"
            "- Tu peux cocher, modifier et completer les colonnes avant generation.\n\n"
            "3. Collec-Science\n"
            "- Peut partir du logiciel ou d'un fichier Excel selon le bouton choisi.\n\n"
            "4. Conseils\n"
            "- Si une ligne ne passe pas, regarder le statut et les erreurs.\n"
            "- Relancer l'application apres une mise a jour si un changement n'apparait pas."
        )
        QMessageBox.information(self, "Manuel d'utilisation", text)

    def _import_base_path(self) -> Path:
        return self.out_path if self.out_path.exists() else self._internal_target_base_path()

    def _internal_target_base_path(self) -> Path:
        return _app_data_dir() / "cache" / "COLISA_base_integree.xlsx"

    def _ensure_internal_target_base(self) -> Path:
        base_path = self._internal_target_base_path()

        openpyxl, _ = core.ensure_deps()
        return create_internal_target_workbook(base_path, openpyxl, template_path=None)

    def _ensure_output_initialized(self) -> Path | None:
        if self.out_path.exists():
            return self.out_path
        try:
            base_path = self._ensure_internal_target_base()
            self.out_path.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(base_path, self.out_path)
            return self.out_path
        except Exception:
            return None

    def _refresh_type_options(self) -> None:
        target = self._import_base_path()
        opts = core.load_type_echantillon_options(target)
        opts.extend(SAMPLE_TYPE_LABELS.keys())
        normalized_opts = {normalize_sample_type_code(code) for code in opts if str(code).strip()}
        normalized_opts.discard("EC MONTEE")
        normalized_opts.discard("EC BRUTE")
        normalized_opts.add("EC")
        self.type_options = sorted(normalized_opts)

        previous_value = self._current_bulk_type_value()
        self.bulk_type.clear()
        for code in self.type_options:
            label = SAMPLE_TYPE_LABELS.get(code, "").strip()
            display = f"{code} - {label}" if label else code
            self.bulk_type.addItem(display, code)
        if previous_value:
            index = self.bulk_type.findData(previous_value)
            if index >= 0:
                self.bulk_type.setCurrentIndex(index)
            else:
                self.bulk_type.setEditText(previous_value)
        if hasattr(self, "table") and self.table is not None:
            self._setup_table_delegates()

    def _current_bulk_type_value(self) -> str:
        data = self.bulk_type.currentData()
        if data is not None and str(data).strip():
            return normalize_sample_type_code(str(data).strip())
        text = self.bulk_type.currentText().strip()
        return normalize_sample_type_code(text)

    def select_source_file(self) -> None:
        start_dir = _default_source_start_dir(self.source_path if str(self.source_path).strip() else None, self.imports_dir)
        p, _ = QFileDialog.getOpenFileName(self, "Choisir source .xls", start_dir, "Excel (*.xls *.xlsx)")
        if p:
            self.source_path = Path(p)
            self._refresh_context_labels()
            self._save_settings()

    def configure_source_format(self) -> None:
        from presentation.source_format_dialog import SourceFormatDialog

        if not str(self.source_path).strip():
            QMessageBox.warning(self, "Source", "Choisis d'abord un fichier source.")
            return
        if not self.source_path.exists():
            QMessageBox.warning(self, "Source", f"Fichier source introuvable: {self.source_path}")
            return

        dialog = SourceFormatDialog(
            source_path=self.source_path,
            current_mode=self.source_mode,
            current_sheet=self.source_sheet,
            current_mapping=self.source_mapping,
            parent=self,
        )
        if dialog.exec() != QDialog.Accepted:
            return

        result = dialog.get_result()
        self.source_mode = str(result.get("mode") or "pac_final")
        if result.get("sheet_name"):
            self.source_sheet = str(result["sheet_name"])
        self.source_mapping = result.get("mapping", {}) or {}
        self._refresh_context_labels()
        self._save_settings()

    def select_other_excel_source(self) -> None:
        p, _ = QFileDialog.getOpenFileName(
            self,
            "Choisir un autre type de source",
            _default_source_start_dir(self.source_path if str(self.source_path).strip() else None, self.imports_dir),
            "Excel (*.xls *.xlsx)",
        )
        if not p:
            return

        self.source_path = Path(p)
        from presentation.source_format_dialog import SourceFormatDialog

        dialog = SourceFormatDialog(
            source_path=self.source_path,
            current_mode="custom",
            current_sheet=self.source_sheet,
            current_mapping=self.source_mapping,
            force_custom=True,
            parent=self,
        )
        if dialog.exec() != QDialog.Accepted:
            return

        result = dialog.get_result()
        self.source_mode = "custom"
        if result.get("sheet_name"):
            self.source_sheet = str(result["sheet_name"])
        self.source_mapping = result.get("mapping", {}) or {}
        self._refresh_context_labels()
        self._save_settings()

    def set_source_sheet(self) -> None:
        txt, ok = QInputDialog.getText(self, "Onglet source", "Nom onglet source:", text=self.source_sheet)
        if ok and txt.strip():
            self.source_sheet = txt.strip()
            self._refresh_context_labels()
            self._save_settings()

    def select_target_file(self) -> None:
        from presentation.dialogs import InfoDialog

        InfoDialog.show(
            self,
            "COLISA",
            "Aucun fichier externe n'est necessaire.",
        )

    def set_target_sheet(self) -> None:
        txt, ok = QInputDialog.getText(self, "Onglet cible", "Nom onglet cible:", text=self.target_sheet)
        if ok and txt.strip():
            self.target_sheet = txt.strip()
            self._refresh_context_labels()
            self._save_settings()

    def set_output_file(self) -> None:
        p, _ = QFileDialog.getSaveFileName(self, "Fichier sortie", str(self.out_path), "Excel (*.xlsx)")
        if p:
            self.out_path = Path(p)
            self._refresh_type_options()
            self._refresh_context_labels()
            self._save_settings()

    def set_history_file(self) -> None:
        p, _ = QFileDialog.getSaveFileName(self, "Fichier historique", str(self.history_path), "JSON (*.json)")
        if p:
            self.history_path = Path(p)
            self._refresh_context_labels()
            self._save_settings()

    def set_selection_csv(self) -> None:
        p, _ = QFileDialog.getSaveFileName(self, "Fichier selection CSV", str(self.selection_csv), "CSV (*.csv)")
        if p:
            self.selection_csv = Path(p)
            self._refresh_context_labels()
            self._save_settings()

    def refaire_fichier_import(self) -> None:
        """Recreate COLISA en cours from the built-in COLISA base."""
        from presentation.dialogs import ConfirmationDialog, InfoDialog

        if self.out_path.exists():
            confirmed = ConfirmationDialog.ask(
                self,
                "Refaire COLISA",
                "Le fichier COLISA en cours sera recree. Continuer ?",
            )
            if not confirmed:
                return

        try:
            base_path = self._ensure_internal_target_base()
            self.out_path.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(base_path, self.out_path)
            self._refresh_type_options()
            self._refresh_context_labels()
            self._save_settings()
            self.lbl_status.setText("COLISA en cours refait")
            InfoDialog.show(self, "COLISA", f"COLISA en cours a bien ete recree.\n\nEmplacement : {self.out_path}")
            # Demander le code échantillon de départ
            code, ok = QInputDialog.getText(
                self,
                "Code échantillon de départ",
                "Par quel code échantillon commencer ?\n(ex : T001, CA950...)",
                text=self.ed_start.text().strip(),
            )
            if ok and code.strip():
                self.ed_start.setText(code.strip().upper())
                self._save_settings()
        except Exception as exc:
            from presentation.dialogs import ErrorDialog
            ErrorDialog.show(self, "COLISA", exc)

    def closeEvent(self, event: Any) -> None:
        self._save_settings()
        super().closeEvent(event)

    def add_new_type(self) -> None:
        target = self._ensure_output_initialized()
        if target is None:
            QMessageBox.warning(self, "Type", "Impossible d'initialiser COLISA en cours.")
            return
        value, ok = QInputDialog.getText(self, "Nouveau type", "Code type echantillon:")
        if not ok or not value.strip():
            return
        if core.append_type_echantillon_option(target, value.strip()):
            self._refresh_type_options()
            self.lbl_status.setText(f"Type ajoute: {value.strip()}")
        else:
            QMessageBox.warning(self, "Type", "Impossible d'ajouter le type")

    def load_range(self) -> None:
        """
        Load and filter source data in background thread (UI responsive).

        OLD: Blocking operation = UI freeze for 10+ seconds with large files
        NEW: Background worker = UI stays responsive, can cancel operation
        """
        start_ref = self.ed_start.text().strip()
        end_ref = self.ed_end.text().strip()
        default_type = self._current_bulk_type_value() or "EC"

        if not start_ref or not end_ref:
            from presentation.dialogs import WarningDialog
            WarningDialog.show(self, "Chargement", "Veuillez saisir les références de début et fin")
            return
        if self.source_mode == "custom" and not (self.source_mapping.get("columns") or {}):
            from presentation.dialogs import WarningDialog
            WarningDialog.show(self, "Chargement", "Configure d'abord le format de l'autre Excel dans Fichier > Format source...")
            return

        # Disable UI during load
        self._set_ui_enabled(False)
        self.lbl_status.setText("Chargement en cours...")

        # Create and start worker
        from presentation.workers import LoadRangeWorker
        self.load_worker = LoadRangeWorker(
            source_path=self.source_path,
            source_sheet=self.source_sheet,
            source_mode=self.source_mode,
            source_mapping=self.source_mapping,
            start_ref=start_ref,
            end_ref=end_ref,
            default_type_echantillon=default_type
        )

        # Connect signals
        self.load_worker.finished.connect(self._on_load_finished)
        self.load_worker.error.connect(self._on_load_error)
        self.load_worker.progress.connect(self._on_load_progress)

        # Start background thread
        self.load_worker.start()

    def _lire_individus_fichier_sortie(self) -> None:
        """
        Lit le fichier de sortie COLISA et met a jour le cache des individus EC/OT deja importes.
        Colonnes utilisees (positions standard) :
          5  = Numero individu
          6  = Code type echantillon
          17 = Ecailles brutes
          18 = Montees
          20 = Otolithes
          23 = Empreintes
        """
        self._individus_sortie = {}
        if not self.out_path.exists():
            return
        try:
            import openpyxl
            from config.constants import DEFAULT_TARGET_SHEET
            wb = openpyxl.load_workbook(str(self.out_path), read_only=True, data_only=True)
            ws = wb[DEFAULT_TARGET_SHEET] if DEFAULT_TARGET_SHEET in wb.sheetnames else wb[wb.sheetnames[0]]

            # Lire l'en-tete pour trouver les colonnes dynamiquement
            col_map = {}
            CIBLES = {
                "numero individu": "num_individu",
                "code type echantillon": "code_type_echantillon",
                "ecailles brutes": "ecailles_brutes",
                "montees": "montees",
                "otolithes": "otolithes",
                "empreintes": "empreintes",
            }
            for cell in next(ws.iter_rows(min_row=1, max_row=1)):
                val = str(cell.value or "").strip().lower()
                for cle, champ in CIBLES.items():
                    if cle in val:
                        col_map[champ] = cell.column
                        break

            if "num_individu" not in col_map or "code_type_echantillon" not in col_map:
                wb.close()
                return

            for excel_row_num, data_row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                num_ind = str(data_row[col_map["num_individu"] - 1] or "").strip()
                code_type = str(data_row[col_map["code_type_echantillon"] - 1] or "").strip()
                if not num_ind or not code_type:
                    continue
                if not _famille_ec_ot(code_type):
                    continue

                def _val(champ):
                    idx = col_map.get(champ)
                    return str(data_row[idx - 1] or "").strip() if idx else ""

                self._individus_sortie[num_ind] = {
                    "row": str(excel_row_num),
                    "code_type_echantillon": code_type,
                    "montees":        _val("montees"),
                    "ecailles_brutes": _val("ecailles_brutes"),
                    "otolithes":      _val("otolithes"),
                    "empreintes":     _val("empreintes"),
                }
            wb.close()
        except Exception as exc:
            self.logger.warning(f"Lecture fichier sortie pour doublons EC/OT : {exc}")

    def _on_load_finished(self, result: Dict[str, Any]) -> None:
        """Handle load completion."""
        from presentation.dialogs import InfoDialog

        # Mettre a jour le cache des individus deja dans le fichier de sortie
        self._lire_individus_fichier_sortie()

        # Fusionner automatiquement les lignes EC et OT du meme individu
        rows_brutes = result["rows"]
        self.rows = fusionner_lignes_ec_ot(rows_brutes)
        self.missing_codes = result["missing_codes"]

        # Update UI
        self._render_table()
        self.lbl_count.setText(f"Lignes: {result['found_count']}")
        self.lbl_pending.setText(f"A reimporter: {result['pending_count']}")
        self.lbl_missing.setText(f"Codes manquants: {len(self.missing_codes)}")
        self.lbl_status.setText(f"Lignes chargees: {len(self.rows)}")

        # Re-enable UI
        self._set_ui_enabled(True)

        missing_date_count = int(result.get("missing_date_count", 0) or 0)
        if missing_date_count > 0:
            InfoDialog.show(
                self,
                "Chargement",
                f"{missing_date_count} ligne(s) sans date ont ete decochees automatiquement pour l'import.",
            )

    def _on_load_error(self, error_msg: str) -> None:
        """Handle load error."""
        from presentation.dialogs import ErrorDialog
        ErrorDialog.show(self, "Chargement", error_msg)
        self.lbl_status.setText("Erreur de chargement")
        self._set_ui_enabled(True)

    def _on_load_progress(self, current: int, total: int) -> None:
        """Handle load progress updates."""
        self.lbl_status.setText(f"Chargement... ({current}/{total})")

    def _colonne_cible_ec_ot(self, row_data: Dict[str, Any], existant: Dict[str, str]) -> str:
        """Retourne la colonne a completer dans le fichier de sortie pour un cas EC/OT deja present."""
        code_type = str(row_data.get("code_type_echantillon", "") or "").strip().upper()
        if "OT" in code_type or "OTOLITH" in code_type:
            return "otolithes"
        if "BRUTE" in code_type:
            return "ecailles brutes"
        if "MONTEE" in code_type:
            return "montees"

        for champ, libelle in (
            ("montees", "montees"),
            ("ecailles_brutes", "ecailles brutes"),
            ("otolithes", "otolithes"),
            ("empreintes", "empreintes"),
        ):
            val = str(row_data.get(champ, "") or "").strip()
            if val and val != "0":
                return libelle

        famille = _famille_ec_ot(code_type)
        if famille == "OT":
            return "otolithes"
        return "montees"

    def _preparer_alertes_ec_ot_import(self, rows: List[Dict[str, Any]]) -> tuple[List[Dict[str, Any]], List[str]]:
        """
        Repere les lignes EC/OT deja presentes dans le fichier de sortie.
        Ces lignes sont retirees de l'import et un message indique ou completer manuellement.
        """
        self._lire_individus_fichier_sortie()
        lignes_import: List[Dict[str, Any]] = []
        alertes: List[str] = []

        for row in rows:
            row_copy = dict(row)
            if not bool(row_copy.get("selected")):
                lignes_import.append(row_copy)
                continue

            famille = _famille_ec_ot(row_copy.get("code_type_echantillon", ""))
            num_ind = str(row_copy.get("num_individu", "") or "").strip()
            existant = self._individus_sortie.get(num_ind)
            if famille and num_ind and existant:
                ligne_existante = str(existant.get("row", "") or "").strip() or "?"
                colonne = self._colonne_cible_ec_ot(row_copy, existant)
                type_existant = str(existant.get("code_type_echantillon", "") or "").strip() or "EC/OT"
                alertes.append(
                    f"{num_ind} -> ligne {ligne_existante} -> rajouter dans la colonne {colonne} (type actuel : {type_existant})"
                )
                row_copy["selected"] = False
            lignes_import.append(row_copy)

        return lignes_import, alertes

    def _show_ec_ot_import_dialog(self, alertes: List[str]) -> None:
        """Affiche les lignes EC/OT a completer manuellement avant import."""
        from PySide6.QtWidgets import QDialog, QVBoxLayout, QLabel, QTextEdit, QPushButton, QHBoxLayout

        dlg = QDialog(self)
        dlg.setWindowTitle("Lignes EC/OT a completer")
        dlg.setMinimumWidth(620)
        dlg.setMinimumHeight(380)

        layout = QVBoxLayout(dlg)
        titre = QLabel(f"{len(alertes)} ligne(s) ne seront pas importee(s).")
        titre.setStyleSheet("font-weight: bold; font-size: 12pt;")
        titre.setWordWrap(True)
        layout.addWidget(titre)

        sub = QLabel(
            "Ces echantillons existent deja dans le fichier de sortie. "
            "Rajoute-les manuellement sur la ligne indiquee, dans la bonne colonne."
        )
        sub.setWordWrap(True)
        layout.addWidget(sub)

        txt = QTextEdit()
        txt.setReadOnly(True)
        txt.setPlainText("\n".join(alertes))
        layout.addWidget(txt)

        btns = QHBoxLayout()
        btns.addStretch()
        btn_ok = QPushButton("OK - Compris")
        btn_ok.clicked.connect(dlg.accept)
        btns.addWidget(btn_ok)
        layout.addLayout(btns)
        dlg.exec()

    def _set_ui_enabled(self, enabled: bool) -> None:
        """Enable/disable UI controls during background operations."""
        # This is a placeholder - specific widgets can be enabled/disabled as needed
        # For now, we don't disable everything to allow cancel operations in the future
        pass

    def show_missing_codes(self) -> None:
        """Show missing codes dialog."""
        from presentation.dialogs import MissingCodesDialog
        MissingCodesDialog.show(self, self.missing_codes)

    def _render_table(self) -> None:
        """
        Render table using model (performance optimized).

        OLD METHOD: Created QComboBox for every cell in combo columns (700+ widgets for 100 rows!)
        NEW METHOD: Model stores data, delegates create combo boxes only during editing.

        Performance: 100+ rows render in <1s instead of 10s+
        """
        self._table_sync_guard = True

        # Simply pass data to model - model handles all rendering efficiently
        self.table_model.set_rows(self.rows)

        self._table_sync_guard = False

    def _on_table_data_changed(self, top_left: QModelIndex, bottom_right: QModelIndex, roles: List[int]) -> None:
        """Handle data changes in the table model."""
        if self._table_sync_guard:
            return

        # Sync model changes back to self.rows
        self.rows = self.table_model.get_rows()

        # Handle column-specific business rules
        from presentation.table_model import COLUMNS
        otolithes_col = COLUMNS.index("otolithes")

        self._table_sync_guard = True  # Prevent recursion

        fusion_declenchee = False

        for row in range(top_left.row(), bottom_right.row() + 1):
            row_data = self.table_model.get_row(row)
            if not row_data:
                continue

            modified = False

            # Handle "otolithes" column changes (sync autre_oss)
            if top_left.column() <= otolithes_col <= bottom_right.column():
                ot = core.normalize(row_data.get("otolithes", ""))
                autre_oss = "OUI" if (ot and ot != "0") else "NON"
                if row_data.get("autre_oss") != autre_oss:
                    row_data["autre_oss"] = autre_oss
                    modified = True

            if modified:
                self.table_model.update_row(row, row_data)

        self._table_sync_guard = False

        # Si un doublon EC/OT a ete detecte, relancer la fusion sur tout le tableau
        if fusion_declenchee:
            self._table_sync_guard = True
            rows_fusionnes = fusionner_lignes_ec_ot(self.rows)
            self.rows = rows_fusionnes
            self.table_model.set_rows(self.rows)
            self._table_sync_guard = False

    # NOTE: _set_selected_item_enabled and _on_table_item_changed removed
    # (were for QTableWidget, now handled by model/delegate pattern)

    def _read_table(self) -> List[Dict[str, Any]]:
        """
        Read table data from model.

        With QTableView + Model, we just get data directly from model (no widget traversal needed).
        """
        # Get data from model
        rows = self.table_model.get_rows()

        # Apply business rule: autre_oss based on otolithes
        for row in rows:
            ot = core.normalize(row.get("otolithes", ""))
            row["autre_oss"] = "OUI" if (ot and ot != "0") else "NON"

        self.rows = rows
        return rows

    def _selected_row_indexes(self) -> List[int]:
        return [i for i, r in enumerate(self._read_table()) if bool(r.get("selected"))]

    def _active_row_indexes(self) -> List[int]:
        indexes = []
        if hasattr(self, "table") and self.table is not None:
            indexes = sorted({index.row() for index in self.table.selectionModel().selectedRows()})
        return indexes or self._selected_row_indexes()

    def _select_all(self, value: bool) -> None:
        """Select or deselect all rows."""
        rows = self.table_model.get_rows()
        for row in rows:
            row["selected"] = value

        # Update model with modified rows
        self.table_model.set_rows(rows)
        self.rows = rows

    # NOTE: _sync_autre_oss_from_otolithes removed (now handled automatically in _on_table_data_changed)

    def apply_bulk(self) -> None:
        """Apply bulk operations to selected rows (model-based)."""
        idxs = self._selected_row_indexes()
        if not idxs:
            from presentation.dialogs import InfoDialog
            InfoDialog.show(self, "Selection", "Aucune ligne selectionnee")
            return

        rows = self.table_model.get_rows()

        for r in idxs:
            if r >= len(rows):
                continue

            row_data = rows[r]

            # Update fields
            for key, val in [
                ("code_type_echantillon", self._current_bulk_type_value()),
                ("categorie", self.bulk_categorie.currentText()),
                ("type_peche", self.bulk_type_peche.currentText()),
                ("autre_oss", self.bulk_autre.currentText()),
                ("ecailles_brutes", self.bulk_ecailles.currentText()),
                ("montees", self.bulk_montees.currentText()),
                ("empreintes", self.bulk_empreintes.currentText()),
                ("otolithes", self.bulk_otolithes.currentText()),
                ("observation_disponibilite", self.bulk_observation.currentText()),
            ]:
                if val:
                    row_data[key] = val

            # Auto-sync autre_oss from otolithes
            ot = core.normalize(row_data.get("otolithes", ""))
            row_data["autre_oss"] = "OUI" if (ot and ot != "0") else "NON"

        # Update model with modified rows
        self.table_model.set_rows(rows)
        self.rows = rows

        self.lbl_status.setText(f"Bulk applique sur {len(idxs)} ligne(s)")

    def assign_sample_types_individually(self) -> None:
        """Open a dialog to assign sample types row by row."""
        from presentation.dialogs import InfoDialog
        from presentation.sample_type_assignment_dialog import SampleTypeAssignmentDialog

        idxs = self._active_row_indexes()
        if not idxs:
            InfoDialog.show(self, "Types", "Selectionne au moins une ligne dans le tableau ou coche la colonne Selection.")
            return

        rows = self.table_model.get_rows()
        selected_rows = [dict(rows[i]) for i in idxs if 0 <= i < len(rows)]

        dlg = SampleTypeAssignmentDialog(selected_rows, self.type_options, parent=self)
        if dlg.exec() != QDialog.Accepted:
            return

        updated_rows = dlg.get_updated_rows()
        for source_index, updated_row in zip(idxs, updated_rows):
            if 0 <= source_index < len(rows):
                rows[source_index]["code_type_echantillon"] = normalize_sample_type_code(
                    updated_row.get("code_type_echantillon", "")
                )

        self.table_model.set_rows(rows)
        self.rows = rows
        self.lbl_status.setText(f"Type echantillon mis a jour sur {len(updated_rows)} ligne(s)")

    def save_csv(self) -> None:
        try:
            data = self._read_table()
            out = self.selection_csv
            out.parent.mkdir(parents=True, exist_ok=True)

            headers = [
                "selected", "status", "ref", "num_individu", "date_capture", "code_espece",
                "lac_riviere", "pays_capture", "pecheur", "categorie", "type_peche", "observation_disponibilite", "autre_oss",
                "ecailles_brutes", "montees", "empreintes", "otolithes", "longueur_mm", "poids_g", "maturite", "sexe",
                "age_total", "code_type_echantillon",
                "sous_espece", "nom_operateur", "lieu_capture", "maille_mm", "code_stade",
                "presence_otolithe_gauche", "presence_otolithe_droite", "nb_opercules",
                "information_stockage", "age_riviere", "age_lac", "nb_fraie",
                "ecailles_regenerees", "observations",
                "errors",
            ]
            with out.open("w", newline="", encoding="utf-8") as f:
                w = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
                w.writeheader()
                for row in data:
                    payload = dict(row)
                    payload["selected"] = "1" if bool(payload.get("selected")) else "0"
                    for key, value in list(payload.items()):
                        payload[key] = coerce_internal_value(str(key), value)
                    w.writerow(payload)

            self.lbl_status.setText("Preparation des donnees terminee")
        except Exception as exc:
            from presentation.dialogs import ErrorDialog
            ErrorDialog.show(self, "Preparation", exc)

    def _choisir_cible_puis_importer(self) -> None:
        """Demande le fichier COLISA cible avant de lancer l'import."""
        from PySide6.QtWidgets import QMessageBox
        from presentation.dialogs import ErrorDialog

        box = QMessageBox(self)
        box.setWindowTitle("Lancer import")
        box.setText("Dans quel fichier COLISA importer ?")
        btn_existant = box.addButton("Fichier existant", QMessageBox.AcceptRole)
        btn_nouveau = box.addButton("Nouveau fichier", QMessageBox.DestructiveRole)
        box.addButton("Annuler", QMessageBox.RejectRole)
        box.exec()
        clicked = box.clickedButton()
        if clicked is None or clicked not in (btn_existant, btn_nouveau):
            return

        if clicked == btn_existant:
            p, _ = QFileDialog.getOpenFileName(
                self,
                "Choisir le fichier COLISA cible",
                str(self.imports_dir),
                "Excel (*.xlsx)",
            )
            if not p:
                return
            self.out_path = Path(p)
        else:
            p, _ = QFileDialog.getSaveFileName(
                self,
                "Créer un nouveau fichier COLISA",
                str(self.imports_dir / "COLISA en cours.xlsx"),
                "Excel (*.xlsx)",
            )
            if not p:
                return
            dest = Path(p)
            if dest.suffix.lower() != ".xlsx":
                dest = dest.with_suffix(".xlsx")
            if dest.exists():
                from presentation.dialogs import WarningDialog
                WarningDialog.show(self, "Lancer import", f"Ce fichier existe déjà et ne sera pas écrasé :\n{dest}\n\nChoisis un autre nom.")
                return
            try:
                base_path = self._ensure_internal_target_base()
                dest.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(base_path, dest)
                self.out_path = dest
            except Exception as exc:
                ErrorDialog.show(self, "Lancer import", f"Impossible de créer le fichier :\n{exc}")
                return

        self._refresh_context_labels()
        self._save_settings()
        self.run_import()

    def _lire_max_code_echantillon(self) -> int:
        """Lit le numéro max du code_echantillon dans out_path. Retourne 0 si vide ou absent."""
        if not self.out_path.exists():
            return 0
        try:
            import re as _re
            import unicodedata as _uni
            openpyxl, _ = core.ensure_deps()
            wb = openpyxl.load_workbook(str(self.out_path), read_only=True, data_only=True)
            ws = wb.active
            pattern = _re.compile(r"^[A-Za-z]*(\d+)$")

            def _norm_hdr(s: object) -> str:
                txt = str(s or "").strip().lower()
                txt = _uni.normalize("NFKD", txt)
                txt = "".join(ch for ch in txt if not _uni.combining(ch))
                return _re.sub(r"[^a-z0-9 _]+", "", txt).strip()

            code_col = None
            header_row_idx = None
            for r_idx, row in enumerate(ws.iter_rows(max_row=20, values_only=True), start=1):
                for c_idx, val in enumerate(row, start=1):
                    if _norm_hdr(val) in ("code echantillon", "code_echantillon"):
                        code_col = c_idx
                        header_row_idx = r_idx
                        break
                if code_col:
                    break
            max_num = 0
            if code_col and header_row_idx:
                for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                    val = str(row[code_col - 1] or "").strip()
                    m = pattern.match(val)
                    if m:
                        max_num = max(max_num, int(m.group(1)))
            wb.close()
            return max_num
        except Exception:
            return 0

    def _demander_numero_depart(self) -> int:
        """Retourne le numéro de départ pour les codes_echantillon.

        - Fichier existant (out_path existe) : toujours 0 → init_sequence_from_workbook
          lit le max dans le fichier sans demander quoi que ce soit.
        - Fichier vide/nouveau (out_path n'existe pas encore, ou header introuvable) :
          demande le numéro de départ à l'utilisateur.
        Retourne -1 si l'utilisateur annule.
        """
        # Si le fichier existe déjà, on laisse init_sequence_from_workbook trouver le max
        if self.out_path.exists():
            return 0

        # Nouveau fichier vide → demander par quel code commencer
        from PySide6.QtWidgets import QInputDialog
        text, ok = QInputDialog.getText(
            self,
            "Code echantillon de depart",
            "Nouveau COLISA vide.\nPar quel code echantillon commencer ?\n\n"
            "Entrez le numero (ex: si vous commencez a T59863, entrez 59863) :",
            text="1",
        )
        if not ok:
            return -1  # annulé
        # Accepter aussi un code complet type "T59863" → extraire le numero
        raw = text.strip()
        digits = "".join(c for c in raw if c.isdigit())
        try:
            return max(1, int(digits)) if digits else -1
        except ValueError:
            from presentation.dialogs import WarningDialog
            WarningDialog.show(self, "Code de depart", f"Valeur invalide : '{text}'")
            return -1

    def run_import(self) -> None:
        """
        Run import in background thread (UI responsive).

        OLD: Blocking operation with StringIO hack = UI freeze
        NEW: Background worker = UI stays responsive, proper result handling
        """
        try:
            # Validate
            data = self._read_table()
            data, alertes_ec_ot = self._preparer_alertes_ec_ot_import(data)
            selected_count = sum(1 for r in data if bool(r.get("selected")))
            if selected_count == 0:
                if alertes_ec_ot:
                    self._show_ec_ot_import_dialog(alertes_ec_ot)
                    self.lbl_status.setText("Import bloque pour lignes EC/OT deja presentes")
                from presentation.dialogs import WarningDialog
                WarningDialog.show(self, "Import", "Aucune ligne n'est selectionnee pour l'import.")
                return

            if alertes_ec_ot:
                self._show_ec_ot_import_dialog(alertes_ec_ot)

            output_already_exists = self.out_path.exists()

            start_numero = self._import_start_numero
            if start_numero == 0 and not output_already_exists:
                start_numero = self._demander_numero_depart()
                if start_numero < 0:
                    from presentation.dialogs import WarningDialog
                    WarningDialog.show(self, "Import annule", "Import annule : aucun numero de depart choisi.")
                    return

            target = self._ensure_output_initialized()
            if target is None:
                from presentation.dialogs import WarningDialog
                WarningDialog.show(self, "Import", "Impossible d'initialiser COLISA en cours.")
                return
            self._import_start_numero = 0  # reset

            # Disable UI during import
            self._set_ui_enabled(False)
            self.lbl_status.setText("Import en cours...")

            # Create and start worker
            from presentation.workers import ImportWorker
            self.import_worker = ImportWorker(
                selection_csv=self.selection_csv,
                selection_rows=data,
                target_path=Path(target),
                target_sheet=self.target_sheet,
                out_target=self.out_path,
                history_path=self.history_path,
                default_organisme=self.ed_org.text(),
                default_country=self.ed_country.text(),
                default_code_unite_gestionnaire=self.ed_code_unite.text(),
                default_site_atelier=self.ed_site_atelier.text(),
                default_numero_correspondant=self.ed_num_correspondant.text(),
                on_duplicate=self.cb_duplicate.currentText(),
                start_numero=start_numero,
            )

            # Connect signals
            self.import_worker.finished.connect(self._on_import_finished)
            self.import_worker.error.connect(self._on_import_error)
            self.import_worker.progress.connect(self._on_import_progress)

            # Start background thread
            self.import_worker.start()

        except Exception as exc:
            from presentation.dialogs import ErrorDialog
            ErrorDialog.show(self, "Import", exc)
            self._set_ui_enabled(True)

    def _on_import_finished(self, result: Dict[str, Any]) -> None:
        """Handle import completion."""
        try:
            duplicates = int(result.get("duplicates", 0) or 0)
            imported = int(result.get("imported", 0) or 0)
            issues = sum(int(result.get(key, 0) or 0) for key in ("duplicates", "skipped_validation", "skipped_manual"))

            # Avertissement doublons AVANT le dialog de résultat
            if duplicates > 0:
                duplicate_refs = result.get("duplicate_refs", [])
                self._show_doublons_dialog(duplicate_refs)

            from presentation.dialogs import ImportResultDialog
            ImportResultDialog.show(self, result)
            if imported > 0 and issues == 0:
                self.lbl_status.setText("Import reussi sans probleme")
            elif duplicates > 0 and imported == 0:
                self.lbl_status.setText(f"{duplicates} doublon(s) detecte(s) - aucune nouvelle ligne")
            elif imported > 0:
                self.lbl_status.setText("Import termine avec alertes")
            else:
                self.lbl_status.setText("Import termine sans nouvelle ligne")
        except Exception as exc:
            import traceback
            from presentation.dialogs import ErrorDialog
            ErrorDialog.show(self, "Import resultat", str(exc) + "\n" + traceback.format_exc())
        finally:
            self._set_ui_enabled(True)

        if self._post_import_pipeline:
            self._post_import_pipeline = False
            import traceback
            from presentation.dialogs import ErrorDialog
            # Etape 2 : Collect-Science - depuis le fichier COLISA genere (contient les numero_identification)
            try:
                self.lbl_status.setText("Pipeline : génération Collect-Science...")
                self._pipeline_collec_science_depuis_colisa()
            except Exception as exc:
                ErrorDialog.show(self, "Pipeline – Collect-Science", str(exc) + "\n" + traceback.format_exc())
                self.lbl_status.setText("Erreur pipeline Collect-Science")
            # Étape 3 : COLISA logiciel (toujours tenté même si collect-science a échoué)
            try:
                self.lbl_status.setText("Pipeline : génération COLISA logiciel...")
                self._generer_colisa_logiciel_depuis_logiciel()
            except Exception as exc:
                ErrorDialog.show(self, "Pipeline – COLISA logiciel", str(exc) + "\n" + traceback.format_exc())
                self.lbl_status.setText("Erreur pipeline COLISA logiciel")

    def _show_doublons_dialog(self, duplicate_refs: list) -> None:
        """Affiche un dialog clair listant les numeros en double, un par ligne."""
        from PySide6.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout,
                                        QLabel, QTextEdit, QPushButton, QFrame)
        from PySide6.QtGui import QFont

        dark = self.dark_mode
        bg_header  = "#8b2a00" if dark else "#c0440a"   # orange brûlé
        bg_dialog  = "#1a0a00" if dark else "#fff5f0"   # fond saumon très clair
        bg_list    = "#120600" if dark else "#ffffff"
        txt_header = "#ffffff"
        txt_body   = "#e8c8b8" if dark else "#3a1000"
        border_col = "#6b1800" if dark else "#e06030"

        dlg = QDialog(self)
        dlg.setWindowTitle("Echantillons deja importes")
        dlg.setMinimumWidth(440)
        dlg.setMinimumHeight(380)
        dlg.setStyleSheet(f"""
            QDialog {{ background-color: {bg_dialog}; }}
            QLabel  {{ color: {txt_body}; }}
            QTextEdit {{
                background-color: {bg_list};
                color: {"#ff9966" if dark else "#8b1a00"};
                border: 2px solid {border_col};
                border-radius: 6px;
                padding: 10px;
                font-family: 'Courier New', monospace;
                font-size: 13pt;
                font-weight: bold;
            }}
            QPushButton {{
                background-color: {bg_header};
                color: white;
                border-radius: 8px;
                padding: 7px 20px;
                font-weight: 700;
                font-size: 10pt;
            }}
            QPushButton:hover {{ background-color: {"#6b1800" if dark else "#8b2a00"}; }}
        """)

        layout = QVBoxLayout(dlg)
        layout.setSpacing(0)
        layout.setContentsMargins(0, 0, 0, 14)

        # ── Bandeau d'en-tête coloré ─────────────────────────────────────────
        header = QFrame()
        header.setStyleSheet(f"background-color: {bg_header}; border-radius: 0px;")
        header_lay = QVBoxLayout(header)
        header_lay.setContentsMargins(18, 14, 18, 14)

        titre = QLabel(f"  {len(duplicate_refs)} echantillon(s) deja present(s) dans le COLISA")
        font_titre = QFont()
        font_titre.setPointSize(13)
        font_titre.setBold(True)
        titre.setFont(font_titre)
        titre.setStyleSheet(f"color: {txt_header}; background: transparent;")
        titre.setWordWrap(True)
        header_lay.addWidget(titre)

        sub = QLabel("Ces numeros sont deja dans le fichier — barre-les sur ta liste physique.")
        sub.setStyleSheet(f"color: #ffd0b8; background: transparent; font-size: 9pt;")
        sub.setWordWrap(True)
        header_lay.addWidget(sub)
        layout.addWidget(header)

        # ── Liste des numéros ────────────────────────────────────────────────
        inner = QVBoxLayout()
        inner.setContentsMargins(14, 12, 14, 0)
        inner.setSpacing(8)

        txt = QTextEdit()
        txt.setReadOnly(True)
        lines = "\n".join(str(r) for r in duplicate_refs) if duplicate_refs else "(aucun numero)"
        txt.setPlainText(lines)
        inner.addWidget(txt)

        # Bouton OK
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        btn_ok = QPushButton("OK - Compris")
        btn_ok.setMinimumWidth(130)
        btn_ok.setFixedHeight(36)
        btn_ok.clicked.connect(dlg.accept)
        btn_layout.addWidget(btn_ok)
        inner.addLayout(btn_layout)
        layout.addLayout(inner)

        dlg.exec()

    def _on_import_error(self, error_msg: str) -> None:
        """Handle import error."""
        from presentation.dialogs import ErrorDialog
        ErrorDialog.show(self, "Import", error_msg)
        self.lbl_status.setText("Erreur d'import")
        self._set_ui_enabled(True)

    def _on_import_progress(self, current: int, total: int) -> None:
        """Handle import progress updates."""
        self.lbl_status.setText(f"Import... ({current}/{total})")

    def generer_collec_science(self) -> None:
        """Generate Collect-Science from a selected Excel file."""
        self.generer_collec_science_depuis_excel()

    def _choose_collec_science_source(self) -> str | None:
        from PySide6.QtWidgets import QMessageBox

        data = self._read_table()
        has_current = any(bool(r.get("selected")) for r in data)

        box = QMessageBox(self)
        box.setWindowTitle("Collec-Science")
        box.setText("Choisir la source des echantillons")
        box.setInformativeText("Tu peux utiliser les echantillons du logiciel ou choisir un fichier Excel.")

        btn_current = None
        if has_current:
            btn_current = box.addButton("Depuis le logiciel", QMessageBox.ActionRole)
        btn_excel = box.addButton("Depuis un fichier Excel", QMessageBox.ActionRole)
        btn_cancel = box.addButton("Annuler", QMessageBox.RejectRole)
        box.exec()

        clicked = box.clickedButton()
        if clicked == btn_current:
            return "current"
        if clicked == btn_excel:
            return "excel"
        if clicked == btn_cancel:
            return None
        return None

    def _choose_colisa_logiciel_source(self) -> str | None:
        from PySide6.QtWidgets import QMessageBox

        data = self._read_table()
        has_current = any(bool(r.get("selected")) for r in data)

        box = QMessageBox(self)
        box.setWindowTitle("Generer format COLISA")
        box.setText("Choisir les echantillons a utiliser")
        box.setInformativeText(
            "Ce bouton sert uniquement a preparer le fichier COLISA a generer, "
            "depuis le tableau du logiciel ou depuis un fichier Excel deja au format COLISA brut."
        )

        btn_current = None
        if has_current:
            btn_current = box.addButton("Depuis le tableau du logiciel", QMessageBox.ActionRole)
        btn_excel = box.addButton("Depuis un fichier COLISA brut", QMessageBox.ActionRole)
        btn_cancel = box.addButton("Annuler", QMessageBox.RejectRole)
        box.exec()

        clicked = box.clickedButton()
        if clicked == btn_current:
            return "current"
        if clicked == btn_excel:
            return "excel"
        if clicked == btn_cancel:
            return None
        return None

    def _generer_collec_science_depuis_logiciel(self) -> None:
        """Generate Collect-Science from samples currently loaded in the software."""
        self._flush_table_edits()

        data = self._read_table()
        selected_rows = [r for r in data if bool(r.get("selected"))]
        if not selected_rows:
            from presentation.dialogs import WarningDialog
            WarningDialog.show(self, "Collec-Science", "Selectionne au moins un echantillon dans le tableau.")
            return

        lignes_sans_date = []
        lignes_sans_espece = []
        for r in selected_rows:
            ref = r.get("ref", "?")
            if not r.get("date_capture"):
                lignes_sans_date.append(ref)
            if not r.get("code_espece"):
                lignes_sans_espece.append(ref)

        forcer_anomalies = False
        if lignes_sans_date or lignes_sans_espece:
            from presentation.collec_science_dialog import AnomaliesDialog
            dlg_anom = AnomaliesDialog(
                parent=self,
                sans_date=lignes_sans_date,
                sans_espece=lignes_sans_espece,
            )
            if dlg_anom.exec() != QDialog.Accepted:
                return
            forcer_anomalies = dlg_anom.forcer_injection

        # Mode logiciel : génération directe depuis les rows, sans passer par COLISA
        # On utilise out_path uniquement pour proposer un nom de fichier de sortie par défaut
        self._run_collec_science_export(
            source_excel=self.out_path,
            forcer_anomalies=forcer_anomalies,
            software_rows=selected_rows,
            output_name_hint="echantillons_selectionnes",
            source_label=f"{len(selected_rows)} echantillon(s) selectionne(s) depuis le logiciel",
        )

    def _pipeline_collec_science_depuis_colisa(self) -> None:
        """Étape pipeline : génère Collect-Science depuis le fichier COLISA fraîchement importé (self.out_path).
        Contrairement à _generer_collec_science_depuis_logiciel, utilise le fichier Excel afin que
        le champ numero_identification (LT...) généré lors de l'import soit disponible."""
        from presentation.dialogs import WarningDialog
        if not self.out_path.exists():
            WarningDialog.show(
                self,
                "Pipeline – Collect-Science",
                f"Le fichier COLISA est introuvable :\n{self.out_path}",
            )
            return
        self._run_collec_science_export(
            source_excel=self.out_path,
            output_name_hint=self.out_path.stem,
            source_label=f"Fichier COLISA : {self.out_path.name}",
        )

    def _resolve_collec_science_source_excel(self, selected_rows: List[Dict[str, Any]]) -> Path | None:
        if self.out_path.exists():
            return self.out_path
        return self._build_collec_science_temp_source(selected_rows)

    def _build_collec_science_temp_source(self, selected_rows: List[Dict[str, Any]]) -> Path | None:
        try:
            from application.import_service import ImportService
            from domain.models import ImportConfig
            from infrastructure.csv_repository import CSVRepository
            from infrastructure.excel_reader import ExcelReader
            from infrastructure.excel_writer import ExcelWriter
            from infrastructure.history_repository import HistoryRepository

            openpyxl, xlrd = core.ensure_deps()
            base_path = self._ensure_internal_target_base()
            temp_source = _app_data_dir() / "COLISA_collect_science_temp.xlsx"
            temp_history = _app_data_dir() / "collect_science_temp_history.json"

            service = ImportService(
                excel_reader=ExcelReader(xlrd),
                excel_writer=ExcelWriter(openpyxl),
                csv_repo=CSVRepository(),
                history_repo=HistoryRepository(),
            )
            config = ImportConfig(
                selection_csv=self.selection_csv,
                target_path=base_path,
                target_sheet=self.target_sheet,
                output_path=temp_source,
                history_path=temp_history,
                default_organisme=self.ed_org.text().strip(),
                default_country=self.ed_country.text().strip(),
                on_duplicate="ignore",
                default_code_unite_gestionnaire=self.ed_code_unite.text().strip(),
                default_site_atelier=self.ed_site_atelier.text().strip(),
                default_numero_correspondant=self.ed_num_correspondant.text().strip(),
                selection_rows=selected_rows,
            )
            service.import_selection(config)
            return temp_source
        except Exception:
            return None

    def _flush_table_edits(self) -> None:
        """Force any in-progress cell edit to be committed before reading the table."""
        focused_widget = QApplication.focusWidget()
        if focused_widget is not None:
            focused_widget.clearFocus()
        if hasattr(self, "table") and self.table is not None:
            self.table.clearFocus()
        QApplication.processEvents()

    def _build_collec_science_type_counts(self, rows: List[Dict[str, Any]]) -> tuple[list[str], Dict[str, int]]:
        def _sample_value_present(value: Any) -> bool:
            if value is None:
                return False
            if isinstance(value, bool):
                return bool(value)
            normalized = core.normalize(value)
            if not normalized:
                return False
            if normalized.upper() in {"NON", "NO", "FALSE", "FAUX", "N", "0"}:
                return False
            if normalized.upper() in {"OUI", "YES", "TRUE", "VRAI", "Y", "X"}:
                return True
            try:
                return float(normalized.replace(",", ".")) >= 1
            except (ValueError, TypeError):
                return True

        active_types = []
        sample_counts: Dict[str, int] = {}
        for key, col in {"ecailles_brutes": "ecailles_brutes", "montees": "montees", "empreintes": "empreintes", "otolithes": "otolithes"}.items():
            count = 0
            for r in rows:
                if _sample_value_present(r.get(col)):
                    count += 1
            if count > 0:
                active_types.append(key)
                sample_counts[key] = count
        return active_types, sample_counts

    def _run_collec_science_export(
        self,
        source_excel: Path | None,
        active_types: list[str] | None = None,
        sample_counts: Dict[str, int] | None = None,
        forcer_anomalies: bool = False,
        allowed_num_individus: set[str] | None = None,
        software_rows: list | None = None,
        output_name_hint: str | None = None,
        source_label: str | None = None,
    ) -> None:
        from presentation.collec_science_dialog import CollecScienceSourceDialog
        from generer_collec_science import generer_collec_science as gen_cs
        from generer_collec_science import generer_collec_science_depuis_rows as gen_cs_rows
        from presentation.dialogs import ErrorDialog, InfoDialog, WarningDialog
        import traceback

        # Dialog unique : prévisualisation + sélection + contenants
        # software_rows est fourni quand on vient du logiciel (données déjà chargées)
        dlg = CollecScienceSourceDialog(
            parent=self,
            colisa_path=source_excel if software_rows is None else None,
            software_rows=software_rows,
            source_label=source_label,
        )
        if dlg.exec() != QDialog.Accepted:
            return
        containers = dlg.build_containers_dict()
        use_fixed_md_num_individu_column = dlg.use_fixed_md_num_individu_column()
        md_num_individu_column_index = dlg.get_md_num_individu_column_index()
        selected_software_rows = dlg.get_selected_software_rows() if software_rows is not None else None
        selected_num_individus = None
        if software_rows is None:
            selected_num_individus = set()
            for row_index in dlg.get_selected_row_indices():
                model = dlg._preview_model
                headers = getattr(model, "_headers", [])
                rows = getattr(model, "_rows", [])
                if row_index < 0 or row_index >= len(rows):
                    continue
                row_values = rows[row_index]
                row_dict = {
                    headers[col_idx]: row_values[col_idx]
                    for col_idx in range(min(len(headers), len(row_values)))
                }
                num_value = core.normalize(
                    row_dict.get("Numero individu")
                    or row_dict.get("Numero individu ")
                    or row_dict.get("num_individu")
                    or row_dict.get("Code echantillon")
                    or row_dict.get("code_echantillon")
                )
                if num_value:
                    selected_num_individus.add(num_value)
            if not selected_num_individus:
                selected_num_individus = None

        name_values: List[str] = []
        if software_rows is not None:
            rows_for_name = selected_software_rows or software_rows
            for row in rows_for_name:
                value = row.get("code_echantillon") or row.get("ref") or row.get("num_individu")
                if value:
                    name_values.append(str(value))
        elif selected_num_individus:
            name_values = sorted(selected_num_individus)

        out_path = self._choose_collec_science_output_path(
            source_excel,
            output_name_hint,
            name_values=name_values,
        )
        if out_path is None:
            return

        try:
            self.lbl_status.setText("Génération Collec-Science en cours...")

            if software_rows is not None:
                # Mode logiciel : génération directe depuis les rows, sans passer par COLISA
                result = gen_cs_rows(
                    rows=selected_software_rows or software_rows,
                    output_path=out_path,
                    containers=containers,
                    forcer_anomalies=forcer_anomalies,
                )
            else:
                # Mode fichier Excel externe : lecture du fichier source tel quel
                result = gen_cs(
                    colisa_path=source_excel,
                    output_path=out_path,
                    containers=containers,
                    forcer_anomalies=forcer_anomalies,
                    allowed_num_individus=selected_num_individus or allowed_num_individus,
                    prefer_fixed_num_individu_column=use_fixed_md_num_individu_column,
                    md_num_individu_column_index=md_num_individu_column_index,
                )

            rows_written = int(result.get("rows_written", 0) or 0) if result else 0
            nb_csv = len(result.get("csv_files", [])) if result else 0
            skipped_details = result.get("skipped_details", []) if result else []
            if rows_written <= 0:
                detail_lines = skipped_details[:8]
                details_text = ""
                if detail_lines:
                    details_text = "\n\nLignes a corriger :\n- " + "\n- ".join(detail_lines)
                    if len(skipped_details) > len(detail_lines):
                        details_text += f"\n- ... et {len(skipped_details) - len(detail_lines)} autre(s)"
                WarningDialog.show(
                    self,
                    "Collec-Science",
                    "Aucun echantillon n'a ete cree dans le fichier Excel.\n\n"
                    "Verifie la selection, les types d'echantillons et les donnees obligatoires."
                    + details_text,
                )
                self.lbl_status.setText("Aucun echantillon cree dans Collec-Science")
                return
            self.lbl_status.setText(f"✅ Collec-Science généré : {out_path.name}")
            info_lines = [
                "✅ Collec-Science généré avec succès !",
                f"Fichier Collect-Science Excel : {out_path}",
                f"Echantillons crees : {rows_written}",
                f"Nombre de CSV générés : {nb_csv}",
            ]
            if skipped_details:
                preview_lines = skipped_details[:5]
                info_lines.append(f"Lignes ignorees : {len(skipped_details)}")
                info_lines.append("A corriger : " + " | ".join(preview_lines))
            if result and result.get("csv_files"):
                info_lines.append("Fichiers Collect-Science CSV : " + ", ".join(result["csv_files"]))
            InfoDialog.show(self, "Collec-Science", "\n".join(info_lines))
        except Exception as exc:
            ErrorDialog.show(
                self,
                "Collec-Science",
                f"Erreur pendant la génération Collec-Science.\n"
                f"Emplacement : {out_path}\n\n"
                f"{exc}\n{traceback.format_exc()}",
            )
            self.lbl_status.setText("❌ Erreur génération Collec-Science")

    def _choose_collec_science_output_path(
        self,
        source_excel: Path | None,
        output_name_hint: str | None = None,
        name_values: List[str] | None = None,
    ) -> Path | None:
        range_hint = self._build_collect_science_range_hint(name_values or [])
        if range_hint:
            default_name = f"Collect-Science_{range_hint}.xlsx"
        else:
            default_name = "Collect-Science.xlsx"
        if source_excel is not None and source_excel.parent.exists():
            default_dir = source_excel.parent
        elif self.imports_dir.exists():
            default_dir = self.imports_dir
        else:
            default_dir = self.imports_dir
        default_path = default_dir / default_name
        selected_path, _ = QFileDialog.getSaveFileName(
            self,
            "📁 Où enregistrer le fichier Collect-Science ?",
            str(default_path),
            "Excel (*.xlsx)",
        )
        if not selected_path:
            return None
        output_path = Path(selected_path)
        if output_path.suffix.lower() != ".xlsx":
            output_path = output_path.with_suffix(".xlsx")
        return output_path

    def _build_collect_science_range_hint(self, name_values: List[str]) -> str:
        tokens: list[str] = []
        for raw_value in name_values:
            normalized = core.normalize(raw_value).upper()
            if not normalized:
                continue
            match = re.search(r"T(\d+)", normalized)
            if match:
                tokens.append(f"T{int(match.group(1))}")
                continue
            compact = re.sub(r"[^A-Z0-9]+", "", normalized)
            if compact:
                tokens.append(compact[:24])

        if not tokens:
            return ""

        unique_tokens = list(dict.fromkeys(tokens))
        if all(re.fullmatch(r"T\d+", token) for token in unique_tokens):
            sorted_tokens = sorted(unique_tokens, key=lambda token: int(token[1:]))
            if len(sorted_tokens) == 1:
                return sorted_tokens[0]
            return f"{sorted_tokens[0]}-{sorted_tokens[-1]}"

        if len(unique_tokens) == 1:
            return unique_tokens[0]
        return f"{unique_tokens[0]}-{unique_tokens[-1]}"

    def _source_rows_to_collect_science_rows(
        self,
        candidates: List[Any],
        default_type_echantillon: str,
    ) -> List[Dict[str, Any]]:
        rows: List[Dict[str, Any]] = []
        for r in candidates:
            errs = core.validate_row(r, default_type_echantillon)
            has_date = bool(core.normalize(r.date_capture))
            rows.append({
                "selected": has_date,
                "ref": core.normalize_ref_code(r.ref),
                "code_type_echantillon": default_type_echantillon,
                "categorie": core.normalize(r.categorie),
                "type_peche": core.normalize(r.type_peche),
                "autre_oss": "NON",
                "ecailles_brutes": core.normalize(getattr(r, "ecailles_brutes", "")),
                "montees": core.normalize(getattr(r, "montees", "")),
                "empreintes": core.normalize(getattr(r, "empreintes", "")),
                "otolithes": core.normalize(getattr(r, "otolithes", "")),
                "observation_disponibilite": core.normalize(r.observation_disponibilite),
                "source_row": r.source_row_index,
                "num_individu": core.normalize_ref_code(r.num_individu),
                "date_capture": core.normalize(r.date_capture),
                "code_espece": core.normalize(r.code_espece),
                "lac_riviere": core.normalize(r.lac_riviere),
                "pays_capture": core.normalize(r.pays_capture),
                "pecheur": core.normalize(r.pecheur),
                "longueur_mm": core.normalize(r.longueur_mm),
                    "poids_g": core.normalize(r.poids_g),
                    "maturite": core.normalize(r.maturite),
                    "sexe": core.normalize(r.sexe),
                    "age_total": core.normalize(r.age_total),
                    "status": "a_reimporter" if errs else "pret",
                    "errors": " | ".join(errs) if errs else "",
                })
        return rows

    def _load_collect_science_rows_from_excel(self, source_excel: Path) -> List[Dict[str, Any]] | None:
        from presentation.source_format_dialog import SourceFormatDialog

        sheet_names = core.get_workbook_sheet_names(source_excel)
        if not sheet_names:
            raise ValueError("Le fichier Excel choisi ne contient aucun onglet exploitable.")

        same_as_current_source = False
        try:
            same_as_current_source = source_excel.resolve() == self.source_path.resolve()
        except Exception:
            same_as_current_source = source_excel == self.source_path

        sheet_name = (
            self.source_sheet
            if same_as_current_source and self.source_sheet in sheet_names
            else sheet_names[0]
        )
        mode = self.source_mode if same_as_current_source else ("pac_final" if source_excel.suffix.lower() == ".xls" else "custom")
        mapping = self.source_mapping if same_as_current_source else {}

        if mode == "custom" and not (mapping.get("columns") or {}):
            dialog = SourceFormatDialog(
                source_path=source_excel,
                current_mode="custom",
                current_sheet=sheet_name,
                current_mapping=mapping,
                force_custom=True,
                parent=self,
            )
            if dialog.exec() != QDialog.Accepted:
                return None
            result = dialog.get_result()
            sheet_name = str(result.get("sheet_name") or sheet_name)
            mapping = result.get("mapping", {}) or {}

            if same_as_current_source:
                self.source_mode = "custom"
                self.source_sheet = sheet_name
                self.source_mapping = mapping
                self._refresh_context_labels()
                self._save_settings()

        source_rows, datemode = core.read_any_source_rows(source_excel, sheet_name)
        if mode == "custom":
            candidates = core.find_candidate_rows_from_mapping(source_rows, datemode, mapping)
        else:
            candidates = core.find_candidate_rows(source_rows, datemode)

        default_type = self._current_bulk_type_value() or "EC"
        return self._source_rows_to_collect_science_rows(candidates, default_type)

    def generer_collec_science_depuis_excel(self) -> None:
        """Generate Collect-Science from a selected Excel file."""
        from presentation.dialogs import WarningDialog

        p, _ = QFileDialog.getOpenFileName(
            self,
            "Choisir un fichier Excel",
            str(self.imports_dir),
            "Excel (*.xls *.xlsx)",
        )
        if not p:
            return

        source_excel = Path(p)
        if not source_excel.exists():
            WarningDialog.show(self, "Collec-Science", "Le fichier Excel choisi est introuvable.")
            return

        try:
            is_colisa_source = False
            if source_excel.suffix.lower() == ".xlsx":
                openpyxl, _ = core.ensure_deps()
                workbook = openpyxl.load_workbook(source_excel, read_only=True, data_only=True)
                try:
                    is_colisa_source, _ = validate_collect_science_source_workbook(workbook)
                finally:
                    workbook.close()

            if is_colisa_source:
                self._run_collec_science_export(
                    source_excel=source_excel,
                    output_name_hint=source_excel.stem,
                )
                return

            rows = self._load_collect_science_rows_from_excel(source_excel)
            if rows is None:
                return
            if not rows:
                WarningDialog.show(
                    self,
                    "Collec-Science",
                    "Aucun echantillon exploitable n'a ete trouve dans le fichier choisi.",
                )
                return

            self._run_collec_science_export(
                source_excel=source_excel,
                software_rows=rows,
                output_name_hint=source_excel.stem,
                source_label=f"{len(rows)} echantillon(s) depuis {source_excel.name}",
            )
        except Exception as exc:
            import traceback
            from presentation.dialogs import ErrorDialog
            ErrorDialog.show(self, "Collec-Science", str(exc) + "\n" + traceback.format_exc())
            self.lbl_status.setText("Erreur generation Collec-Science")

    def generer_format_colisa_logiciel(self) -> None:
        """Generate the COLISA logiciel format from a selected Excel (COLISA brut) file."""
        from presentation.dialogs import ErrorDialog, WarningDialog
        from generer_colisa_logiciel import lire_rows_depuis_excel_colisa

        try:
            default_open_dir = self.imports_dir
            p, _ = QFileDialog.getOpenFileName(
                self,
                "Choisir un fichier Excel deja au format COLISA brut",
                str(default_open_dir),
                "Excel (*.xlsx *.xls)",
            )
            if not p:
                return
            source_excel = Path(p)
            rows = lire_rows_depuis_excel_colisa(source_excel)
            if not rows:
                WarningDialog.show(
                    self,
                    "Format COLISA logiciel",
                    "Aucun echantillon a generer. Le fichier choisi doit deja etre un fichier Excel au format COLISA brut.",
                )
                return
            self._run_colisa_logiciel_export(rows=rows, source_name=source_excel.stem)
        except Exception as exc:
            import traceback
            ErrorDialog.show(self, "Format COLISA logiciel", str(exc) + "\n" + traceback.format_exc())
            self.lbl_status.setText("Erreur generation format COLISA logiciel")

    def _generer_colisa_logiciel_depuis_logiciel(self) -> None:
        """Generate the COLISA logiciel format from currently selected rows in the software."""
        from presentation.dialogs import ErrorDialog, WarningDialog

        try:
            self._flush_table_edits()
            data = self._read_table()
            rows = [r for r in data if bool(r.get("selected"))]
            if not rows:
                WarningDialog.show(self, "Format COLISA logiciel", "Selectionne au moins un echantillon dans le tableau.")
                return
            self._run_colisa_logiciel_export(rows=rows, source_name="echantillons_selectionnes")
        except Exception as exc:
            import traceback
            ErrorDialog.show(self, "Format COLISA logiciel", str(exc) + "\n" + traceback.format_exc())
            self.lbl_status.setText("Erreur generation format COLISA logiciel")

    def _run_colisa_logiciel_export(self, rows: List[Dict[str, Any]], source_name: str) -> None:
        """Common logic: preparation dialog + file save + generation for COLISA logiciel."""
        from presentation.dialogs import ErrorDialog, InfoDialog, WarningDialog
        from generer_colisa_logiciel import generer_colisa_logiciel_depuis_rows
        from presentation.colisa_logiciel_dialog import ColisaLogicielPreparationDialog

        if not self.colisa_logiciel_template_path.exists():
            tpl = self._demander_et_installer_template_colisa_logiciel()
            if tpl is None:
                return
            self.colisa_logiciel_template_path = tpl

        dlg = ColisaLogicielPreparationDialog(
            rows=rows,
            template_path=self.colisa_logiciel_template_path,
            default_site_atelier=self.ed_site_atelier.text().strip(),
            parent=self,
        )
        if dlg.exec() != QDialog.Accepted:
            return
        rows = dlg.get_rows()
        if not rows:
            WarningDialog.show(self, "Format COLISA logiciel", "Aucune ligne n'est cochee pour la generation.")
            return

        default_dir = self.imports_dir
        default_path = default_dir / f"{source_name}_format_colisa_logiciel.xlsx"
        selected_path, _ = QFileDialog.getSaveFileName(
            self,
            "📁 Où enregistrer le format COLISA logiciel ?",
            str(default_path),
            "Excel (*.xlsx)",
        )
        if not selected_path:
            return

        output_path = Path(selected_path)
        if output_path.suffix.lower() != ".xlsx":
            output_path = output_path.with_suffix(".xlsx")

        result = generer_colisa_logiciel_depuis_rows(
            rows=rows,
            template_path=self.colisa_logiciel_template_path,
            output_path=output_path,
            default_code_unite_gestionnaire=self.ed_code_unite.text().strip(),
            default_site_atelier=self.ed_site_atelier.text().strip(),
            default_numero_correspondant=self.ed_num_correspondant.text().strip(),
            default_organisme=self.ed_org.text().strip(),
        )
        rows_written = int(result.get("rows_written", 0) or 0)
        if rows_written <= 0:
            WarningDialog.show(self, "Format COLISA logiciel", "Aucun echantillon n'a ete ecrit dans le fichier genere.")
            self.lbl_status.setText("Aucun echantillon genere dans le format COLISA")
            return
        self.lbl_status.setText(f"Format COLISA logiciel genere : {output_path.name}")
        InfoDialog.show(
            self,
            "Format COLISA logiciel",
            f"Fichier Excel : {output_path}\nLignes generees : {rows_written}",
        )

    def generer_pipeline_complet(self) -> None:
        """Lance l'import puis, une fois termine, genere Collec-Science et COLISA logiciel depuis le logiciel."""
        from PySide6.QtWidgets import QMessageBox
        from presentation.dialogs import ErrorDialog

        box = QMessageBox(self)
        box.setWindowTitle("Export complet logiciel")
        box.setText("Importer dans le fichier COLISA existant ou repartir sur un nouveau fichier vierge ?")
        btn_existant = box.addButton("Dans l'existant", QMessageBox.AcceptRole)
        btn_nouveau = box.addButton("Nouveau fichier", QMessageBox.DestructiveRole)
        box.addButton("Annuler", QMessageBox.RejectRole)
        box.exec()
        clicked = box.clickedButton()
        if clicked is None or clicked not in (btn_existant, btn_nouveau):
            return
        if clicked == btn_existant:
            default_open_dir = self.imports_dir
            p, _ = QFileDialog.getOpenFileName(
                self,
                "📁 Choisir le fichier COLISA existant",
                str(default_open_dir),
                "Excel (*.xlsx)",
            )
            if not p:
                return
            self.out_path = Path(p)
            self._refresh_context_labels()
            self._save_settings()
        if clicked == btn_nouveau:
            nouveau_path, _ = QFileDialog.getSaveFileName(
                self,
                "Choisir l'emplacement du nouveau fichier COLISA",
                str(self.imports_dir / "COLISA en cours.xlsx"),
                "Excel (*.xlsx)",
            )
            if not nouveau_path:
                return
            dest = Path(nouveau_path)
            if dest.suffix.lower() != ".xlsx":
                dest = dest.with_suffix(".xlsx")
            if dest.exists():
                from presentation.dialogs import WarningDialog
                WarningDialog.show(
                    self,
                    "Export complet logiciel",
                    f"Ce fichier existe déjà et ne sera pas écrasé :\n{dest}\n\nChoisis un autre nom.",
                )
                return
            try:
                base_path = self._ensure_internal_target_base()
                dest.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(base_path, dest)
                self.out_path = dest
                self._refresh_context_labels()
            except Exception as exc:
                ErrorDialog.show(self, "Export complet logiciel", f"Impossible de creer le nouveau fichier :\n{exc}")
                return

        self._post_import_pipeline = True
        self.run_import()

    def show_history(self) -> None:
        """Show import history dialog."""
        from presentation.dialogs import HistoryDialog, InfoDialog, WarningDialog
        try:
            if not self.history_path.exists():
                InfoDialog.show(self, "Suivi", "Aucun suivi disponible pour le moment.")
                return
            payload = json.loads(self.history_path.read_text(encoding="utf-8"))
            HistoryDialog.show(self, payload)
        except Exception as exc:
            WarningDialog.show(self, "Suivi", str(exc))


def main() -> None:
    app = QApplication(sys.argv)
    w = MainWindow()
    w.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()


