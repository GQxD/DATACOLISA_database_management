#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import os
import re
import sys
import unicodedata
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


REQUIRED_IMPORTS = {
    "openpyxl": "openpyxl",
    "xlrd": "xlrd",
}

DEFAULT_SOURCE_SHEET = "Travail4avril2012"
DEFAULT_TARGET_SHEET = "Feuil1"
TYPE_SHEET_CANDIDATES = ["Type echantillon", "Type ?chantillon", "Type_echantillon"]

# Mapping V1: source positional columns (1-based indexes in source data row)
# based on current project context.
SOURCE_POSITIONS = {
    "num_individu_primary": 2,   # REF
    "pecheur": 4,               # C4 (nom pecheur)
    "contexte": 5,              # C5 (code se terminant FR/CH)
    "code_espece": 6,            # C6
    "date_capture": 7,           # C7
    "lac_riviere": 10,           # C10
    "engin_source": 14,         # C14 / colonne N
    "longueur_mm": 16,           # C16
    "poids_g": 17,               # C17
    "maturite": 20,              # C20
    "sexe": 21,                  # C21
    "num_individu_fallback": 26, # C26
    "age_total": 36,             # C36
}

TARGET_HEADERS = {
    "code_unite_gestionnaire": ["Code unit? gestionnaire", "Code unite gestionnaire"],
    "site_atelier": ["Site Atelier", "Site atelier"],
    "numero_correspondant": ["Numero du correspondant", "Num?ro du correspondant"],
    "code_type_echantillon": ["Code type échantillon", "Code type echantillon"],
    "code_echantillon": ["Code échantillon", "Code echantillon"],
    "code_espece": ["Code espèce", "Code espece"],
    "organisme": ["Organisme préleveur", "Organisme preleveur"],
    "pays": ["Pays capture"],
    "date_capture": ["Date capture", "Date de capture"],
    "lac_riviere": ["Lac/rivière", "Lac/riviere"],
    "categorie": ["Catégorie pêcheur", "Categorie pecheur"],
    "type_peche": ["Type pêche/engin", "Type peche/engin"],
    "code_compose": ["Ligne LE02", "LE02"],
    "num_individu": ["Numéro individu", "Numero individu", "Numero individu (numero de capture)"],
    "longueur_mm": ["Longueur totale (mm)"],
    "poids_g": ["Poids (g)"],
    "maturite": ["Code maturité sexuelle", "Code maturite sexuelle"],
    "sexe": ["Code sexe"],
    "age_total": ["Age total"],
    "ecailles_brutes": ["Ecailles brutes", "Écailles brutes"],
    "montees": ["Montées", "Montees"],
    "otolithes": ["Otolithes"],
    "autre_oss": ["Autre ?chantillon osseuses collect?e sur l'individu OUI/NON", "Autre echantillon osseuses collectee sur l'individu OUI/NON"],
    "observation_disponibilite": ["Observation disponibilit?", "Observation disponibilite"],
}


def resolve_target_sheet(wb: Any, requested_sheet: str) -> Any:
    if requested_sheet in wb.sheetnames:
        ws = wb[requested_sheet]
        if ws.max_row > 1 or ws.max_column > 1:
            return ws

    req_norm = normalize(requested_sheet)
    candidates = []
    for name in wb.sheetnames:
        if normalize(name) == req_norm:
            ws = wb[name]
            candidates.append((ws.max_row * ws.max_column, ws))
    if candidates:
        candidates.sort(key=lambda x: x[0], reverse=True)
        return candidates[0][1]

    fatal(f"Onglet cible introuvable: {requested_sheet}")


@dataclass
class SourceRow:
    source_row_index: int
    ref: str
    code_espece: Any
    date_capture: Any
    lac_riviere: Any
    num_individu: Any
    longueur_mm: Any
    poids_g: Any
    maturite: Any
    sexe: Any
    age_total: Any
    type_peche: Any
    categorie: Any
    pecheur: Any
    pays_capture: Any
    pecheur_source: Any
    observation_disponibilite: Any


def fatal(msg: str) -> None:
    print(f"ERREUR: {msg}", file=sys.stderr)
    sys.exit(1)


def ensure_deps() -> Tuple[Any, Any]:
    missing = []
    loaded = {}
    for mod_name in REQUIRED_IMPORTS:
        try:
            loaded[mod_name] = __import__(mod_name)
        except Exception:
            missing.append(mod_name)
    if missing:
        fatal(
            "Modules manquants: "
            + ", ".join(missing)
            + ". Installez-les avec: pip install -r requirements.txt"
        )
    return loaded["openpyxl"], loaded["xlrd"]


def normalize(s: Any) -> str:
    if s is None:
        return ""
    return str(s).strip()


def normalize_header_name(s: Any) -> str:
    txt = normalize(s)
    if not txt:
        return ""
    txt = unicodedata.normalize("NFKD", txt)
    txt = "".join(ch for ch in txt if not unicodedata.combining(ch))
    txt = txt.lower()
    txt = re.sub(r"[^a-z0-9]+", " ", txt)
    return re.sub(r"\s+", " ", txt).strip()


def parse_ref_parts(code: str) -> Optional[Tuple[str, int]]:
    code = normalize(code).upper()
    m = re.match(r"^([A-Z]+)\s*0*(\d+)$", code)
    if not m:
        return None
    return m.group(1), int(m.group(2))


def in_ref_range(code: str, start: str, end: str) -> bool:
    p_code = parse_ref_parts(code)
    p_start = parse_ref_parts(start)
    p_end = parse_ref_parts(end)
    if not p_code or not p_start or not p_end:
        return False
    if p_start[0] != p_end[0] or p_code[0] != p_start[0]:
        return False
    return p_start[1] <= p_code[1] <= p_end[1]


def read_source_rows(xlrd: Any, source_file: Path, sheet_name: str) -> Tuple[List[List[Any]], int]:
    wb = xlrd.open_workbook(str(source_file), formatting_info=False)
    if sheet_name not in wb.sheet_names():
        fatal(f"Onglet source introuvable: {sheet_name}")
    ws = wb.sheet_by_name(sheet_name)
    rows: List[List[Any]] = []
    for r in range(ws.nrows):
        rows.append([ws.cell_value(r, c) if c < ws.ncols else None for c in range(ws.ncols)])
    return rows, int(getattr(wb, "datemode", 0))


def get_pos(row: List[Any], one_based_col: int) -> Any:
    idx = one_based_col - 1
    if idx < 0 or idx >= len(row):
        return None
    return row[idx]


def _excel_date_to_date(value: Any, datemode: int = 0) -> Optional[dt.date]:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, (int, float)):
        try:
            import xlrd  # type: ignore

            tup = xlrd.xldate_as_tuple(float(value), datemode)
            return dt.date(tup[0], tup[1], tup[2])
        except Exception:
            pass
        try:
            base = dt.date(1899, 12, 30)
            return base + dt.timedelta(days=int(float(value)))
        except Exception:
            return None
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
            try:
                return dt.datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def format_capture_date(value: Any, datemode: int = 0) -> str:
    d = _excel_date_to_date(value, datemode)
    if d:
        return d.strftime("%d/%m/%y")
    return normalize(value)


def derive_type_and_categorie_from_source(raw_engin: Any) -> Tuple[str, str]:
    txt = normalize(raw_engin)
    if not txt:
        return "", ""

    low = txt.lower()
    up = txt.upper()

    # Type peche: conserver la valeur source, sauf abreviations T/F/S normalisees
    if low == "t":
        type_peche = "TRAINE"
    elif low == "f":
        type_peche = "FILET"
    elif low == "s":
        type_peche = "SONDE"
    elif low == "_":
        type_peche = ""
    elif "traine" in low:
        type_peche = "TRAINE"
    elif "filet" in low:
        type_peche = "FILET"
    elif "sonde" in low:
        type_peche = "SONDE"
    elif "ligne" in low:
        type_peche = "LIGNE"
    else:
        type_peche = up

    categorie = ""
    if "pic" in low:
        categorie = "PRO"
    elif "sci" in low:
        categorie = "SCIENTIFIQUE"
    elif type_peche == "TRAINE" or "traine" in low:
        categorie = "AMATEUR"
    elif "pro" in low:
        categorie = "PRO"
    elif "amat" in low:
        categorie = "AMATEUR"

    return type_peche, categorie




def derive_country_from_contexte(raw_contexte: Any) -> str:
    txt = normalize(raw_contexte).upper()
    if not txt:
        return ""
    m = re.search(r"(CH|FR)\s*$", txt)
    if not m:
        return ""
    code = m.group(1)
    if code == "CH":
        return "Suisse"
    if code == "FR":
        return "France"
    return ""

def find_candidate_rows(rows: List[List[Any]], datemode: int = 0) -> List[SourceRow]:
    out: List[SourceRow] = []
    for i, row in enumerate(rows, start=1):
        ref = normalize(get_pos(row, SOURCE_POSITIONS["num_individu_primary"]))
        if not ref:
            continue
        if not re.search(r"\d", ref):
            continue

        num_individu = ref or normalize(get_pos(row, SOURCE_POSITIONS["num_individu_fallback"]))
        engin_raw = get_pos(row, SOURCE_POSITIONS["engin_source"])
        contexte_raw = get_pos(row, SOURCE_POSITIONS["contexte"])
        type_peche, categorie = derive_type_and_categorie_from_source(engin_raw)
        pays_capture = derive_country_from_contexte(contexte_raw)

        poids_src = normalize(get_pos(row, SOURCE_POSITIONS["poids_g"]))
        sexe_src = normalize(get_pos(row, SOURCE_POSITIONS["sexe"]))
        if poids_src == "_":
            poids_src = ""
        if sexe_src == "_":
            sexe_src = ""

        out.append(
            SourceRow(
                source_row_index=i,
                ref=ref,
                code_espece=get_pos(row, SOURCE_POSITIONS["code_espece"]),
                date_capture=format_capture_date(get_pos(row, SOURCE_POSITIONS["date_capture"]), datemode),
                lac_riviere=get_pos(row, SOURCE_POSITIONS["lac_riviere"]),
                num_individu=num_individu,
                longueur_mm=get_pos(row, SOURCE_POSITIONS["longueur_mm"]),
                poids_g=poids_src if poids_src else "N",
                maturite=get_pos(row, SOURCE_POSITIONS["maturite"]),
                sexe=sexe_src if sexe_src else "N",
                age_total=get_pos(row, SOURCE_POSITIONS["age_total"]),
                type_peche=type_peche,
                categorie=categorie,
                pecheur=normalize(get_pos(row, SOURCE_POSITIONS["pecheur"])),
                pays_capture=pays_capture,
                pecheur_source=normalize(engin_raw),
                observation_disponibilite="",
            )
        )
    return out


def validate_row(r: SourceRow, default_type_echantillon: str) -> List[str]:
    missing = []
    if not normalize(r.date_capture):
        missing.append("Date de capture (entrée manquante pour T11)")
    if not normalize(r.code_espece):
        missing.append("Taxon / Code espèce (entrée manquante pour T6)")
    if not normalize(r.num_individu):
        missing.append("Numéro individu (entrée manquante pour T19)")
    if not normalize(default_type_echantillon):
        missing.append("Code type échantillon (T4)")
    # Champs calculés en cible: on vérifie les entrées, pas la formule finale.
    if not normalize(r.num_individu):
        missing.append("Entrée requise pour génération Code échantillon")
    if not normalize(r.code_espece):
        missing.append("Entrée requise pour génération ligne LE02")
    return missing


def write_extract_csv(rows: List[SourceRow], path: Path, default_type_echantillon: str) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(
            [
                "include",
                "status",
                "source_row",
                "ref",
                "num_individu",
                "date_capture",
                "code_espece",
                "lac_riviere",
                "pays_capture",
                "pecheur",
                "pecheur_source",
                "categorie",
                "type_peche",
                "observation_disponibilite",
                "autre_oss",
                "ecailles_brutes",
                "montees",
                "otolithes",
                "longueur_mm",
                "poids_g",
                "maturite",
                "sexe",
                "age_total",
                "code_type_echantillon",
                "errors",
            ]
        )
        for r in rows:
            errs = validate_row(r, default_type_echantillon)
            w.writerow(
                [
                    "1",
                    "a_reimporter" if errs else "pret",
                    r.source_row_index,
                    r.ref,
                    r.num_individu,
                    r.date_capture,
                    r.code_espece,
                    r.lac_riviere,
                    r.pays_capture,
                    r.pecheur,
                    r.pecheur_source,
                    r.categorie,
                    r.type_peche,
                    r.observation_disponibilite,
                    "",
                    "",
                    "",
                    "",
                    r.longueur_mm,
                    r.poids_g,
                    r.maturite,
                    r.sexe,
                    r.age_total,
                    default_type_echantillon,
                    " | ".join(errs),
                ]
            )


def find_header_row_and_map(ws: Any) -> Tuple[int, Dict[str, int]]:
    header_row = None
    header_map: Dict[str, int] = {}
    alias_norm: Dict[str, set[str]] = {}
    for key, aliases in TARGET_HEADERS.items():
        alias_norm[key] = {normalize_header_name(a) for a in aliases if normalize_header_name(a)}

    for r in range(1, 51):
        values = [normalize(ws.cell(r, c).value) for c in range(1, min(ws.max_column, 300) + 1)]
        for idx, val in enumerate(values, start=1):
            val_norm = normalize_header_name(val)
            if not val_norm:
                continue
            for key, aliases in alias_norm.items():
                if val_norm in aliases:
                    header_map[key] = idx
        if "num_individu" in header_map and "code_echantillon" in header_map:
            header_row = r
            break
        header_map.clear()

    if header_row is None:
        fatal("Impossible de trouver la ligne d'en-tête de Feuil1 (Code échantillon + Numéro individu)")

    return header_row, header_map


def first_empty_row(ws: Any, header_row: int, anchor_col: int) -> int:
    r = header_row + 1
    while ws.cell(r, anchor_col).value not in (None, ""):
        r += 1
    return r


def build_existing_index(ws: Any, header_row: int, header_map: Dict[str, int]) -> Dict[Tuple[str, str], Dict[str, Any]]:
    idx: Dict[Tuple[str, str], Dict[str, Any]] = {}
    num_col = header_map.get("num_individu")
    type_col = header_map.get("code_type_echantillon")
    if not num_col or not type_col:
        return idx

    for r in range(header_row + 1, ws.max_row + 1):
        num = normalize(ws.cell(r, num_col).value)
        typ = normalize(ws.cell(r, type_col).value)
        if not num:
            continue
        idx[(num, typ)] = {
            "row": r,
            "date_capture": normalize(ws.cell(r, header_map.get("date_capture", 1)).value),
            "code_espece": normalize(ws.cell(r, header_map.get("code_espece", 1)).value),
        }
    return idx


def init_code_sequence(ws: Any, header_row: int, code_col: Optional[int]) -> Tuple[str, int, int]:
    if not code_col:
        return "T", 0, 0

    pattern = re.compile(r"^([A-Za-z]*)(\d+)$")
    best_prefix = "T"
    best_num = 0
    best_width = 0

    for r in range(header_row + 1, ws.max_row + 1):
        raw = normalize(ws.cell(r, code_col).value)
        if not raw:
            continue
        m = pattern.match(raw)
        if not m:
            continue
        prefix, num_s = m.group(1), m.group(2)
        num = int(num_s)
        width = len(num_s)
        if num > best_num:
            best_num = num
            best_prefix = prefix or best_prefix
            best_width = width

    if best_width == 0:
        best_width = 5
    return best_prefix or "T", best_num, best_width


def assign_next_code_if_missing(
    ws: Any,
    target_row: int,
    code_col: Optional[int],
    seq_state: Dict[str, Any],
) -> None:
    if not code_col:
        return
    cell = ws.cell(target_row, code_col)
    if normalize(cell.value):
        return

    seq_state["num"] += 1
    prefix = seq_state.get("prefix", "T") or "T"
    width = int(seq_state.get("width", 5))
    cell.value = f"{prefix}{str(seq_state['num']).zfill(width)}"


def excel_date_from_xlrd(value: Any, datemode: int = 0) -> Any:
    try:
        if isinstance(value, (int, float)) and value > 0:
            import xlrd  # type: ignore
            tup = xlrd.xldate_as_tuple(value, datemode)
            return dt.datetime(*tup).date()
    except Exception:
        pass
    return value


def parse_capture_date_for_target(value: Any) -> Any:
    d = _excel_date_to_date(value, 0)
    if d:
        return d
    return normalize(value)


def apply_target_row(
    ws: Any,
    r: int,
    row: Dict[str, str],
    header_map: Dict[str, int],
    default_org: str,
    default_country: str,
) -> None:
    def set_if_header(key: str, value: Any) -> None:
        c = header_map.get(key)
        if c:
            ws.cell(r, c).value = value

    set_if_header("code_type_echantillon", row.get("code_type_echantillon", ""))
    # code_echantillon and code_compose are formula-driven in target; keep untouched.
    set_if_header("code_espece", row.get("code_espece", ""))
    set_if_header("organisme", default_org)
    country_val = normalize(row.get("pays_capture", "")) or default_country
    set_if_header("pays", country_val)

    capture_val = parse_capture_date_for_target(row.get("date_capture", ""))
    set_if_header("date_capture", capture_val)
    date_col = header_map.get("date_capture")
    if date_col and isinstance(capture_val, (dt.date, dt.datetime)):
        ws.cell(r, date_col).number_format = "dd/mm/yy"

    lac_val = normalize(row.get("lac_riviere", "")).upper()
    if lac_val == "LEMAN":
        lac_val = "L"
    set_if_header("lac_riviere", lac_val)
    lac_col = header_map.get("lac_riviere")
    if lac_col:
        ws.cell(r, lac_col).number_format = "@"

    categorie_val = normalize(row.get("categorie", ""))
    type_peche_val = normalize(row.get("type_peche", ""))

    # Regles metier automatiques
    src_pecheur = normalize(row.get("pecheur_source", ""))
    derived_type, derived_cat = derive_type_and_categorie_from_source(src_pecheur)
    if derived_type:
        type_peche_val = derived_type
    if derived_cat:
        categorie_val = derived_cat

    set_if_header("categorie", categorie_val)
    set_if_header("type_peche", type_peche_val)

    otolithes_val = normalize(row.get("otolithes", ""))
    autre_oss_val = normalize(row.get("autre_oss", ""))
    if otolithes_val and otolithes_val != "0":
        autre_oss_val = "OUI"
    else:
        autre_oss_val = "NON"
    set_if_header("autre_oss", autre_oss_val)
    set_if_header("ecailles_brutes", row.get("ecailles_brutes", ""))
    set_if_header("montees", row.get("montees", ""))
    set_if_header("otolithes", row.get("otolithes", ""))
    observation_val = row.get("observation_disponibilite")
    if observation_val in (None, ""):
        observation_val = row.get("observation_disponible", "")
    set_if_header("observation_disponibilite", observation_val)

    set_if_header("num_individu", row.get("num_individu", ""))
    set_if_header("longueur_mm", row.get("longueur_mm", ""))
    poids_val = normalize(row.get("poids_g", ""))
    sexe_val = normalize(row.get("sexe", ""))
    if poids_val in ("", "_"):
        poids_val = "N"
    if sexe_val in ("", "_"):
        sexe_val = "N"
    set_if_header("poids_g", poids_val)
    set_if_header("maturite", row.get("maturite", ""))
    set_if_header("sexe", sexe_val)
    set_if_header("age_total", row.get("age_total", ""))


def propagate_formula_from_previous_row(openpyxl: Any, ws: Any, target_row: int, col_index: Optional[int], min_row: int = 2) -> None:
    if not col_index or target_row <= min_row:
        return

    current_cell = ws.cell(target_row, col_index)
    if current_cell.value not in (None, ""):
        return

    source_row = None
    source_formula = None
    for r in range(target_row - 1, min_row - 1, -1):
        candidate = ws.cell(r, col_index).value
        if isinstance(candidate, str) and candidate.startswith("="):
            source_row = r
            source_formula = candidate
            break

    if source_row is None or source_formula is None:
        return

    try:
        from openpyxl.formula.translate import Translator  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore

        col_letter = get_column_letter(col_index)
        origin = f"{col_letter}{source_row}"
        dest = f"{col_letter}{target_row}"
        current_cell.value = Translator(source_formula, origin=origin).translate_formula(dest)
    except Exception:
        current_cell.value = source_formula


def propagate_missing_formulas_row(openpyxl: Any, ws: Any, target_row: int, min_row: int = 2) -> None:
    # Propage toutes les formules manquantes pour la ligne (colonnes sans mapping explicite incluses).
    for col in range(1, ws.max_column + 1):
        propagate_formula_from_previous_row(openpyxl, ws, target_row, col, min_row=min_row)


def propagate_incremental_code_from_previous_row(ws: Any, target_row: int, col_index: Optional[int], min_row: int = 2) -> None:
    if not col_index or target_row <= min_row:
        return

    cell = ws.cell(target_row, col_index)
    if cell.value not in (None, ""):
        return

    pattern = re.compile(r"^(.*?)(\d+)$")
    for r in range(target_row - 1, min_row - 1, -1):
        prev_val = normalize(ws.cell(r, col_index).value)
        if not prev_val:
            continue
        m = pattern.match(prev_val)
        if not m:
            continue
        prefix, num = m.group(1), m.group(2)
        next_num = str(int(num) + 1).zfill(len(num))
        cell.value = f"{prefix}{next_num}"
        return


def copy_context_values_from_previous_row(
    ws: Any,
    target_row: int,
    header_map: Dict[str, int],
    min_row: int = 2,
    source_rows: Optional[List[int]] = None,
    expected_type: str = "",
    expected_ref_prefix: str = "",
) -> None:
    # Keep context propagation for structural metadata only.
    # Sampling detail fields must stay exactly as entered in the current row.
    keys = ["code_unite_gestionnaire", "site_atelier", "numero_correspondant", "lac_riviere", "categorie", "type_peche", "autre_oss", "observation_disponibilite"]

    type_col = header_map.get("code_type_echantillon")
    num_col = header_map.get("num_individu")

    if source_rows:
        candidates = [r for r in source_rows if min_row <= r < target_row]
        candidates = sorted(candidates, reverse=True)
    else:
        candidates = list(range(target_row - 1, min_row - 1, -1))

    source_row = None
    for r in candidates:
        if type_col and expected_type:
            row_type = normalize(ws.cell(r, type_col).value)
            if row_type != normalize(expected_type):
                continue

        if num_col and expected_ref_prefix:
            num_val = normalize(ws.cell(r, num_col).value)
            parts = parse_ref_parts(num_val)
            prefix = parts[0] if parts else ""
            if prefix != expected_ref_prefix:
                continue

        has_context = False
        for key in keys:
            col = header_map.get(key)
            if col and normalize(ws.cell(r, col).value):
                has_context = True
                break

        if has_context:
            source_row = r
            break

    if source_row is None:
        return

    for key in keys:
        col = header_map.get(key)
        if not col:
            continue
        current = normalize(ws.cell(target_row, col).value)
        if current:
            continue
        prev = ws.cell(source_row, col).value
        if normalize(prev):
            ws.cell(target_row, col).value = prev


def update_history(history_path: Path, rows: List[Dict[str, Any]]) -> None:
    payload = {
        "updated_at": dt.datetime.now().isoformat(timespec="seconds"),
        "rows": rows,
    }
    history_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _resolve_type_sheet(wb: Any) -> Any:
    names = list(getattr(wb, "sheetnames", []))
    for candidate in TYPE_SHEET_CANDIDATES:
        if candidate in names:
            return wb[candidate]

    normalized = {normalize(n).lower(): n for n in names}
    for candidate in TYPE_SHEET_CANDIDATES:
        key = normalize(candidate).lower()
        if key in normalized:
            return wb[normalized[key]]
    return None


def load_type_echantillon_options(workbook_path: Path) -> List[str]:
    if not workbook_path.exists():
        return []
    openpyxl, _ = ensure_deps()
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        ws = _resolve_type_sheet(wb)
        if ws is None:
            return []

        out: List[str] = []
        seen = set()
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            val = normalize(row[0] if row else "")
            if not val:
                continue
            low = val.lower()
            if low in ("type echantillon", "type ?chantillon", "code type echantillon", "code type ?chantillon"):
                continue
            if val not in seen:
                seen.add(val)
                out.append(val)
        return out
    finally:
        wb.close()


def append_type_echantillon_option(workbook_path: Path, new_type: str) -> bool:
    value = normalize(new_type)
    if not value:
        return False
    if not workbook_path.exists():
        return False

    openpyxl, _ = ensure_deps()
    wb = openpyxl.load_workbook(workbook_path)
    try:
        ws = _resolve_type_sheet(wb)
        if ws is None:
            ws = wb.create_sheet(TYPE_SHEET_CANDIDATES[0])
            ws.cell(1, 1).value = "Code type ?chantillon"

        existing = set()
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            existing.add(normalize(row[0] if row else ""))
        if value in existing:
            return True

        next_row = ws.max_row + 1
        if next_row == 1:
            next_row = 2
            ws.cell(1, 1).value = "Code type ?chantillon"
        ws.cell(next_row, 1).value = value
        wb.save(workbook_path)
        return True
    finally:
        wb.close()


def cmd_extract(args: argparse.Namespace) -> None:
    _, xlrd = ensure_deps()

    source_rows, datemode = read_source_rows(xlrd, Path(args.source), args.source_sheet)
    candidates = find_candidate_rows(source_rows, datemode)

    filtered = [r for r in candidates if in_ref_range(r.ref, args.start_ref, args.end_ref)]
    filtered.sort(key=lambda r: parse_ref_parts(r.ref)[1] if parse_ref_parts(r.ref) else 0)

    out_csv = Path(args.out_csv)
    write_extract_csv(filtered, out_csv, args.default_type_echantillon)

    found_codes = {normalize(r.ref).upper() for r in filtered}
    missing = []
    p_start = parse_ref_parts(args.start_ref)
    p_end = parse_ref_parts(args.end_ref)
    if p_start and p_end and p_start[0] == p_end[0]:
        for n in range(p_start[1], p_end[1] + 1):
            code = f"{p_start[0]}{n}"
            if code not in found_codes:
                missing.append(code)

    summary = {
        "range": f"{args.start_ref}..{args.end_ref}",
        "found": len(filtered),
        "missing": missing,
        "extract_csv": str(out_csv),
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))


def cmd_import(args: argparse.Namespace) -> None:
    openpyxl, _ = ensure_deps()

    csv_path = Path(args.selection_csv)
    if not csv_path.exists():
        fatal(f"CSV introuvable: {csv_path}")

    wb = openpyxl.load_workbook(args.target)
    ws = resolve_target_sheet(wb, args.target_sheet)

    header_row, header_map = find_header_row_and_map(ws)
    existing = build_existing_index(ws, header_row, header_map)

    prefix, max_num, width = init_code_sequence(ws, header_row, header_map.get("code_echantillon"))
    seq_state = {"prefix": prefix, "num": max_num, "width": width}

    imported = []
    skipped_manual = []
    skipped_validation = []
    duplicates = []
    history_rows = []
    run_rows: List[int] = []

    with csv_path.open("r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            include = normalize(row.get("include", "1")) in ("1", "true", "TRUE", "yes", "YES")
            if not include:
                row["status"] = "non_importe_manuel"
                skipped_manual.append(row)
                history_rows.append({"ref": row.get("ref"), "status": row["status"], "reason": "Décoche utilisateur"})
                continue

            errs = []
            if not normalize(row.get("date_capture")):
                errs.append("Date capture manquante")
            if not normalize(row.get("code_espece")):
                errs.append("Code espèce manquant")
            if not normalize(row.get("num_individu")):
                errs.append("Numéro individu manquant")
            if not normalize(row.get("code_type_echantillon")):
                errs.append("Code type échantillon manquant")

            if errs:
                row["status"] = "a_reimporter"
                skipped_validation.append({"row": row, "errors": errs})
                history_rows.append({"ref": row.get("ref"), "status": row["status"], "reason": " | ".join(errs)})
                continue

            key = (normalize(row.get("num_individu")), normalize(row.get("code_type_echantillon")))
            if key in existing:
                duplicates.append({"row": row, "existing_row": existing[key]["row"]})
                action = args.on_duplicate
                if action == "ignore":
                    row["status"] = "ignore_doublon"
                    history_rows.append({"ref": row.get("ref"), "status": row["status"], "reason": "Doublon exact"})
                    continue
                if action == "alert":
                    row["status"] = "a_reimporter"
                    history_rows.append({"ref": row.get("ref"), "status": row["status"], "reason": "Doublon exact (alerte)"})
                    continue
                if action == "replace":
                    target_row = existing[key]["row"]
                    apply_target_row(ws, target_row, row, header_map, args.default_organisme, args.default_country)
                    propagate_formula_from_previous_row(openpyxl, ws, target_row, header_map.get("code_echantillon"), min_row=header_row + 1)
                    propagate_formula_from_previous_row(openpyxl, ws, target_row, header_map.get("code_compose"), min_row=header_row + 1)
                    propagate_missing_formulas_row(openpyxl, ws, target_row, min_row=header_row + 1)
                    propagate_incremental_code_from_previous_row(ws, target_row, header_map.get("code_echantillon"), min_row=header_row + 1)
                    assign_next_code_if_missing(ws, target_row, header_map.get("code_echantillon"), seq_state)
                    parts = parse_ref_parts(normalize(row.get("num_individu")))
                    expected_prefix = parts[0] if parts else ""
                    copy_context_values_from_previous_row(
                        ws,
                        target_row,
                        header_map,
                        min_row=header_row + 1,
                        source_rows=run_rows,
                        expected_type=normalize(row.get("code_type_echantillon")),
                        expected_ref_prefix=expected_prefix,
                    )
                    run_rows.append(target_row)
                    row["status"] = "importe_remplace"
                    imported.append(row)
                    history_rows.append({"ref": row.get("ref"), "status": row["status"], "reason": "Doublon remplacé"})
                    continue

            target_row = first_empty_row(ws, header_row, header_map["num_individu"])
            apply_target_row(ws, target_row, row, header_map, args.default_organisme, args.default_country)
            # Keep formula-driven columns alive for newly appended rows.
            propagate_formula_from_previous_row(openpyxl, ws, target_row, header_map.get("code_echantillon"), min_row=header_row + 1)
            propagate_formula_from_previous_row(openpyxl, ws, target_row, header_map.get("code_compose"), min_row=header_row + 1)
            propagate_missing_formulas_row(openpyxl, ws, target_row, min_row=header_row + 1)

            # Fallback: if there is no formula to propagate (value-only workbook), increment from previous code.
            propagate_incremental_code_from_previous_row(ws, target_row, header_map.get("code_echantillon"), min_row=header_row + 1)
            assign_next_code_if_missing(ws, target_row, header_map.get("code_echantillon"), seq_state)

            # Preserve context fields used across batches when current source is empty.
            parts = parse_ref_parts(normalize(row.get("num_individu")))
            expected_prefix = parts[0] if parts else ""
            copy_context_values_from_previous_row(
                ws,
                target_row,
                header_map,
                min_row=header_row + 1,
                source_rows=run_rows,
                expected_type=normalize(row.get("code_type_echantillon")),
                expected_ref_prefix=expected_prefix,
            )
            existing[key] = {"row": target_row}
            run_rows.append(target_row)
            row["status"] = "importe"
            imported.append(row)
            history_rows.append({"ref": row.get("ref"), "status": row["status"], "reason": "OK"})

    out_file = Path(args.out_target)
    wb.save(out_file)
    update_history(Path(args.history), history_rows)

    summary = {
        "imported": len(imported),
        "skipped_manual": len(skipped_manual),
        "skipped_validation": len(skipped_validation),
        "duplicates": len(duplicates),
        "target_out": str(out_file),
        "history": str(args.history),
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))


def cmd_reimport(args: argparse.Namespace) -> None:
    hist = Path(args.history)
    if not hist.exists():
        fatal(f"Historique introuvable: {hist}")

    payload = json.loads(hist.read_text(encoding="utf-8"))
    rows = payload.get("rows", [])
    pending = [r for r in rows if r.get("status") in ("a_reimporter", "non_importe_manuel")]

    if not pending:
        print("Aucune ligne à réimporter.")
        return

    selected_refs = set(args.refs)
    selected = [r for r in pending if not selected_refs or normalize(r.get("ref")) in selected_refs]

    print(json.dumps({"pending": len(pending), "selected": selected}, ensure_ascii=False, indent=2))


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="datacolisa-importer",
        description="Import Excel -> Excel pour DATACOLISA (MVP context master)",
    )
    sub = p.add_subparsers(dest="cmd", required=True)

    e = sub.add_parser("extract", help="Filtre la source par plage REF et génère un CSV de sélection")
    e.add_argument("--source", required=True, help="Fichier source .xls")
    e.add_argument("--source-sheet", default=DEFAULT_SOURCE_SHEET)
    e.add_argument("--start-ref", required=True)
    e.add_argument("--end-ref", required=True)
    e.add_argument("--out-csv", default="selection_import.csv")
    e.add_argument("--default-type-echantillon", default="EC MONTEE")
    e.set_defaults(func=cmd_extract)

    i = sub.add_parser("import", help="Importe vers la cible Feuil1 selon le CSV de sélection")
    i.add_argument("--selection-csv", required=True)
    i.add_argument("--target", required=True, help="Fichier cible .xlsx")
    i.add_argument("--target-sheet", default=DEFAULT_TARGET_SHEET)
    i.add_argument("--out-target", default="COLISA_imported.xlsx")
    i.add_argument("--history", default="import_history.json")
    i.add_argument("--default-organisme", default="INRAE")
    i.add_argument("--default-country", default="France")
    i.add_argument("--on-duplicate", choices=["alert", "ignore", "replace"], default="alert")
    i.set_defaults(func=cmd_import)

    r = sub.add_parser("reimport", help="Liste les lignes marquées non importées ou à réimporter")
    r.add_argument("--history", default="import_history.json")
    r.add_argument("--refs", nargs="*", default=[])
    r.set_defaults(func=cmd_reimport)

    return p


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
